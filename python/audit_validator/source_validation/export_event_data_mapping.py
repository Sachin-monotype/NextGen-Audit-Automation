"""Per-operation field-level data mapping for every UI-Navigation event.

Produces ONE workbook with:
  - Summary
  - Index (operation → sheet name + field count)
  - Event Coverage (all sheet events + gaps)
  - One sheet per validated operation (activateFamily, activateList, …)

Each operation sheet answers: for this event's enriched JSON field, where does
the value come from (``service>table>column``), how do you query/validate it,
how is it transformed, and did the live source check PASS.

Data source = ``reports/comparison-latest.json`` — success-state enriched
samples only (raw + enriched pair), with live UMS/CMS/AMS/Typesense checks.

Usage::

    PYTHONPATH=python python -m audit_validator.source_validation.export_event_data_mapping
"""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..event_categories import resolve_category
from .export_field_mappings import (
    _envelope_section,
    _query_for,
    _query_key,
    _safe_sheet_name,
    _sort_key,
    _source_shortcut,
    _transform_for,
)

_UI_NAV_XLSX = "docs/UI Navigation of Event (2).xlsx"
_COMPARISON_LATEST = "reports/comparison-latest.json"

# Per-operation sheet columns (mirrors Font_Sync_*_field_mappings.xlsx + Match)
_OP_HEADER = (
    "#",
    "Enriched JSON path",
    "Section",
    "Source",
    "Query",
    "Transformation",
    "Enriched value",
    "Source value",
    "Match",
    "Notes",
)

_COVERAGE_HEADER = (
    "#",
    "Event (from sheet)",
    "Matched operation",
    "Section",
    "UI Navigation",
    "Validated?",
    "Fields",
    "PASS",
    "SKIP",
    "FAIL",
    "Status / gap reason",
)

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_SECTION_FILLS = {
    "event": PatternFill("solid", fgColor="EEF2FF"),
    "source": PatternFill("solid", fgColor="ECFDF5"),
    "subject": PatternFill("solid", fgColor="FFF7ED"),
    "subject.enrichedSnapshot": PatternFill("solid", fgColor="FEF3C7"),
    "actor": PatternFill("solid", fgColor="F0F9FF"),
    "actor.enrichedSnapshot": PatternFill("solid", fgColor="E0F2FE"),
}
_MATCH_FILL = {
    "PASS": PatternFill("solid", fgColor="ECFDF5"),
    "SKIP": PatternFill("solid", fgColor="FEF3C7"),
    "FAIL": PatternFill("solid", fgColor="FEE2E2"),
    "N/A": PatternFill("solid", fgColor="F3F4F6"),
}


def _norm(s: str | None) -> str:
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


@dataclass
class SheetEvent:
    operation: str
    section: str
    ui_nav: str


def _flatten_nav(cells) -> str:
    parts = [str(c).replace("\n", " ").strip() for c in cells if c]
    return " | ".join(p for p in parts if p)


def load_sheet_events(xlsx_path: Path) -> dict[str, SheetEvent]:
    """Distinct events from every tab of the UI-Navigation workbook."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, data_only=True)
    events: dict[str, SheetEvent] = {}

    def _add(op: str, section: str, ui_nav: str) -> None:
        op = (op or "").strip()
        if not op:
            return
        existing = events.get(op)
        if existing is None:
            events[op] = SheetEvent(op, section.strip(), ui_nav.strip())
        else:
            if not existing.section and section:
                existing.section = section.strip()
            if not existing.ui_nav and ui_nav:
                existing.ui_nav = ui_nav.strip()

    if "UI Navigation by Section" in wb.sheetnames:
        ws = wb["UI Navigation by Section"]
        cur_sec = ""
        for r in list(ws.iter_rows(values_only=True))[1:]:
            if r and r[0]:
                cur_sec = str(r[0]).strip()
            op = str(r[1]).strip() if len(r) > 1 and r[1] else ""
            _add(op, cur_sec, _flatten_nav(r[2:8]))

    if "App based Events" in wb.sheetnames:
        ws = wb["App based Events"]
        cur_sec = ""
        for r in list(ws.iter_rows(values_only=True))[1:]:
            if r and r[0]:
                cur_sec = str(r[0]).strip()
            op = str(r[1]).strip() if len(r) > 1 and r[1] else ""
            _add(op, cur_sec, _flatten_nav(r[2:5]))

    if "Touch Points" in wb.sheetnames:
        ws = wb["Touch Points"]
        cur = ""
        for r in list(ws.iter_rows(values_only=True))[1:]:
            if r and r[0]:
                cur = str(r[0]).strip()
            _add(cur, "Touch Points", "")

    return events


def load_validated_rows(comparison_path: Path) -> dict[str, list[dict]]:
    """operation -> list of validated field rows from comparison-latest.json."""
    if not comparison_path.is_file():
        return {}
    data = json.loads(comparison_path.read_text(encoding="utf-8"))
    out: dict[str, list[dict]] = {}
    for op, payload in data.items():
        if not isinstance(payload, dict):
            continue
        rows = payload.get("rows") or []
        if isinstance(rows, list) and rows:
            out[str(op)] = [r for r in rows if isinstance(r, dict) and r.get("field_path")]
    return out


def _style_header(ws, header: tuple[str, ...]) -> None:
    ws.append(list(header))
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def _ordered_operations(
    validated: dict[str, list[dict]],
    events: dict[str, SheetEvent],
    norm_to_op: dict[str, str],
) -> list[str]:
    ordered: list[str] = []
    seen: set[str] = set()
    for ev in events.values():
        op = norm_to_op.get(_norm(ev.operation))
        if op and op not in seen and validated.get(op):
            ordered.append(op)
            seen.add(op)
    for op in sorted(validated):
        if op not in seen:
            ordered.append(op)
            seen.add(op)
    return ordered


def _write_operation_sheet(
    wb: Workbook,
    operation: str,
    rows: list[dict],
    used_names: set[str],
) -> str:
    sheet_name = _safe_sheet_name(operation, used_names)
    ws = wb.create_sheet(sheet_name)
    _style_header(ws, _OP_HEADER)

    sorted_rows = sorted(rows, key=lambda r: _sort_key(str(r.get("field_path") or "")))
    seen_query_keys: set[str] = set()
    for i, r in enumerate(sorted_rows, 1):
        fp = str(r.get("field_path") or "")
        section = _envelope_section(fp)
        src = _source_shortcut(fp, str(r.get("source_system") or ""), str(r.get("source_api") or ""))
        qkey = _query_key(src, fp)
        full_q = _query_for(src, fp)
        if not full_q:
            query_cell = ""
        elif qkey and qkey in seen_query_keys:
            query_cell = "\u21bb same as above"
        else:
            query_cell = full_q
            if qkey:
                seen_query_keys.add(qkey)
        match = str(r.get("match_status") or "")
        ws.append(
            [
                i,
                fp,
                section,
                src,
                query_cell,
                _transform_for(src, fp),
                str(r.get("value_in_enriched") or "")[:600],
                str(r.get("value_in_source") or "")[:600],
                match,
                str(r.get("notes") or "")[:400],
            ]
        )
        fill = _SECTION_FILLS.get(section)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        match_fill = _MATCH_FILL.get(match)
        if match_fill:
            ws.cell(row=ws.max_row, column=9).fill = match_fill

    widths = [6, 56, 24, 42, 72, 48, 40, 40, 8, 44]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_OP_HEADER))}{max(ws.max_row, 1)}"
    return sheet_name


def _write_index_sheet(
    wb: Workbook,
    ops: list[str],
    sheet_by_op: dict[str, str],
    counts: dict[str, int],
    events: dict[str, SheetEvent],
    norm_to_op: dict[str, str],
) -> None:
    # Reverse: validated op → UI sheet event
    op_to_event: dict[str, SheetEvent] = {}
    for ev in events.values():
        op = norm_to_op.get(_norm(ev.operation))
        if op and op not in op_to_event:
            op_to_event[op] = ev

    ws = wb.create_sheet("Index", 0)
    ws.append(["Report", "UI-Navigation event data mapping — one sheet per operation"])
    ws.append(
        [
            "Note",
            "Source = service>table>column. Query once per unique SQL/curl; later rows "
            "use '↻ same as above'. Match = live source check on success-state enrich.",
        ]
    )
    ws.append(["Operations", len(ops)])
    ws.append([])
    header = ("#", "Operation", "Sheet", "Category", "UI Section", "Field count")
    ws.append(list(header))
    for cell in ws[5]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for i, op in enumerate(ops, 1):
        ev = op_to_event.get(op)
        ws.append(
            [
                i,
                op,
                sheet_by_op.get(op, op[:31]),
                resolve_category(op),
                (ev.section if ev else "(not in UI sheet)"),
                counts.get(op, 0),
            ]
        )
    for col, w in zip("ABCDEF", (6, 36, 32, 28, 28, 12), strict=False):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:F{max(ws.max_row, 5)}"


def _tracked_norms() -> set[str]:
    try:
        from ..operation_registry import tracked_operations

        return {_norm(o) for o in tracked_operations()}
    except Exception:
        return set()


def _write_coverage_sheet(
    wb: Workbook,
    validated: dict[str, list[dict]],
    events: dict[str, SheetEvent],
    norm_to_op: dict[str, str],
) -> None:
    ws = wb.create_sheet("Event Coverage")
    _style_header(ws, _COVERAGE_HEADER)
    tracked = _tracked_norms()
    for i, ev in enumerate(sorted(events.values(), key=lambda e: (e.section, e.operation)), 1):
        op = norm_to_op.get(_norm(ev.operation))
        rows = validated.get(op or "") or []
        counts = Counter(str(r.get("match_status") or "") for r in rows)
        validated_flag = "Yes" if rows else "No"
        if rows:
            status = "Validated against source (executed query/API)"
        elif _norm(ev.operation) in tracked:
            status = "Tracked operation — no success-state enriched sample captured yet"
        else:
            status = "UI-flow label / step — not a distinct audit operation"
        ws.append(
            [
                i,
                ev.operation,
                op or "",
                ev.section,
                ev.ui_nav,
                validated_flag,
                len(rows),
                counts.get("PASS", 0),
                counts.get("SKIP", 0),
                counts.get("FAIL", 0),
                status,
            ]
        )
    widths = [6, 30, 26, 24, 46, 10, 8, 8, 8, 8, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_COVERAGE_HEADER))}{max(ws.max_row, 1)}"


def _write_summary_sheet(
    wb: Workbook,
    validated: dict[str, list[dict]],
    events: dict[str, SheetEvent],
    norm_to_op: dict[str, str],
    total_field_rows: int,
    n_op_sheets: int,
) -> None:
    ws = wb.create_sheet("Summary", 0)
    matched = sum(1 for ev in events.values() if validated.get(norm_to_op.get(_norm(ev.operation)) or ""))
    tracked = _tracked_norms()
    tracked_uncaptured = sum(
        1
        for ev in events.values()
        if not validated.get(norm_to_op.get(_norm(ev.operation)) or "")
        and _norm(ev.operation) in tracked
    )
    ui_labels = len(events) - matched - tracked_uncaptured
    dist: Counter = Counter()
    for rows in validated.values():
        for r in rows:
            dist[str(r.get("match_status") or "")] += 1
    lines = [
        ("Report", "UI-Navigation event → enriched-JSON field data mapping"),
        (
            "Layout",
            f"One sheet per operation ({n_op_sheets} operation sheets) + Index + Coverage",
        ),
        (
            "Note",
            "Source = service>table>column. Match reflects a live query/API check "
            "on the success-state enriched sample (raw+enriched pair in Mongo).",
        ),
        ("", ""),
        ("Distinct events in UI sheet", len(events)),
        ("Events with validated success-state mapping", matched),
        ("Tracked ops in sheet — no success sample captured yet", tracked_uncaptured),
        ("UI-flow labels / steps (not distinct audit ops)", ui_labels),
        ("Operation sheets", n_op_sheets),
        ("Total field rows (all operation sheets)", total_field_rows),
        ("", ""),
        ("Field checks — PASS", dist.get("PASS", 0)),
        ("Field checks — SKIP", dist.get("SKIP", 0)),
        ("Field checks — FAIL", dist.get("FAIL", 0)),
        ("Field checks — N/A", dist.get("N/A", 0)),
    ]
    for k, v in lines:
        ws.append([k, v])
    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 80
    for row in ws.iter_rows():
        row[0].font = Font(bold=True)
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def build_event_data_mapping_workbook(
    *,
    project_root: Path,
    out_path: Path,
) -> tuple[Path, int, int]:
    events = load_sheet_events(project_root / _UI_NAV_XLSX)
    validated = load_validated_rows(project_root / _COMPARISON_LATEST)
    norm_to_op = {_norm(op): op for op in validated}
    ordered = _ordered_operations(validated, events, norm_to_op)

    wb = Workbook()
    # Drop the default empty sheet — we rebuild from Summary / Index
    default = wb.active
    wb.remove(default)

    used_names: set[str] = {"Summary", "Index", "Event Coverage"}
    sheet_by_op: dict[str, str] = {}
    counts: dict[str, int] = {}
    total_rows = 0

    for op in ordered:
        rows = validated[op]
        sheet_by_op[op] = _write_operation_sheet(wb, op, rows, used_names)
        counts[op] = len(rows)
        total_rows += len(rows)

    _write_coverage_sheet(wb, validated, events, norm_to_op)
    _write_index_sheet(wb, ordered, sheet_by_op, counts, events, norm_to_op)
    _write_summary_sheet(wb, validated, events, norm_to_op, total_rows, len(ordered))

    # Tab order: Summary, Index, Event Coverage, then operation sheets A–Z
    desired = ["Summary", "Index", "Event Coverage"] + [
        sheet_by_op[op] for op in ordered
    ]
    for i, name in enumerate(desired):
        if name in wb.sheetnames:
            wb.move_sheet(name, offset=i - wb.sheetnames.index(name))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path, len(ordered), total_rows


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", type=Path, default=Path("."))
    parser.add_argument(
        "--out", type=Path, default=Path("docs/mappings/UI_Navigation_event_data_mapping.xlsx")
    )
    args = parser.parse_args(argv)
    root = args.root.resolve()
    out, n_ops, n_rows = build_event_data_mapping_workbook(project_root=root, out_path=args.out)
    print(f"Wrote {out} — {n_ops} operation sheet(s), {n_rows} field rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
