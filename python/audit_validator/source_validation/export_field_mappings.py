"""Export per-category Excel workbooks mirroring the Results field list.

Columns (copy-friendly for DB validation)::
  # | Enriched JSON path | Section | Source | Query | Transformation

``Source`` uses shortcuts like ``ums>vw_profile_details>email`` or
``audit-service>enricher>source`` (never ``unknown`` / ``resolver``).

Usage::

    PYTHONPATH=python python -m audit_validator.source_validation.export_field_mappings
"""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from ..event_categories import CATEGORIES, resolve_category
from ..operation_registry import tracked_operations
from .comparison_rows import ComparisonRow, build_comparison_rows
from .config import load_source_validation_config
from .runner import _load_enriched_sample

_ENVELOPE_RANK = {
    "event": 0,
    "source": 1,
    "subject": 2,
    "subject.enrichedSnapshot": 3,
    "actor": 4,
    "actor.enrichedSnapshot": 5,
}

_HEADER = (
    "#",
    "Enriched JSON path",
    "Section",
    "Source",
    "Query",
    "Transformation",
)

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
_SECTION_FILLS = {
    "event": PatternFill("solid", fgColor="EEF2FF"),
    "source": PatternFill("solid", fgColor="ECFDF5"),
    "subject": PatternFill("solid", fgColor="FFF7ED"),
    "subject.enrichedSnapshot": PatternFill("solid", fgColor="FEF3C7"),
    "actor": PatternFill("solid", fgColor="F0F9FF"),
    "actor.enrichedSnapshot": PatternFill("solid", fgColor="E0F2FE"),
}


@dataclass(frozen=True)
class MappingRow:
    field_path: str
    source_system: str
    source_api: str
    value_in_enriched: str
    match_status: str
    notes: str


def _envelope_section(path: str) -> str:
    if path.startswith("actor.enrichedSnapshot."):
        return "actor.enrichedSnapshot"
    if path.startswith("actor.") or path == "actor":
        return "actor"
    if path.startswith("subject.enrichedSnapshot."):
        return "subject.enrichedSnapshot"
    if path.startswith("subject.") or path == "subject":
        return "subject"
    if path.startswith("source.") or path == "source":
        return "source"
    return "event"


def _sort_key(path: str) -> tuple:
    return (_ENVELOPE_RANK.get(_envelope_section(path), 9), path)


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name.strip())[:31] or "op"
    base = cleaned
    n = 2
    while cleaned in used:
        suffix = f"_{n}"
        cleaned = (base[: 31 - len(suffix)] + suffix)[:31]
        n += 1
    used.add(cleaned)
    return cleaned


def _safe_filename(category: str) -> str:
    cleaned = re.sub(r"[^\w\-]+", "_", category.strip()).strip("_")
    return cleaned or "Other"


def _leaf(path: str) -> str:
    seg = path.rsplit(".", 1)[-1]
    return re.sub(r"\[\d+\]$", "", seg) or seg


def _source_shortcut(field_path: str, source_system: str, source_api: str = "") -> str:
    """Compact source: service>table>column (never unknown)."""
    sys = (source_system or "").strip()
    leaf = _leaf(field_path)
    low = field_path.lower()
    api = (source_api or "").lower()
    sys_l = sys.lower()
    sys_u = sys.upper()

    # Envelope / subject raw fields
    if low in {"xcorrelationid", "eventid", "eventversion", "enrichmentversion", "enrichedat"}:
        if low == "enrichedat":
            return "audit-service>enricher>enrichedAt"
        return f"raw>envelope>{leaf}"
    if low.startswith("source."):
        return f"raw>source>{leaf}"
    if low == "subject.type":
        return "graphql>mutation>subject.type"
    if low.startswith("subject.id"):
        return "graphql>mutation>subject.id"
    if low.startswith("actor.") and leaf.lower() in {
        "globaluserid",
        "globalcustomerid",
        "orgid",
        "parentcustomerid",
    }:
        return f"jwt>token>{leaf}"

    # Treat old labels
    if "resolver" in sys_l or sys_u in {"RESOLVER", "RESOLVER (GENERATED)"}:
        sys_u = "AUDIT-SERVICE"
        sys_l = "audit-service"
    if sys_u == "UNKNOWN" or not sys:
        # Last-chance path inference
        if "fontdetails" in low or ".family." in low or ".styles[" in low:
            return f"typesense>styles>{leaf}"
        if ".customer." in low or ".subscription." in low:
            return f"cms>customers>{leaf}"
        if ".profile." in low or ".user." in low:
            return f"ums>vw_profile_details>{leaf}"
        if ".asset." in low:
            return f"ams>assets>{leaf}"
        if low.startswith("subject.") or low.startswith("actor."):
            return f"raw>envelope>{leaf}"
        return f"audit-service>enricher>{leaf}"

    if sys_u == "CMS" or "cms" in sys_l:
        if "subscription" in low or "subscription" in api:
            if "plandefinition" in low:
                return f"cms>customer_subscription>plan_definition.{leaf}"
            col = {
                "createdat": "created_on",
                "istrial": "is_trial",
                "isactive": "is_deleted",
                "customerid": "customer_id",
                "producttype": "product_type",
                "seatsavailable": "seats_available",
                "terminationdate": "termination_date",
            }.get(leaf.lower(), leaf)
            return f"cms>customer_subscription>{col}"
        col = {
            "displayname": "display_name",
            "createdat": "created_on",
            "modifiedat": "modified_on",
            "parentid": "parent_id",
            "metadata": "meta_data",
        }.get(leaf.lower(), leaf)
        return f"cms>customers>{col}"

    if sys_u == "UMS" or "ums" in sys_l:
        if "permission" in low:
            return "ums>role_permissions_mapping>permission_id"
        if ".role." in low or "role" in api:
            col = {
                "displayname": "display_name",
                "typeid": "type_id",
            }.get(leaf.lower(), leaf)
            return f"ums>roles>{col}"
        if "deleted" in api:
            return f"ums>deleted_profiles>{leaf}"
        col = {
            "id": "profile_Id_uuid",
            "customerid": "customer_id_uuid",
            "userid": "user_id_uuid",
            "idpuserid": "idp_user_id",
            "firstname": "first_name",
            "lastname": "last_name",
            "isactive": "is_active",
            "createdat": "created_on",
            "displayname": "role_name",
        }.get(leaf.lower(), leaf)
        return f"ums>vw_profile_details>{col}"

    if sys_u == "AMS" or "ams" in sys_l or "asset" in sys_l:
        if leaf.lower() in {"name", "updatedat", "description", "parentid"}:
            col = {"updatedat": "updated_at", "parentid": "parent_id"}.get(leaf.lower(), leaf)
            return f"ams>projects>{col}"
        if "accessid" in leaf.lower() or "accessids" in low:
            return "ams>asset_user_access>access_id"
        col = {
            "id": "asset_id",
            "assettype": "asset_type",
            "createdby": "created_by",
            "globalcustomerid": "global_customer_id",
            "createdat": "created_at",
            "assetpath": "asset_path",
            "depth": "asset_level",
            "metadata": "meta_data",
        }.get(leaf.lower(), leaf)
        return f"ams>assets>{col}"

    if "typesense" in sys_l or "discovery" in sys_l:
        return f"typesense>styles_or_variations>{leaf}"

    if "jwt" in sys_l or "bearer" in sys_l:
        return f"jwt>token>{leaf}"

    if "audit" in sys_l or "resolver" in sys_l:
        return f"audit-service>enricher>{leaf}"

    if sys_u in {"RAW", "N/A"} or "raw" in sys_l:
        return f"raw>envelope>{leaf}"

    if "byof" in sys_l or "batch" in sys_l or "contract" in low:
        return f"byof-license>contract_or_batch>{leaf}"

    safe = re.sub(r"[^\w]+", "-", sys_l) or "audit-service"
    return f"{safe}>{leaf}"


def _discovery_base() -> str:
    import os

    return (
        os.getenv("DISCOVERY_BASE_URL")
        or "https://mtc-middleware-discovery.monotype-pp.com"
    ).rstrip("/")


def _query_key(source: str, field_path: str) -> str:
    """Stable key for deduping identical Query text within a sheet."""
    low = field_path.lower()
    if source.startswith("cms>customer_subscription"):
        return "cms.customer_subscription"
    if source.startswith("cms>customers"):
        return "cms.customers"
    if source.startswith("ums>role_permissions_mapping"):
        return "ums.role_permissions"
    if source.startswith("ums>roles"):
        return "ums.roles"
    if source.startswith("ums>vw_profile_details"):
        return "ums.profile"
    if source.startswith("ums>deleted_profiles"):
        return "ums.deleted_profiles"
    if source.startswith("ams>projects"):
        return "ams.projects"
    if source.startswith("ams>asset_user_access"):
        return "ams.asset_user_access"
    if source.startswith("ams>assets"):
        return "ams.assets"
    if source.startswith("typesense"):
        if "variation" in low or "md5" in low:
            return "discovery.variations"
        return "discovery.styles"
    # raw / jwt / audit-service / byof → no shared executable query
    return ""


def _query_for(source: str, field_path: str) -> str:
    """Copy-paste Query with {placeholders}. Blank when not independently queryable."""
    low = field_path.lower()

    # Envelope / JWT / enricher — no source query (process is audit pipeline, not Mongo docs)
    if source.startswith("raw>") or source.startswith("jwt>") or source.startswith("audit-service>"):
        return ""

    if source.startswith("cms>customer_subscription"):
        return (
            "SELECT customer_id, plan_definition, product_type, seats_available, "
            "termination_date, is_trial, is_deleted, created_on "
            "FROM customer_management.customer_subscription "
            "WHERE customer_id = '{gcid}' "
            "AND (is_deleted = 0 OR is_deleted IS NULL) "
            "ORDER BY id DESC LIMIT 1;"
        )
    if source.startswith("cms>customers"):
        return (
            "SELECT id, name, display_name, source, parent_id, meta_data, "
            "is_predelivery_enabled, is_test_demo, created_on, modified_on "
            "FROM customer_management.customers "
            "WHERE id = '{gcid}' LIMIT 1;"
        )
    if source.startswith("ums>role_permissions_mapping"):
        return (
            "SELECT LOWER(BIN_TO_UUID(role_id)) AS role_id, permission_id "
            "FROM user_management.role_permissions_mapping "
            "WHERE role_id = UUID_TO_BIN('{role_id}') "
            "ORDER BY permission_id;"
        )
    if source.startswith("ums>roles"):
        return (
            "SELECT LOWER(BIN_TO_UUID(id)) AS id, display_name, type_id, description "
            "FROM user_management.roles "
            "WHERE id = UUID_TO_BIN('{role_id}') LIMIT 1;"
        )
    if source.startswith("ums>vw_profile_details"):
        return (
            "SELECT profile_Id_uuid, customer_id_uuid, user_id_uuid, email, "
            "first_name, last_name, idp_user_id, is_active, role_id_uuid, "
            "role_name, created_on, meta "
            "FROM user_management.vw_profile_details "
            "WHERE profile_Id_uuid = '{profile_id}' "
            "AND (is_deleted = 0 OR is_deleted IS NULL) LIMIT 1;"
        )
    if source.startswith("ums>deleted_profiles"):
        return (
            "SELECT * FROM user_management.deleted_profiles "
            "WHERE idp_user_id = '{idp_user_id}' LIMIT 1;"
        )
    if source.startswith("ams>projects"):
        return (
            "SELECT LOWER(BIN_TO_UUID(id)) AS id, name, description, "
            "LOWER(BIN_TO_UUID(parent_id)) AS parent_id, created_at, updated_at "
            "FROM asset_management.projects "
            "WHERE id = UUID_TO_BIN('{asset_id}') LIMIT 1;"
        )
    if source.startswith("ams>asset_user_access"):
        return (
            "SELECT asset_id, asset_type, user_id, access_id, created_at "
            "FROM asset_management.asset_user_access "
            "WHERE asset_id = '{asset_id}' AND user_id = '{profile_id}';"
        )
    if source.startswith("ams>assets"):
        return (
            "SELECT asset_id, asset_type, created_by, global_customer_id, "
            "created_at, asset_path, asset_level, meta_data "
            "FROM asset_management.assets "
            "WHERE asset_id = '{asset_id}' LIMIT 1;"
        )
    if source.startswith("typesense"):
        base = _discovery_base()
        if "variation" in low or "md5" in low:
            return (
                f"curl -X GET '{base}/v1/variations?md5s={{md5}}' \\\n"
                f"  -H 'Authorization: Bearer {{discovery_token}}' \\\n"
                f"  -H 'Accept: application/json'"
            )
        return (
            f"curl -X POST '{base}/v1/styles?skipInventoryCheck=true' \\\n"
            f"  -H 'Authorization: Bearer {{discovery_token}}' \\\n"
            f"  -H 'Content-Type: application/json' \\\n"
            f"  -H 'Accept: application/json' \\\n"
            f"  -d '{{\"familyIds\":[\"{{family_id}}\"],\"page\":1,\"per_page\":5}}'"
        )
    if source.startswith("byof"):
        return ""
    return ""


def _transform_for(source: str, field_path: str) -> str:
    leaf = _leaf(field_path).lower()
    low = field_path.lower()

    if "subscription" in low and leaf == "isactive":
        return (
            "DB: is_deleted=0 AND (termination_date IS NULL OR termination_date > now) "
            "→ API isActive (bool)"
        )
    if leaf in {"istrial", "isactive", "ispredeliveryenabled", "istestdemo"}:
        return "MySQL TINYINT 0/1 → Python/JSON bool True/False"
    if leaf.endswith("at") or leaf in {"createdat", "modifiedat", "updatedat", "terminationdate"}:
        return "datetime → ISO-8601 with ms + Z (e.g. 2026-06-15T06:01:32.000Z)"
    if "plandefinition" in low:
        return "MySQL JSON column plan_definition → nested object as-is"
    if "metadata" in low or leaf == "meta":
        return "MySQL JSON → object; missing keys treated as absent"
    if "permission" in low:
        return "permission_id rows → [{id: n}, …] sorted by id"
    if "accessids" in low or leaf.startswith("accessid"):
        return (
            "asset_user_access (+ parent path + Company Admin SuperAdmin) "
            "→ int array; compare per index"
        )
    if source.startswith("typesense"):
        return (
            "Typesense hit fields → enriched catalog/visual leaves; "
            "case-insensitive / trailing ' test' stripped on names"
        )
    if source.startswith("jwt>"):
        return "Echo from JWT claims — no transform"
    if source.startswith("raw>"):
        return "Echo from audit envelope (producer) — no transform"
    if source.startswith("audit-service>"):
        return "Enricher constant or derived flag — compare as echoed string/bool"
    if leaf in {"displayname", "name", "firstname", "lastname", "name_en", "title_en"}:
        return "String trim + casefold for compare; optional strip trailing ' test'"
    return "Direct equality after normalize_compare (trim / stringify)"


def _comparison_latest_path(project_root: Path) -> Path:
    return project_root / "reports" / "comparison-latest.json"


def _load_from_comparison_latest(project_root: Path) -> dict[str, list[MappingRow]]:
    path = _comparison_latest_path(project_root)
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}
    out: dict[str, list[MappingRow]] = {}
    ops_block = data.get("operations") if isinstance(data, dict) else None
    if not isinstance(ops_block, dict):
        ops_block = data if isinstance(data, dict) else {}
    for op, payload in ops_block.items():
        if not isinstance(payload, dict):
            continue
        raw_rows = payload.get("rows") or []
        if not isinstance(raw_rows, list):
            continue
        mapped: list[MappingRow] = []
        for r in raw_rows:
            if not isinstance(r, dict):
                continue
            fp = str(r.get("field_path") or "").strip()
            if not fp:
                continue
            mapped.append(
                MappingRow(
                    field_path=fp,
                    source_system=str(r.get("source_system") or ""),
                    source_api=str(r.get("source_api") or ""),
                    value_in_enriched=str(r.get("value_in_enriched") or "")[:500],
                    match_status=str(r.get("match_status") or ""),
                    notes=str(r.get("notes") or "")[:300],
                )
            )
        if mapped:
            out[str(op)] = mapped
    return out


def _rows_from_enrich_sample(project_root: Path, operation: str) -> list[MappingRow]:
    cfg = load_source_validation_config(project_root)
    enriched = _load_enriched_sample(cfg, operation, sample_source="fresh")
    if not enriched:
        return []
    rows: list[ComparisonRow] = build_comparison_rows(operation, enriched, live={})
    return [
        MappingRow(
            field_path=r.field_path,
            source_system=r.source_system,
            source_api=r.source_api,
            value_in_enriched=r.value_in_enriched,
            match_status=r.match_status,
            notes=r.notes,
        )
        for r in rows
    ]


def collect_operation_rows(
    operation: str,
    *,
    project_root: Path,
    latest_cache: dict[str, list[MappingRow]],
) -> list[MappingRow]:
    rows = latest_cache.get(operation) or _rows_from_enrich_sample(project_root, operation)
    return sorted(rows, key=lambda r: _sort_key(r.field_path))


def _write_index(wb: Workbook, category: str, ops: list[str], counts: dict[str, int]) -> None:
    ws = wb.active
    ws.title = "Index"
    ws.append(["Category", category])
    ws.append(["Operations", len(ops)])
    ws.append(
        [
            "Note",
            "Source = service>table>column. Query is written once per unique "
            "SQL/curl; later rows use '↻ same as above'. raw/jwt/audit-service left blank.",
        ]
    )
    ws.append([])
    ws.append(["#", "Operation", "Sheet", "Field count"])
    for cell in ws[5]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for i, op in enumerate(ops, 1):
        ws.append([i, op, op[:31], counts.get(op, 0)])
    for col, w in zip("ABCD", (6, 36, 36, 12), strict=False):
        ws.column_dimensions[col].width = w


_SAME_AS_ABOVE = "↻ same as above"


def _write_operation_sheet(
    wb: Workbook,
    operation: str,
    rows: list[MappingRow],
    used_names: set[str],
) -> None:
    ws = wb.create_sheet(_safe_sheet_name(operation, used_names))
    ws.append(list(_HEADER))
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # Write each unique query once; later rows sharing the same source query → "↻ same as above"
    seen_query_keys: set[str] = set()

    for i, r in enumerate(rows, 1):
        section = _envelope_section(r.field_path)
        source = _source_shortcut(r.field_path, r.source_system, r.source_api)
        qkey = _query_key(source, r.field_path)
        full_q = _query_for(source, r.field_path)
        if not full_q:
            query_cell = ""
        elif qkey and qkey in seen_query_keys:
            query_cell = _SAME_AS_ABOVE
        else:
            query_cell = full_q
            if qkey:
                seen_query_keys.add(qkey)
        ws.append(
            [
                i,
                r.field_path,
                section,
                source,
                query_cell,
                _transform_for(source, r.field_path),
            ]
        )
        fill = _SECTION_FILLS.get(section)
        if fill:
            for cell in ws[ws.max_row]:
                cell.fill = fill
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = [6, 56, 24, 42, 72, 48]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADER))}{max(ws.max_row, 1)}"


def export_category_workbook(
    category: str,
    operations: list[str],
    out_dir: Path,
    *,
    project_root: Path,
    latest_cache: dict[str, list[MappingRow]],
) -> Path | None:
    ops = sorted(operations)
    if not ops:
        return None
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"{_safe_filename(category)}_field_mappings.xlsx"

    rows_by_op: dict[str, list[MappingRow]] = {}
    for op in ops:
        rows_by_op[op] = collect_operation_rows(
            op, project_root=project_root, latest_cache=latest_cache
        )
    ops_with_fields = [op for op in ops if rows_by_op.get(op)]
    if not ops_with_fields:
        return None

    wb = Workbook()
    _write_index(
        wb,
        category,
        ops_with_fields,
        {op: len(rows_by_op[op]) for op in ops_with_fields},
    )
    used: set[str] = {"Index"}
    for op in ops_with_fields:
        _write_operation_sheet(wb, op, rows_by_op[op], used)
    wb.save(path)
    return path


def export_all_category_workbooks(
    *,
    out_dir: Path,
    project_root: Path | None = None,
    operations: list[str] | None = None,
) -> list[Path]:
    root = project_root or Path(".")
    ops = list(operations or tracked_operations())
    latest = _load_from_comparison_latest(root)
    print(f"  comparison-latest.json ops: {len(latest)}")

    by_cat: dict[str, list[str]] = defaultdict(list)
    for op in ops:
        by_cat[resolve_category(op)].append(op)

    written: list[Path] = []
    ordered = list(CATEGORIES) + sorted(c for c in by_cat if c not in CATEGORIES)
    for cat in ordered:
        if cat not in by_cat:
            continue
        path = export_category_workbook(
            cat,
            by_cat[cat],
            out_dir,
            project_root=root,
            latest_cache=latest,
        )
        if path:
            written.append(path)
            print(f"  {cat} → {path.name}")
    return written


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--out", type=Path, default=Path("docs/mappings"))
    parser.add_argument("--root", type=Path, default=Path("."))
    parser.add_argument("--operations", type=str, default="")
    args = parser.parse_args(argv)
    ops = [o.strip() for o in args.operations.split(",") if o.strip()] or None
    print(f"Writing field mappings → {args.out}")
    paths = export_all_category_workbooks(
        out_dir=args.out, project_root=args.root.resolve(), operations=ops
    )
    print(f"Done: {len(paths)} workbook(s)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
