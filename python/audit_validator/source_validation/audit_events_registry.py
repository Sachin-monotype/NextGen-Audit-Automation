"""Parse audit-events.xlsx — master event → enricher → API mapping."""

from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

DEFAULT_AUDIT_EVENTS_XLSX = (
    Path(__file__).resolve().parent.parent / "data" / "audit-events.xlsx"
)
_FALLBACK_AUDIT_EVENTS_XLSX = (
    Path.home() / "Downloads" / "MT Connect NextGen" / "audit-events.xlsx"
)


def resolve_audit_events_xlsx(path: Path | None = None) -> Path:
    if path and path.is_file():
        return path
    if DEFAULT_AUDIT_EVENTS_XLSX.is_file():
        return DEFAULT_AUDIT_EVENTS_XLSX
    return _FALLBACK_AUDIT_EVENTS_XLSX


@dataclass(frozen=True)
class ApiCatalogEntry:
    service: str
    function_name: str
    method: str
    url: str
    headers: str
    body: str
    used_by: str
    curl_example: str


@dataclass(frozen=True)
class AuditEventSpec:
    index: int
    operation: str
    category: str
    routing_key: str
    enricher_file: str
    produces: str  # S+A | A | S | S+A (opt)
    subject_apis: str
    actor_apis: str
    enrichment_remarks: str = ""

    @property
    def enriches_actor(self) -> bool:
        return self.produces.startswith("S+A") or self.produces == "A"

    @property
    def enriches_subject(self) -> bool:
        return self.produces.startswith("S+A") or self.produces == "S"


@dataclass(frozen=True)
class ActorApiSpec:
    service: str
    endpoint: str
    purpose: str
    snapshot_field: str


def _clean(val: object) -> str:
    if val is None:
        return ""
    return str(val).strip().replace("\xa0", " ")


def _is_index(val: str) -> bool:
    return bool(re.fullmatch(r"\d+", val))


def load_audit_events(path: Path | None = None) -> list[AuditEventSpec]:
    path = resolve_audit_events_xlsx(path)
    if not path.is_file():
        return []
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    events: list[AuditEventSpec] = []
    category = ""
    for row in ws.iter_rows(min_row=15, values_only=True):
        vals = [_clean(c) for c in row] + [""] * 8
        if not vals[1]:
            continue
        if vals[0] in {"#", ""} and vals[1] == "Event":
            continue
        if not _is_index(vals[0]) and vals[1] and not vals[2]:
            category = vals[1]
            continue
        if not _is_index(vals[0]):
            continue
        events.append(
            AuditEventSpec(
                index=int(vals[0]),
                operation=vals[1],
                category=category,
                routing_key=vals[2],
                enricher_file=vals[3],
                produces=vals[4],
                subject_apis=vals[5],
                actor_apis=vals[6],
                enrichment_remarks=vals[7] if len(vals) > 7 else "",
            )
        )
    wb.close()
    return events


def load_api_catalog(path: Path | None = None) -> list[ApiCatalogEntry]:
    path = resolve_audit_events_xlsx(path)
    if not path.is_file():
        return []
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    if "Sheet2" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["Sheet2"]
    out: list[ApiCatalogEntry] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        vals = [_clean(c) for c in row] + [""] * 8
        if not str(vals[0]).isdigit():
            continue
        out.append(
            ApiCatalogEntry(
                service=vals[1],
                function_name=vals[2],
                method=vals[3],
                url=vals[4],
                headers=vals[5],
                body=vals[6],
                used_by=vals[7],
                curl_example=vals[8],
            )
        )
    wb.close()
    return out


def load_actor_api_specs(path: Path | None = None) -> list[ActorApiSpec]:
    path = resolve_audit_events_xlsx(path)
    if not path.is_file():
        return []
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    specs: list[ActorApiSpec] = []
    for row in ws.iter_rows(min_row=8, max_row=13, values_only=True):
        vals = [_clean(c) for c in row]
        if vals[0] in {"Service", ""}:
            continue
        specs.append(
            ActorApiSpec(
                service=vals[0],
                endpoint=vals[1],
                purpose=vals[2],
                snapshot_field=vals[3],
            )
        )
    wb.close()
    return specs


@lru_cache(maxsize=4)
def events_by_operation(path_str: str = "") -> dict[str, AuditEventSpec]:
    path = Path(path_str) if path_str else DEFAULT_AUDIT_EVENTS_XLSX
    return {e.operation: e for e in load_audit_events(path)}


@lru_cache(maxsize=4)
def events_by_category(path_str: str = "") -> dict[str, list[AuditEventSpec]]:
    path = Path(path_str) if path_str else DEFAULT_AUDIT_EVENTS_XLSX
    out: dict[str, list[AuditEventSpec]] = {}
    for e in load_audit_events(path):
        out.setdefault(e.category, []).append(e)
    return out


def operations_with_fresh_captures(enriched_dir: Path) -> list[str]:
    """Operations with a pruned E2E capture: `{operation}.json` under payload/enrich/."""
    if not enriched_dir.is_dir():
        return []
    ops: list[str] = []
    for path in sorted(enriched_dir.glob("*.json")):
        stem = path.stem
        if stem.startswith("unknown-"):
            continue
        if stem.endswith("-mtconnect-api"):
            op = stem.replace("-mtconnect-api", "")
        else:
            op = stem
        if op:
            ops.append(op)
    return sorted(set(ops))


def operations_with_samples(queue_pairs_dir: Path) -> list[str]:
    if not queue_pairs_dir.is_dir():
        return []
    ops: list[str] = []
    for p in sorted(queue_pairs_dir.glob("*EnrichedJson.json")):
        name = p.name.replace("EnrichedJson.json", "")
        if name:
            ops.append(name)
    return ops


def category_for_operation(operation: str, *, xlsx: Path | None = None) -> str:
    spec = events_by_operation(str(xlsx or DEFAULT_AUDIT_EVENTS_XLSX)).get(operation)
    return spec.category if spec else "Unknown"


def summary_stats(events: list[AuditEventSpec]) -> dict[str, Any]:
    cats: dict[str, int] = {}
    for e in events:
        cats[e.category] = cats.get(e.category, 0) + 1
    return {
        "total_events": len(events),
        "categories": len(cats),
        "by_category": cats,
    }
