"""Field mapping registry per operation (QA Excel + resolver enricher templates)."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from .audit_events_registry import (
    DEFAULT_AUDIT_EVENTS_XLSX,
    AuditEventSpec,
    events_by_operation,
    operations_with_samples,
)
from .field_specs import ALL_SAMPLE_OPERATIONS

# QA reference workbook (ActivateFamily + createTeam sheets)
DEFAULT_REFERENCE_XLSX = Path.home() / "Downloads" / "Notification and audit2 (1).xlsx"

# Resolver enricher repo (for documentation / future codegen)
DEFAULT_RESOLVER_ROOT = (
    Path.home()
    / "Documents"
    / "CodeBases"
    / "MT Connect NextGen"
    / "mt-audit-log-resolver-service"
)


@dataclass(frozen=True)
class MappingField:
    field: str
    node: str
    sub_node: str
    attribute: str
    data_mapping: str
    notes: str
    validate: str  # Y | N | empty
    enriched_path: str
    source_system: str
    source_api: str
    layer: str  # event | actor | subject
    discovery_key: str = ""  # Typesense document field name


def _clean(val: object) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in {"\xa0", "None"} else s


def _parse_source(data_mapping: str) -> tuple[str, str]:
    """Return (Source label, API hint) — labels: Typesense, UMS, Raw, CMS, AMS, Resolver, JWT."""
    dm = data_mapping or ""
    lower = dm.lower()
    if "variation" in lower and ("discovery" in lower or "typesense" in lower):
        return "Typesense", "GET /v1/variations"
    if "discovery" in lower or "typesense" in lower:
        return "Typesense", "POST /v1/styles"
    if "ums" in lower or "user-management" in lower:
        return "UMS", "POST/GET /api/v3/customers/{gcid}/profiles"
    if "cms" in lower or "customer-management" in lower:
        return "CMS", "GET /api/v2/customers/{gcid}"
    if "ams" in lower or "asset-management" in lower:
        return "AMS", "GET /api/v3/assets/{id}"
    if "jwt" in lower or "bearer" in lower:
        return "JWT", "Bearer token claims"
    if "auth0" in lower:
        return "Auth0", "GET /api/v2/users/{id}"
    if "resolver" in lower:
        return "Resolver", "enricher constant"
    if "raw" in lower or "mtconnect-api" in lower or "graphql" in lower or "trigger" in lower or "curl" in lower:
        return "Trigger", "GraphQL curl / event trigger"
    return "Unknown", dm[:80] if dm else ""


def _normalize_indexed_path(segment: str) -> str:
    s = segment.replace("fontDetails[]", "fontDetails[0]")
    s = s.replace("styles[]", "styles[0]")
    s = s.replace("variations[]", "variations[0]")
    s = re.sub(r"\[\]", "[0]", s)
    return s


def _discovery_key_from_sub_node(sub_node: str) -> str:
    sn = sub_node.replace("[]", "").strip()
    if sn == "visual_properties.*":
        return "visual_properties"
    if "md5" in sn and "render" in sn:
        return "render_md5"
    if "/" in sn:
        return sn.split("/")[0].strip()
    return sn


def _path_from_hierarchy(
    *,
    field: str,
    node: str,
    sub_node: str,
    attribute: str,
    section: str,
) -> str:
    if field in {"xCorrelationId", "eventId", "eventVersion", "occurredAt", "enrichedEventId",
                 "enrichmentVersion", "enrichedAt"}:
        return field
    if field and "." in field and "(top-level)" not in field:
        return _normalize_indexed_path(field)
    if field.startswith("subject.") or field.startswith("actor."):
        return _normalize_indexed_path(field)

    # fontDetails[].styles[].catalog + font_ps_names[]
    if node.startswith("fontDetails") or "fontDetails" in node:
        prefix = "subject.enrichedSnapshot."
        n = _normalize_indexed_path(node)
        path = prefix + n
        if sub_node:
            sn = sub_node.replace("[]", "[0]").strip()
            if sn in {"foundry.name_en", "foundry.title_en"}:
                path = f"{path}.family.foundry.{sn.split('.', 1)[1]}"
            elif sn in {"variations[0].md5", "variations[].md5"}:
                path = f"{path}.styles[0].variations[0].catalog.md5"
            else:
                dk = _discovery_key_from_sub_node(sub_node)
                if dk:
                    path = f"{path}.{dk}"
        return path

    # actor.enrichedSnapshot: user / profile.id
    if section == "actor.enrichedSnapshot" and node:
        base = "actor.enrichedSnapshot"
        if sub_node and "." in sub_node:
            return f"{base}.{node}.{sub_node}"
        if sub_node:
            return f"{base}.{node}.{sub_node}"
        return f"{base}.{node}" if node else base

    if section == "subject.enrichedSnapshot" and node and not node.startswith("fontDetails"):
        base = "subject.enrichedSnapshot"
        if sub_node:
            return f"{base}.{node}.{sub_node}"
        return f"{base}.{node}"

    if field == "source" and not node:
        return "source"
    if node == "source" and section:
        return f"{section}.source"

    parts: list[str] = []
    if section:
        parts.append(section)
    if node:
        parts.append(_normalize_indexed_path(node))
    if sub_node:
        parts.append(_discovery_key_from_sub_node(sub_node))
    if not parts and field:
        return _normalize_indexed_path(field)
    return ".".join(parts)


def _layer_for_path(path: str) -> str:
    if path.startswith("actor."):
        return "actor"
    if path.startswith("subject."):
        return "subject"
    return "event"


def _rows_from_excel_sheet(ws) -> list[MappingField]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return []
    out: list[MappingField] = []
    section = ""
    for raw in rows[2:]:
        cols = list(raw) + [None] * 8
        field = _clean(cols[0]) or _clean(cols[1])
        node = _clean(cols[2])
        sub_node = _clean(cols[3])
        validate = _clean(cols[4])
        data_mapping = _clean(cols[6] if len(cols) > 6 else cols[5])
        notes = _clean(cols[7] if len(cols) > 7 else "")

        if not any([field, node, sub_node, data_mapping]):
            continue
        if field in {"ACTOR (top-level)", "SUBJECT (top-level)"}:
            section = ""
            continue
        if field in {"actor.enrichedSnapshot", "None"} and node == "user":
            section = "actor.enrichedSnapshot"
        if field == "actor.enrichedSnapshot":
            section = "actor.enrichedSnapshot"
            continue
        if field == "subject.enrichedSnapshot":
            section = "subject.enrichedSnapshot"
            continue
        if field and field.startswith("subject.enrichedSnapshot"):
            section = "subject.enrichedSnapshot"

        # Skip composite / non-scalar mapping rows
        if sub_node and ("/" in sub_node and "is_default" in sub_node):
            continue
        if sub_node and sub_node.startswith("other "):
            continue
        if sub_node and "activationState" in sub_node and "variations" in node:
            continue

        enriched_path = _path_from_hierarchy(
            field=field, node=node, sub_node=sub_node, attribute="", section=section
        )
        if field and not node and not sub_node:
            if field.startswith("subject.") or field.startswith("actor."):
                enriched_path = _normalize_indexed_path(field)
            elif field in {"globalUserId", "globalCustomerId", "orgId", "parentCustomerId"}:
                enriched_path = f"actor.{field}"
            elif field == "source" and not section:
                enriched_path = "source"

        if not enriched_path or enriched_path in {"subject.enrichedSnapshot", "actor.enrichedSnapshot"}:
            continue

        dk = _discovery_key_from_sub_node(sub_node) if sub_node else ""
        src_sys, src_api = _parse_source(data_mapping)
        display_field = field or sub_node or node
        out.append(
            MappingField(
                field=display_field,
                node=node,
                sub_node=sub_node,
                attribute="",
                data_mapping=data_mapping,
                notes=notes,
                validate=validate,
                enriched_path=enriched_path,
                source_system=src_sys,
                source_api=src_api,
                layer=_layer_for_path(enriched_path),
                discovery_key=dk,
            )
        )
    return out


def _font_envelope_fields(operation: str) -> list[MappingField]:
    """Shared font-activation mapping (activate-family.enricher pattern)."""
    op = operation
    base: list[tuple[str, str, str, str, str, str]] = [
        ("xCorrelationId", "", "", "", "Source: GraphQL curl / event trigger → xCorrelationId", "N"),
        ("eventId", "", "", "", "Source: GraphQL curl / event trigger → eventId", "N"),
        ("source", "operation", "", "", f"Source: GraphQL curl / event trigger → {op}", "N"),
        ("actor.enrichedSnapshot", "user", "profile.id", "", "Source: UMS POST profiles", "Y"),
        ("actor.enrichedSnapshot", "user", "profile.email", "", "Source: UMS POST profiles", "Y"),
        ("actor.enrichedSnapshot", "user", "role.displayName", "", "Source: UMS GET roles", "Y"),
        ("actor.enrichedSnapshot", "customer", "id", "", "Source: CMS GET customer", "Y"),
        ("actor.enrichedSnapshot", "customer", "name", "", "Source: CMS GET customer", "Y"),
        ("subject.enrichedSnapshot", "source", "", "", "Source: Resolver → mt-connect-middleware-discovery", "N"),
        ("subject.enrichedSnapshot", "fontDetails[0]", "family.id", "", "Discovery POST /v1/styles → mtc_families_data.id", "Y"),
        ("subject.enrichedSnapshot", "fontDetails[0]", "family.catalog.name_en", "", "Discovery/Typesense name_en", "Y"),
        ("subject.enrichedSnapshot", "fontDetails[0]", "family.catalog.title_en", "", "Discovery/Typesense title_en", "Y"),
        ("subject.enrichedSnapshot", "fontDetails[0]", "family.foundry.name_en", "", "Discovery mtc_foundries_data.name_en", "Y"),
        ("subject.enrichedSnapshot", "fontDetails[0]", "styles[0].id", "", "Discovery style document id", "Y"),
        ("subject.enrichedSnapshot", "fontDetails[0]", "styles[0].variations[0].catalog.md5", "", "Discovery GET /v1/variations md5", "Y"),
        ("subject.metadata", "input", "familyIds[0]", "", "Source: GraphQL mutation input familyIds", "Y"),
        ("subject.metadata", "input", "listIds[0]", "", "Source: GraphQL mutation input listIds (touchpoint)", "Y"),
        ("subject.metadata", "input", "listType", "", "Source: GraphQL mutation input listType (FONTLIST/PROJECT)", "Y"),
        ("subject.metadata", "input", "projectIds[0]", "", "Source: GraphQL mutation input projectIds (touchpoint)", "Y"),
        ("subject.metadata", "input", "activationType", "", "Source: GraphQL mutation input activationType", "Y"),
        ("subject.metadata", "input", "activationMode", "", "Source: GraphQL mutation input activationMode", "Y"),
        ("subject.metadata", "result", "families.totalCount", "", "Source: GraphQL mutation response families.totalCount", "Y"),
    ]
    if operation in {"bulkActivateStyles", "bulkDeactivateStyles"}:
        base.extend([
            ("subject", "batchId", "", "", "Source: GraphQL result batchId", "Y"),
            ("subject", "styleIds[0]", "", "", "Source: GraphQL input styleIds", "Y"),
        ])
    fields: list[MappingField] = []
    for field, node, sub, attr, dm, val in base:
        if node and sub:
            path = f"{field}.{node}.{sub}" if field else f"{node}.{sub}"
        elif node:
            path = f"{field}.{node}" if field else node
        else:
            path = field
        path = path.replace("[]", "[0]")
        src, api = _parse_source(dm)
        fields.append(
            MappingField(field, node, sub, attr, dm, "", val, path, src, api, _layer_for_path(path))
        )
    return fields


def _role_fields(operation: str) -> list[MappingField]:
    return [
        MappingField(
            "subject.enrichedSnapshot", "role", "source", "",
            "Source: Resolver → user-management-service", "", "N",
            "subject.enrichedSnapshot.role.source", "Resolver", "constant", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "role", "id", "",
            "Source: UMS GET roles by id", "", "Y",
            "subject.enrichedSnapshot.role.id", "UMS", "GET roles", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "role", "displayName", "",
            "Source: UMS GET roles", "", "Y",
            "subject.enrichedSnapshot.role.displayName", "UMS", "GET roles", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "role", "permissions[0].id", "",
            "Source: UMS role permissions", "", "Y",
            "subject.enrichedSnapshot.role.permissions[0].id", "UMS", "GET roles", "subject",
        ),
        MappingField(
            "actor.enrichedSnapshot", "user", "profile.id", "",
            "Source: UMS profile", "", "Y",
            "actor.enrichedSnapshot.user.profile.id", "UMS", "POST profiles", "actor",
        ),
        MappingField(
            "actor.enrichedSnapshot", "customer", "id", "",
            "Source: CMS customer", "", "Y",
            "actor.enrichedSnapshot.customer.id", "CMS", "GET customer", "actor",
        ),
    ]


def _team_fields() -> list[MappingField]:
    return [
        MappingField(
            "subject.enrichedSnapshot", "team", "source", "",
            "Source: Resolver → user-management-service", "", "N",
            "subject.enrichedSnapshot.team.source", "Resolver", "constant", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "team", "id", "",
            "Source: UMS GET teams", "", "Y",
            "subject.enrichedSnapshot.team.id", "UMS", "GET teams", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "team", "name", "",
            "Source: UMS GET teams", "", "Y",
            "subject.enrichedSnapshot.team.name", "UMS", "GET teams", "subject",
        ),
        MappingField(
            "actor.enrichedSnapshot", "user", "profile.email", "",
            "Source: UMS profile", "", "Y",
            "actor.enrichedSnapshot.user.profile.email", "UMS", "POST profiles", "actor",
        ),
    ]


def _asset_fields(operation: str) -> list[MappingField]:
    return [
        MappingField(
            "subject.enrichedSnapshot", "asset", "source", "",
            "Source: Resolver constant", "", "N",
            "subject.enrichedSnapshot.asset.source", "Resolver", "constant", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "asset", "id", "",
            "Source: AMS GET asset by id", "", "Y",
            "subject.enrichedSnapshot.asset.id", "AMS", "GET asset", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "asset", "name", "",
            "Source: AMS GET asset", "", "Y",
            "subject.enrichedSnapshot.asset.name", "AMS", "GET asset", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "asset", "assetType", "",
            "Source: AMS GET asset", "", "Y",
            "subject.enrichedSnapshot.asset.assetType", "AMS", "GET asset", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "sharingInfo[0]", "accessLevel", "",
            "Resolver-derived from AMS sharing/bulk + ACCESS_ID_MAP (not probed)", "", "Y",
            "subject.enrichedSnapshot.sharingInfo[0].accessLevel", "Resolver",
            "derived (AMS sharing/bulk + ACCESS_ID_MAP)", "subject",
        ),
    ]


def _tag_fields() -> list[MappingField]:
    return [
        MappingField(
            "subject.enrichedSnapshot", "tags[0]", "id", "",
            "Source: Discovery GET /v1/privateTag/{id}", "", "Y",
            "subject.enrichedSnapshot.tags[0].id", "UMS/Search", "GET /v1/privateTag/{id}", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "tags[0]", "name", "",
            "Source: Discovery GET /v1/privateTag/{id}", "", "Y",
            "subject.enrichedSnapshot.tags[0].name", "UMS/Search", "GET /v1/privateTag/{id}", "subject",
        ),
    ]


def _profile_fields() -> list[MappingField]:
    return [
        MappingField(
            "subject.enrichedSnapshot", "user", "profile.id", "",
            "Source: UMS POST profiles", "", "Y",
            "subject.enrichedSnapshot.user.profile.id", "UMS", "POST profiles", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "user", "profile.email", "",
            "Source: UMS POST profiles", "", "Y",
            "subject.enrichedSnapshot.user.profile.email", "UMS", "POST profiles", "subject",
        ),
        MappingField(
            "subject.enrichedSnapshot", "user", "role.displayName", "",
            "Source: UMS GET role", "", "Y",
            "subject.enrichedSnapshot.user.role.displayName", "UMS", "GET role", "subject",
        ),
    ]


_FONT_CATEGORIES = frozenset({
    "FontActivation", "FontList", "Favorite", "FavoritePair", "FontDownload",
    "FontAccess", "PrivateTag", "ProductionFont", "BYOF",
})
_ASSET_CATEGORIES = frozenset({
    "Asset", "ProjectManagement", "WebProject", "PinnedAsset + AddOn",
    "FontImport", "DocumentScanning",
})
_ROLE_CATEGORIES = frozenset({"UserRole", "Team"})
_PROFILE_CATEGORIES = frozenset({"UserProfile", "Invitation"})
_CUSTOMER_CATEGORIES = frozenset({"Customer"})


def _actor_fields() -> list[MappingField]:
    return [
        MappingField(
            "actor", "globalUserId", "", "",
            "Source: UMS profile id via JWT email / idpUserId", "", "Y",
            "actor.globalUserId", "Bearer token", "JWT + UMS", "actor",
        ),
        MappingField(
            "actor", "globalCustomerId", "", "",
            "Source: JWT claim https://api.monotype.com/gcid", "", "Y",
            "actor.globalCustomerId", "Bearer token", "JWT claim", "actor",
        ),
        MappingField(
            "actor", "orgId", "", "",
            "Source: JWT claim org_id", "", "Y",
            "actor.orgId", "Bearer token", "JWT claim", "actor",
        ),
        MappingField(
            "actor", "parentCustomerId", "", "",
            "Source: JWT claim parentCustomerId", "", "Y",
            "actor.parentCustomerId", "Bearer token", "JWT claim", "actor",
        ),
        MappingField(
            "actor", "inventories", "", "",
            "Source: JWT claim https://api.monotype.com/inventories", "", "Y",
            "actor.inventories", "Bearer token", "JWT claim", "actor",
        ),
        MappingField(
            "actor.enrichedSnapshot", "user", "source", "",
            "Source: Resolver → user-management-service", "", "N",
            "actor.enrichedSnapshot.user.source", "Resolver", "constant", "actor",
        ),
        MappingField(
            "actor.enrichedSnapshot", "user", "profile.id", "",
            "Source: UMS POST profiles", "", "Y",
            "actor.enrichedSnapshot.user.profile.id", "UMS", "POST profiles", "actor",
        ),
        MappingField(
            "actor.enrichedSnapshot", "user", "profile.email", "",
            "Source: UMS POST profiles (email from JWT)", "", "Y",
            "actor.enrichedSnapshot.user.profile.email", "UMS", "POST profiles", "actor",
        ),
        MappingField(
            "actor.enrichedSnapshot", "user", "profile.firstName", "",
            "Source: UMS POST profiles", "", "Y",
            "actor.enrichedSnapshot.user.profile.firstName", "UMS", "POST profiles", "actor",
        ),
        MappingField(
            "actor.enrichedSnapshot", "user", "role.displayName", "",
            "Source: UMS GET roles", "", "Y",
            "actor.enrichedSnapshot.user.role.displayName", "UMS", "GET roles", "actor",
        ),
        MappingField(
            "actor.enrichedSnapshot", "customer", "id", "",
            "Source: CMS GET customer", "", "Y",
            "actor.enrichedSnapshot.customer.id", "CMS", "GET customer", "actor",
        ),
        MappingField(
            "actor.enrichedSnapshot", "customer", "name", "",
            "Source: CMS GET customer", "", "Y",
            "actor.enrichedSnapshot.customer.name", "CMS", "GET customer", "actor",
        ),
        MappingField(
            "actor.enrichedSnapshot", "customer", "displayName", "",
            "Source: CMS GET customer", "", "Y",
            "actor.enrichedSnapshot.customer.displayName", "CMS", "GET customer", "actor",
        ),
    ]


def _event_header_fields(operation: str) -> list[MappingField]:
    """Envelope fields — source is the GraphQL/curl trigger we fired (not Raw Mongo).

    Validation column mirrors the QA ActivateFamily sheet (N = informational /
    compare when trigger capture exists; never fail solely because Raw is absent).
    """
    trigger = "Trigger"
    api = "GraphQL curl / event trigger"
    return [
        MappingField(
            "xCorrelationId", "", "", "",
            "Source: x-correlation-id header we sent on the trigger", "", "Y",
            "xCorrelationId", trigger, api, "event",
        ),
        MappingField(
            "eventId", "", "", "",
            "Source: eventId from trigger / service", "", "N",
            "eventId", trigger, api, "event",
        ),
        MappingField(
            "eventVersion", "", "", "",
            "Source: eventVersion from trigger envelope", "", "N",
            "eventVersion", trigger, api, "event",
        ),
        MappingField(
            "occurredAt", "", "", "",
            "Source: occurredAt from trigger envelope", "", "N",
            "occurredAt", trigger, api, "event",
        ),
        MappingField(
            "source", "operation", "", "",
            f"Source: GraphQL mutation → {operation}", "", "Y",
            "source.operation", trigger, api, "event",
        ),
        MappingField(
            "source", "service", "", "",
            "Source: trigger service (mtconnect-api)", "", "Y",
            "source.service", trigger, api, "event",
        ),
        MappingField(
            "source", "operationState", "", "",
            "Source: mutation success → operationState", "", "Y",
            "source.operationState", trigger, api, "event",
        ),
        MappingField(
            "source", "operationIndex", "", "",
            "Source: trigger batch index (usually 0)", "", "N",
            "source.operationIndex", trigger, api, "event",
        ),
        MappingField(
            "source", "platform", "", "",
            "Source: trigger platform (nextGen)", "", "Y",
            "source.platform", trigger, api, "event",
        ),
        MappingField(
            "source", "platformEnvironment", "", "",
            "Source: trigger platformEnvironment (web)", "", "Y",
            "source.platformEnvironment", trigger, api, "event",
        ),
        MappingField(
            "source", "platformVersion", "", "",
            "Source: trigger platformVersion", "", "N",
            "source.platformVersion", trigger, api, "event",
        ),
        MappingField(
            "source", "actorUserAgent", "", "",
            "Source: User-Agent header on the GraphQL curl", "", "Y",
            "source.actorUserAgent", trigger, api, "event",
        ),
        MappingField(
            "source", "type", "", "",
            "Source: trigger source.type[]", "", "N",
            "source.type", trigger, api, "event",
        ),
    ]


def _subject_structural_field(key: str, source_label: str) -> MappingField:
    path = f"subject.enrichedSnapshot.{key}"
    return MappingField(
        key, "enrichedSnapshot", key, "",
        f"Source: Resolver → {source_label}", "", "Y",
        path, "Resolver", "structural", "subject",
    )


def _mapping_for_event_spec(spec: AuditEventSpec) -> list[MappingField]:
    op = spec.operation
    fields = _event_header_fields(op)
    subj_apis = (spec.subject_apis or "").upper()
    cat = spec.category

    if spec.enriches_subject:
        if cat in _FONT_CATEGORIES or "D:" in subj_apis or "DISCOVERY" in subj_apis:
            fields.extend(_font_envelope_fields(op))
        elif cat in _ASSET_CATEGORIES or "AMS" in subj_apis:
            fields.extend(_asset_fields(op))
            if "D:" in subj_apis:
                fields.extend(
                    f for f in _font_envelope_fields(op)
                    if f.layer == "subject" and "fontDetails" in f.enriched_path
                )
        elif cat in _ROLE_CATEGORIES:
            fields.extend(_role_fields(op))
        elif cat in _PROFILE_CATEGORIES:
            fields.extend(_profile_fields())
        elif cat in _CUSTOMER_CATEGORIES:
            fields.append(_subject_structural_field("customer", "customer-management-service"))
        elif "TAG" in cat.upper() or op.startswith("createPrivate") or "PrivateTag" in cat:
            fields.extend(_tag_fields())
        else:
            fields.append(_subject_structural_field("source", spec.enricher_file))

    if spec.enriches_actor:
        fields.extend(_actor_fields())

    # Deduplicate by enriched_path
    seen: set[str] = set()
    out: list[MappingField] = []
    for f in fields:
        if f.enriched_path in seen:
            continue
        seen.add(f.enriched_path)
        out.append(f)
    return out


_FONT_OPS = frozenset({
    "activateFamily", "activateStyle", "deactivateStyle", "activateVariation",
    "bulkActivateStyles", "bulkDeactivateStyles", "addFavoriteStyles", "addFavoriteFamilies",
})
_EXPORT_OPS = frozenset({
    "exportFontAssets",
    "exportFontProjects",
    "exportFontUsers",
    "exportFontWebkits",
    "exportImportedFonts",
    "exportMyLibrary",
    "exportNotifications",
    "exportRoles",
    "exportTags",
    "exportCompanyLibrary",
    "exportTeams",
    "exportUsers",
    "exportReportingFonts",
    "exportReportingUsers",
    "exportActiveFonts",
})
_ROLE_OPS = frozenset({"createRole", "updateRole"})
_DELETE_ROLE_OPS = frozenset({"deleteRoles"})
_DELETE_TEAM_OPS = frozenset({"deleteTeams"})
_ASSET_OPS = frozenset({
    "createProject", "publishProject", "createAsset", "updateAsset", "createWebProject",
})
_EXCEL_SHEET_BY_OP = {
    "activateFamily": "ActivateFamily",
    "createTeam": "createTeam",
}


def _export_batch_fields(operation: str) -> list[MappingField]:
    """Async batch export mutations — source is GraphQL input + response batchId."""
    fields = _event_header_fields(operation)
    fields.extend(
        [
            MappingField(
                "subject", "batchId", "", "",
                "Source: GraphQL mutation response batchId", "", "Y",
                "subject.batchId", "GraphQL", "mutation response batchId", "subject",
            ),
            MappingField(
                "subject", "type", "", "",
                "Source: GraphQL mutation response subject.type", "", "Y",
                "subject.type", "GraphQL", "mutation response", "subject",
            ),
            MappingField(
                "subject", "metadata", "input", "format",
                "Source: GraphQL mutation input format (CSV)", "", "Y",
                "subject.metadata.input.format", "GraphQL", "mutation input", "subject",
            ),
            MappingField(
                "subject", "metadata", "result", "status", "",
                "Source: GraphQL mutation response status", "", "Y",
                "subject.metadata.result.status", "GraphQL", "mutation response", "subject",
            ),
            MappingField(
                "subject", "metadata", "result", "message", "",
                "Source: GraphQL mutation response message", "", "N",
                "subject.metadata.result.message", "GraphQL", "mutation response", "subject",
            ),
        ]
    )
    fields.extend(_actor_fields())
    return fields


def _delete_teams_fields() -> list[MappingField]:
    fields = _event_header_fields("deleteTeams")
    fields.extend(
        [
            MappingField(
                "subject", "id", "", "",
                "Source: GraphQL mutation input ids", "", "Y",
                "subject.id", "GraphQL", "mutation input ids", "subject",
            ),
            MappingField(
                "subject", "type", "", "",
                "Source: GraphQL mutation subject.type", "", "Y",
                "subject.type", "GraphQL", "mutation response", "subject",
            ),
            MappingField(
                "subject.enrichedSnapshot", "teams", "id", "",
                "Source: GraphQL mutation input ids (teamIds)", "", "Y",
                "subject.enrichedSnapshot.teams[0].id", "GraphQL", "mutation input ids", "subject",
            ),
        ]
    )
    fields.extend(_actor_fields())
    return fields


def _delete_roles_fields() -> list[MappingField]:
    fields = _event_header_fields("deleteRoles")
    fields.extend(
        [
            MappingField(
                "subject", "id", "", "",
                "Source: GraphQL mutation input ids", "", "Y",
                "subject.id", "GraphQL", "mutation input ids", "subject",
            ),
            MappingField(
                "subject", "type", "", "",
                "Source: GraphQL mutation subject.type", "", "Y",
                "subject.type", "GraphQL", "mutation response", "subject",
            ),
            MappingField(
                "subject.enrichedSnapshot", "role", "id", "",
                "Source: GraphQL mutation input ids (deleted role)", "", "Y",
                "subject.enrichedSnapshot.role.id", "GraphQL", "mutation input ids", "subject",
            ),
            MappingField(
                "subject.enrichedSnapshot", "roles", "id", "",
                "Source: GraphQL mutation input ids (deleted role)", "", "Y",
                "subject.enrichedSnapshot.roles[0].id", "GraphQL", "mutation input ids", "subject",
            ),
        ]
    )
    fields.extend(_actor_fields())
    return fields


def _builtin_mapping(operation: str) -> list[MappingField]:
    if operation in _EXPORT_OPS:
        return _export_batch_fields(operation)
    if operation in _FONT_OPS:
        return _font_envelope_fields(operation)
    if operation == "activateList":
        return _font_envelope_fields(operation) + _asset_fields(operation)
    if operation in _DELETE_TEAM_OPS:
        return _delete_teams_fields()
    if operation in _DELETE_ROLE_OPS:
        return _delete_roles_fields()
    if operation in _ROLE_OPS:
        return _role_fields(operation)
    if operation == "createTeam":
        return _team_fields()
    if operation in _ASSET_OPS:
        return _asset_fields(operation)
    if operation == "createPrivateTags":
        return _tag_fields()
    if operation == "updateProfile":
        return _profile_fields()
    return _font_envelope_fields(operation)


def load_reference_excel(path: Path | None = None) -> dict[str, list[MappingField]]:
    path = path or DEFAULT_REFERENCE_XLSX
    if not path.is_file():
        return {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    out: dict[str, list[MappingField]] = {}
    for op, sheet in _EXCEL_SHEET_BY_OP.items():
        if sheet in wb.sheetnames:
            out[op] = _rows_from_excel_sheet(wb[sheet])
    wb.close()
    return out


_FONT_TEMPLATE_OP = "activateFamily"
_excel_cache: dict[str, list[MappingField]] | None = None


def _font_template(reference_xlsx: Path | None = None) -> list[MappingField]:
    global _excel_cache
    if _excel_cache is None:
        _excel_cache = load_reference_excel(reference_xlsx)
    rows = _excel_cache.get(_FONT_TEMPLATE_OP) or _font_envelope_fields(_FONT_TEMPLATE_OP)
    return rows


def get_operation_mapping(
    operation: str,
    *,
    reference_xlsx: Path | None = None,
    audit_events_xlsx: Path | None = None,
) -> list[MappingField]:
    xlsx = audit_events_xlsx or DEFAULT_AUDIT_EVENTS_XLSX
    registry = events_by_operation(str(xlsx))
    spec = registry.get(operation)
    if spec:
        return _mapping_for_event_spec(spec)

    excel_maps = load_reference_excel(reference_xlsx)
    if operation in excel_maps and excel_maps[operation]:
        return excel_maps[operation]
    if operation in _FONT_OPS or operation == "activateList":
        return _font_template(reference_xlsx)
    return _builtin_mapping(operation)


def all_operation_mappings(
    operations: Iterable[str] | None = None,
    *,
    reference_xlsx: Path | None = None,
    audit_events_xlsx: Path | None = None,
) -> dict[str, list[MappingField]]:
    ops = list(operations or ALL_SAMPLE_OPERATIONS)
    return {
        op: get_operation_mapping(
            op, reference_xlsx=reference_xlsx, audit_events_xlsx=audit_events_xlsx
        )
        for op in ops
    }


def categories_for_operations(operations: Iterable[str], *, xlsx: Path | None = None) -> dict[str, list[str]]:
    registry = events_by_operation(str(xlsx or DEFAULT_AUDIT_EVENTS_XLSX))
    out: dict[str, list[str]] = {}
    for op in operations:
        spec = registry.get(op)
        cat = spec.category if spec else "Unknown"
        out.setdefault(cat, []).append(op)
    return out


def resolver_enricher_path(operation: str, resolver_root: Path | None = None) -> Path | None:
    """Best-effort path to resolver enricher TS file for documentation."""
    root = resolver_root or DEFAULT_RESOLVER_ROOT
    if not root.is_dir():
        return None
    kebab = re.sub(r"([a-z0-9])([A-Z])", r"\1-\2", operation).lower()
    matches = list(root.glob(f"src/enrichment/enrichers/**/{kebab}.enricher.ts"))
    return matches[0] if matches else None
