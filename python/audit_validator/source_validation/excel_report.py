"""Excel workbook: merged category tabs (≤15) + Scheduler & Cron."""

from __future__ import annotations

import re
from pathlib import Path

from ..cron.payloads import load_cron_cases
from ..ingress.payloads import load_ingress_cases
from .audit_events_registry import DEFAULT_AUDIT_EVENTS_XLSX, events_by_operation, summary_stats, load_audit_events
from .comparison_rows import ComparisonRow
from .runner import SourceValidationReport

_MAX_SHEETS = 15
_CRON_SHEET = "Scheduler & Cron"
_INGRESS_SHEET = "Ingress API"

# Fine-grained audit-events.xlsx categories → coarse workbook tabs
_CATEGORY_BUCKETS: dict[str, tuple[str, ...]] = {
    "Fonts": (
        "FontActivation",
        "FontDeactivation",
        "FontList",
        "FontImport",
        "FontDownload",
        "ProductionFont",
        "FontTemplate (all actor-only)",
        "FontAccess",
    ),
    "Favorites": ("Favorite", "FavoritePair"),
    "Projects & Assets": (
        "ProjectManagement",
        "Asset",
        "WebProject",
        "PinnedAsset + AddOn",
        "PrivateTag",
    ),
    "Users & Teams": (
        "UserProfile",
        "UserRole",
        "Team",
        "Invitation",
        "ServiceAccount",
        "SSO (all actor-only)",
    ),
    "Customer & BYOF": (
        "Customer",
        "LicenseManagement (all actor-only)",
        "BYOF",
        "CompanyLogo (all actor-only)",
    ),
    "Documents": (
        "DocumentScanning",
        "DocumentMetadata (all actor-only)",
        "StyleDocument/Comment (all actor-only)",
    ),
    "Notifications": ("Notification (all actor-only)", "Scheduled"),
    "Read-only": ("Query/Read (all actor-only)",),
}

_FINE_TO_COARSE: dict[str, str] = {
    fine: coarse for coarse, fines in _CATEGORY_BUCKETS.items() for fine in fines
}

_CRON_OPERATIONS: frozenset[str] = frozenset(c.operation for c in load_cron_cases())
_INGRESS_OPERATIONS: frozenset[str] = frozenset(c.operation for c in load_ingress_cases())


def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[:\\/?*\[\]]", "_", name.strip())
    return cleaned[:31] or "Unknown"


def _node_subnode(spec: MappingField) -> str:
    parts = [p for p in (spec.node, spec.sub_node) if p]
    return " / ".join(parts) if parts else ""


def _node_subnode_from_row(row: ComparisonRow) -> str:
    parts = [p for p in (row.node, row.sub_node) if p]
    if parts:
        return " / ".join(parts)
    if row.field_path.startswith("subject.enrichedSnapshot."):
        return row.field_path[len("subject.enrichedSnapshot.") :]
    if row.field_path.startswith("actor.enrichedSnapshot."):
        return row.field_path[len("actor.enrichedSnapshot.") :]
    return row.field_path


def _coarse_categories(operations: list[str]) -> dict[str, list[str]]:
    """Group operations into ≤12 coarse tabs (+ Cron handled separately)."""
    registry = events_by_operation(str(DEFAULT_AUDIT_EVENTS_XLSX))
    coarse: dict[str, list[str]] = {}
    cron_ops: list[str] = []
    ingress_ops: list[str] = []

    for op in operations:
        if op in _CRON_OPERATIONS:
            cron_ops.append(op)
            continue
        if op in _INGRESS_OPERATIONS and op not in registry:
            ingress_ops.append(op)
            continue
        spec = registry.get(op)
        fine = spec.category if spec else "Unknown"
        bucket = _FINE_TO_COARSE.get(fine, "Other")
        coarse.setdefault(bucket, []).append(op)

    if cron_ops:
        coarse[_CRON_SHEET] = sorted(set(cron_ops))
    if ingress_ops:
        coarse[_INGRESS_SHEET] = sorted(set(ingress_ops))
    return coarse


def _write_category_sheet(
    ws,
    *,
    operations: list[str],
    comparison_rows: list[ComparisonRow],
    blank_row_between_events: bool = True,
) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    hdr = [
        "event",
        "field",
        "node/subnode",
        "value_in_enriched_json",
        "value_in_source_json",
        "source",
        "status",
        "remark",
        "routing_key",
    ]
    ws.append(hdr)
    header_fill = PatternFill("solid", fgColor="4472C4")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    registry = events_by_operation(str(DEFAULT_AUDIT_EVENTS_XLSX))
    rows_by_op: dict[str, list[ComparisonRow]] = {}
    for r in comparison_rows:
        rows_by_op.setdefault(r.operation, []).append(r)

    sorted_ops = sorted(operations)
    for idx, op in enumerate(sorted_ops):
        spec = registry.get(op)
        op_rows = sorted(rows_by_op.get(op, []), key=lambda r: r.field_path)
        if op_rows:
            for comp in op_rows:
                ws.append([
                    op,
                    comp.field or comp.field_path.rsplit(".", 1)[-1],
                    _node_subnode_from_row(comp),
                    comp.value_in_enriched,
                    comp.value_in_source,
                    comp.source_system,
                    comp.match_status,
                    comp.notes,
                    spec.routing_key if spec else "",
                ])
        else:
            ws.append([
                op,
                "",
                "",
                "",
                "",
                "",
                "SKIP",
                "No enriched sample or no scalar fields in enriched JSON",
                spec.routing_key if spec else "",
            ])
        if blank_row_between_events and idx < len(sorted_ops) - 1:
            ws.append([""] * len(hdr))

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 28
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 36
    ws.column_dimensions["I"].width = 24
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = f"A1:I{ws.max_row}"


def write_source_validation_workbook(
    *,
    path: Path,
    report: SourceValidationReport,
    comparison_rows: list[ComparisonRow],
    project_root: Path,
    operations: list[str] | None = None,
) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("openpyxl required: pip install openpyxl") from exc

    ops = list(operations or [r.operation for r in report.operations])
    # Always include cron ops that have an enriched capture on disk
    enrich_dir = project_root / "payload" / "enrich"
    ingress_enrich = project_root / "payload" / "ingress" / "enrich"
    for case in load_cron_cases():
        if (enrich_dir / f"{case.operation}.json").is_file() and case.operation not in ops:
            ops.append(case.operation)
    for case in load_ingress_cases():
        if (ingress_enrich / f"{case.operation}.json").is_file() and case.operation not in ops:
            ops.append(case.operation)

    events = load_audit_events()
    by_coarse = _coarse_categories(ops)

    wb = Workbook()
    ws_sum = wb.active
    ws_sum.title = "Summary"
    ws_sum.append(["metric", "value"])
    stats = summary_stats(events)
    ws_sum.append(["registry_total_events", stats["total_events"]])
    ws_sum.append(["registry_categories", stats["categories"]])
    ws_sum.append(["validated_operations", len(ops)])
    ws_sum.append(["workbook_tabs", len(by_coarse)])
    ws_sum.append(["pass", report.passed])
    ws_sum.append(["fail", report.failed])
    ws_sum.append(["skip", report.skipped])
    ws_sum.append(["comparison_rows", len(comparison_rows)])
    ws_sum.append(["discovery_calls", ", ".join(report.discovery_calls)])
    ws_sum.append(["cron_operations", ", ".join(sorted(_CRON_OPERATIONS & set(ops)))])
    ws_sum.append(["ingress_operations", ", ".join(sorted(_INGRESS_OPERATIONS & set(ops)))])
    if report.pandas_summary:
        for k, v in report.pandas_summary.items():
            ws_sum.append([f"pandas_{k}", v])
    for cell in ws_sum[1]:
        cell.font = Font(bold=True)

    # Scheduler & Cron first (if present), then largest coarse groups, cap total sheets
    sheet_order = sorted(
        by_coarse.keys(),
        key=lambda c: (
            0 if c == _CRON_SHEET else 1 if c == _INGRESS_SHEET else 2,
            -len(by_coarse[c]),
            c,
        ),
    )
    if len(sheet_order) > _MAX_SHEETS - 1:
        # Merge smallest buckets into Other
        while len(sheet_order) > _MAX_SHEETS - 1:
            smallest = min(
                (c for c in sheet_order if c not in (_CRON_SHEET, "Other")),
                key=lambda c: len(by_coarse[c]),
                default=None,
            )
            if not smallest:
                break
            by_coarse.setdefault("Other", []).extend(by_coarse.pop(smallest))
            sheet_order = sorted(
                by_coarse.keys(),
                key=lambda c: (
                    0 if c == _CRON_SHEET else 1 if c == _INGRESS_SHEET else 2,
                    -len(by_coarse[c]),
                    c,
                ),
            )

    for cat in sheet_order:
        sheet_name = _safe_sheet_name(cat)
        ws = wb.create_sheet(sheet_name)
        _write_category_sheet(
            ws,
            operations=by_coarse[cat],
            comparison_rows=comparison_rows,
            blank_row_between_events=True,
        )

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
