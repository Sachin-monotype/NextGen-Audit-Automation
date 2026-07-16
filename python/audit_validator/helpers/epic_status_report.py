"""Epic vs actual enrichment status — detailed sheet with event routing key names."""

from __future__ import annotations

import csv
import json
import logging
from pathlib import Path

from ..coverage.matrix import OperationCoverageRow, PipelineStage
from ..csv_report import _SIMULATION_ALIASES, _load_flows_simulation
from .notification_audit_report import _index_dl_by_correlation, _index_enriched_by_correlation, _load_raw_payload
from ..utility.operation_meta import ui_navigation
from ..rabbitmq.resolver_routing_map import expected_routing_key

log = logging.getLogger(__name__)

CONFLUENCE_PAGE = (
    "https://monotype.atlassian.net/wiki/spaces/ENG/pages/7525368143/"
    "Epic+Vs+Actual+Audit+Logs+Status"
)

FIELDS = [
    "graphql_operation",
    "graphql_mutation_field",
    "epic_doc_status",
    "enriched_event_routing_key",
    "enriched_event_name",
    "event_domain",
    "event_action",
    "received_enriched_routing_key",
    "routing_key_matches_expected",
    "ui_navigation",
    "graphql_simulation",
    "raw_audit_event",
    "enriched_audit_event",
    "dead_letter",
    "enriched_structure_valid",
    "raw_enriched_match",
    "validation_status",
    "enrichment_status",
    "issue_detail",
    "x_correlation_id",
    "enriched_source_service",
    "enriched_operation_state",
]


def _routing_key_display(routing_key: str) -> str:
    if not routing_key:
        return ""
    return " ".join(part.replace("_", " ").title() for part in routing_key.split("."))


def _routing_key_parts(routing_key: str) -> tuple[str, str]:
    if not routing_key or "." not in routing_key:
        return routing_key, ""
    domain, action = routing_key.split(".", 1)
    return domain, action


def _load_manifest(project_root: Path) -> dict[str, dict]:
    path = project_root / "python" / "audit_validator" / "data" / "operation_manifest.json"
    data = json.loads(path.read_text(encoding="utf-8"))
    return {o["auditOperation"]: o for o in data.get("operations", [])}


def _load_coverage(project_root: Path) -> dict[str, OperationCoverageRow]:
    path = project_root / "reports" / "e2e" / "coverage-matrix.json"
    if not path.is_file():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, OperationCoverageRow] = {}
    for row in data.get("operations", []):
        stages = {PipelineStage(k): v for k, v in row["stages"].items()}
        out[row["operation"]] = OperationCoverageRow(
            operation=row["operation"],
            template_id=row["template_id"],
            stages=stages,
            x_correlation_id=row.get("x_correlation_id"),
            enriched_routing_key=row.get("enriched_routing_key"),
            expected_routing_key=row.get("expected_routing_key"),
            enrichment_expected=row.get("enrichment_expected", True),
            validation_status=row.get("validation_status"),
            gaps=row.get("gaps", []),
        )
    return out


def _load_validation(project_root: Path) -> dict[str, str]:
    path = project_root / "reports" / "e2e" / "validation.json"
    if not path.is_file():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, str] = {}
    for r in data.get("results", []):
        if isinstance(r, dict) and r.get("operation"):
            out[r["operation"]] = r.get("status", "")
    return out


def _enriched_meta(
    enriched_dir: Path,
    dl_dir: Path,
    correlation_id: str | None,
) -> tuple[str | None, str | None]:
    if not correlation_id:
        return None, None
    payload = _index_enriched_by_correlation(enriched_dir).get(correlation_id)
    if not payload:
        payload = _index_dl_by_correlation(dl_dir).get(correlation_id)
    if not payload:
        return None, None
    source = payload.get("source") or {}
    service = source.get("service")
    state = source.get("operationState")
    return (
        str(service) if service is not None else None,
        str(state) if state is not None else None,
    )


def _enrichment_status(
    *,
    epic_done: bool,
    expected_rk: str,
    received_rk: str,
    raw: str,
    enriched: str,
    dead_letter: str,
    match: str,
    structure: str,
    gaps: list[str],
) -> tuple[str, str]:
    rk = received_rk or expected_rk or "(no routing key)"
    gap = "; ".join(gaps) if gaps else ""

    if not epic_done:
        if enriched == "YES" and match == "YES":
            return (
                f"NOT IN EPIC — enriched event '{rk}' works but not marked Done in doc",
                gap,
            )
        return ("NOT IN EPIC — enrichment not expected / not verified", gap)

    if enriched == "YES" and match == "YES":
        return (
            f"WORKING — enriched event '{rk}' received, structure valid, raw/enriched match",
            "",
        )
    if enriched == "YES" and structure == "NO":
        return (
            f"FAIL — enriched event '{rk}' received but structure validation failed",
            gap or "Enriched structure validation failed",
        )
    if enriched == "YES" and match == "NO":
        return (
            f"FAIL — enriched event '{rk}' received but raw/enriched fields mismatch",
            gap or "Raw vs enriched field mismatch",
        )
    if dead_letter == "YES":
        return (
            f"FAIL — expected enriched event '{expected_rk}' but resolver dead-lettered raw event",
            gap or "Event dead-lettered — enrichment failed",
        )
    if raw == "YES" and enriched == "NO":
        return (
            f"FAIL — raw audit event emitted for '{expected_rk}' but no enriched event on queue",
            gap or "Raw event without enriched counterpart",
        )
    if raw == "NO":
        return (
            f"FAIL — expected enriched event '{expected_rk}' but no raw audit event on queue",
            gap or "GraphQL operation did not produce raw audit event",
        )
    return (f"FAIL — expected enriched event '{expected_rk}'", gap)


def build_epic_status_rows(project_root: Path) -> tuple[list[dict[str, str]], str]:
    manifest = _load_manifest(project_root)
    cov_by = _load_coverage(project_root)
    val_by = _load_validation(project_root)

    flows_path = project_root / "reports" / "e2e" / "flows-results.json"
    sim, _sim_errors = _load_flows_simulation(flows_path) if flows_path.is_file() else ({}, {})

    raw_dir = project_root / "payload" / "raw"
    enriched_dir = project_root / "payload" / "enrich"
    dl_dir = project_root / "dl_events"

    generated_at = ""
    cov_path = project_root / "reports" / "e2e" / "coverage-matrix.json"
    if cov_path.is_file():
        generated_at = json.loads(cov_path.read_text(encoding="utf-8")).get("generatedAt", "")

    all_ops = sorted(manifest.keys())
    rows: list[dict[str, str]] = []

    for op in all_ops:
        meta = manifest[op]
        cov = cov_by.get(op)
        stages = cov.stages if cov else {}
        gaps = list(cov.gaps) if cov else []
        cid = cov.x_correlation_id if cov else None

        expected_rk = expected_routing_key(op) or (cov.expected_routing_key if cov else "") or ""
        received_rk = (cov.enriched_routing_key if cov else None) or ""
        domain, action = _routing_key_parts(expected_rk)

        raw_payload = _load_raw_payload(raw_dir, op, cid)
        if raw_payload and not cid:
            cid = raw_payload.get("xCorrelationId")

        epic_done = "Done" if meta.get("resolverMapped") else "Not Done"
        raw = "YES" if stages.get(PipelineStage.RAW_QUEUE) else "NO"
        dead_letter = "YES" if stages.get(PipelineStage.DEAD_LETTER) else "NO"
        if not meta.get("resolverMapped") and cov and not cov.enrichment_expected:
            enriched = "N/A"
        elif stages.get(PipelineStage.ENRICHED_QUEUE):
            enriched = "YES"
        else:
            enriched = "NO"
        structure = "YES" if stages.get(PipelineStage.STRUCTURE_VALID) else "NO"
        match = "YES" if stages.get(PipelineStage.RAW_ENRICHED_MATCH) else "NO"
        rk_match = "YES" if received_rk and expected_rk and received_rk == expected_rk else (
            "N/A" if not received_rk else "NO"
        )

        service, op_state = _enriched_meta(enriched_dir, dl_dir, cid)
        status, issue = _enrichment_status(
            epic_done=epic_done == "Done",
            expected_rk=expected_rk,
            received_rk=received_rk,
            raw=raw,
            enriched=enriched,
            dead_letter=dead_letter,
            match=match,
            structure=structure,
            gaps=gaps,
        )

        rows.append(
            {
                "graphql_operation": op,
                "graphql_mutation_field": meta.get("rootField", op),
                "epic_doc_status": epic_done,
                "enriched_event_routing_key": expected_rk,
                "enriched_event_name": _routing_key_display(expected_rk),
                "event_domain": domain,
                "event_action": action.replace("_", " "),
                "received_enriched_routing_key": received_rk,
                "routing_key_matches_expected": rk_match,
                "ui_navigation": ui_navigation(op),
                "graphql_simulation": sim.get(op, "NOT_RUN"),
                "raw_audit_event": raw,
                "enriched_audit_event": enriched,
                "dead_letter": dead_letter,
                "enriched_structure_valid": structure,
                "raw_enriched_match": match,
                "validation_status": val_by.get(op, cov.validation_status if cov else "") or "",
                "enrichment_status": status,
                "issue_detail": issue,
                "x_correlation_id": cid or "",
                "enriched_source_service": service or "",
                "enriched_operation_state": op_state or "",
            }
        )

    status_order = {
        "WORKING": 0,
        "FAIL": 1,
        "NOT IN EPIC": 2,
    }

    def sort_key(row: dict[str, str]) -> tuple:
        prefix = row["enrichment_status"].split(" —", 1)[0]
        epic_rank = 0 if row["epic_doc_status"] == "Done" else 1
        return (
            epic_rank,
            status_order.get(prefix, 9),
            row["enriched_event_routing_key"],
            row["graphql_operation"],
        )

    rows.sort(key=sort_key)
    return rows, generated_at


def write_epic_status_csv(*, path: Path, project_root: Path) -> int:
    rows, _ = build_epic_status_rows(project_root)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=FIELDS)
        writer.writeheader()
        writer.writerows(rows)
    log.info("Wrote epic status CSV %s (%d rows)", path, len(rows))
    return len(rows)


def write_epic_status_workbook(*, path: Path, project_root: Path) -> int:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError as exc:
        raise RuntimeError("openpyxl required: pip install openpyxl") from exc

    rows, generated_at = build_epic_status_rows(project_root)
    done_rows = [r for r in rows if r["epic_doc_status"] == "Done"]
    working = [r for r in done_rows if r["enrichment_status"].startswith("WORKING")]
    failing = [r for r in done_rows if not r["enrichment_status"].startswith("WORKING")]

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.append(["Epic Vs Actual — Enriched Event Status"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append(["Confluence", CONFLUENCE_PAGE])
    ws.append(["Done column source", "resolverMapped from mt-audit-log-resolver-service"])
    ws.append(["E2E run at", generated_at])
    ws.append([])
    ws.append(["Metric", "Count"])
    ws.append(["Total GraphQL operations", len(rows)])
    ws.append(["Done in epic (expect enriched event)", len(done_rows)])
    ws.append(["Done + enriched event WORKING", len(working)])
    ws.append(["Done + enriched event FAILING", len(failing)])
    ws.append([])
    ws.append(["Unique enriched event routing keys (Done ops)", ""])
    seen: set[str] = set()
    for r in done_rows:
        rk = r["enriched_event_routing_key"]
        if rk and rk not in seen:
            seen.add(rk)
            ok = sum(
                1
                for x in done_rows
                if x["enriched_event_routing_key"] == rk
                and x["enrichment_status"].startswith("WORKING")
            )
            total = sum(1 for x in done_rows if x["enriched_event_routing_key"] == rk)
            ws.append([rk, f"{ok}/{total} operations working"])

    for title, data in [
        ("All Operations", rows),
        ("Done — Working Events", working),
        ("Done — Failing Events", failing),
        ("Not Done", [r for r in rows if r["epic_doc_status"] == "Not Done"]),
    ]:
        sheet = wb.create_sheet(title)
        sheet.append(FIELDS)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        for row in data:
            sheet.append([row[f] for f in FIELDS])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    log.info("Wrote epic status workbook %s", path)
    return len(rows)
