#!/usr/bin/env python3
"""Build combined web/app audit coverage workbook (covered, missing, manual-only)."""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from audit_validator.operation_sources import operation_source_report
from audit_validator.touchpoint.scenarios import short_touchpoint
from audit_validator.ui_case_recipes import short_touch

_OPERATION_ALIASES: dict[str, str] = {
    "deactivatelist": "deactivatelist",
    "bulkactivatelist": "bulkactivatelists",
    "bulkdeactivatelist": "bulkdeactivatelists",
    "activatefamilies": "activatefamily",
    "deleterole": "deleteroles",
    "bulkaddpairstofavorite": "bulkaddpairstofavorite",
    "bulkremovepairsfromfavorite": "bulkremovepairsfromfavorite",
}

_APP_SCENARIO_EQUIV: dict[str, set[str]] = {
    "global": {"global_app"},
    "global_app": {"global"},
}

_APP_ONLY_SCENARIOS = frozenset({"global_app", "app"})
_PAIRING_OPS = frozenset(
    {
        "addfavoritepair",
        "removefavoritepair",
        "bulkaddpairstofavorite",
        "bulkremovepairsfromfavorite",
    }
)

_SCENARIO_DISPLAY = {
    "global": "global",
    "favourite": "Favourite",
    "list": "List",
    "project": "Project",
    "project_list": "Project > List",
    "global_app": "App (global)",
    "pairing": "pairing",
    "roles": "roles",
    "teams": "teams",
    "manage": "manage",
    "app": "app",
    "reporting": "reporting",
    "company_library": "company_library",
    "company_settings": "company_settings",
    "library_assets": "library_assets",
    "user_access": "user_access",
}


def _norm_scenario(raw: str) -> str:
    s = (raw or "global").strip()
    s = s.replace(">", "_").replace(" ", "_").lower()
    if s in {"project>list", "project_list"}:
        return "project_list"
    return short_touch(s) or short_touchpoint(s) or s or "global"


def _display_scenario(raw: str) -> str:
    sc = _norm_scenario(raw)
    return _SCENARIO_DISPLAY.get(sc, sc.replace("_", " ").title())


def _canonical_ops(catalog: list[dict]) -> dict[str, str]:
    out: dict[str, str] = {}
    for item in catalog:
        op = str(item.get("operation") or "").strip()
        if op:
            out[op.lower()] = op
    return out


def _norm_operation(raw: str, canonical: dict[str, str]) -> str:
    s = str(raw or "").strip()
    if not s:
        return ""
    low = s.lower()
    alias = _OPERATION_ALIASES.get(low, low)
    return canonical.get(alias, canonical.get(low, s))


def _expand_scenarios(operation: str, scenario: str) -> set[str]:
    op = operation.lower()
    sc = _norm_scenario(scenario)
    out = {sc}
    if op in _PAIRING_OPS and sc == "pairing":
        out.update({"global", "favourite"})
    return out


def _literal_covers_target(
    cov_op: str,
    cov_sc: str,
    target_op: str,
    target_sc: str,
    *,
    channel: str,
) -> bool:
    if cov_op.lower() != target_op.lower():
        return False
    if cov_sc == target_sc:
        return True
    if cov_sc == "pairing" and cov_op.lower() in _PAIRING_OPS:
        return target_sc in {"global", "favourite", "pairing"}
    if cov_sc == "roles" and cov_op.lower() in {"createrole", "updaterole", "deleteroles"}:
        return target_sc in {"global", "manage", "roles", "user_access"}
    if cov_sc == "teams" and cov_op.lower() in {"createteam", "updateteam", "deleteteams"}:
        return target_sc in {"global", "manage", "user_access", "teams"}
    if cov_sc == "list" and cov_op.lower() in {"pinasset", "updateasset"}:
        return target_sc in {"global", "list"}
    if cov_sc in {"manage", "global"} and target_sc in {"manage", "global"}:
        if cov_op.lower() in {
            "createrole",
            "updaterole",
            "createteam",
            "updateteam",
            "deleteroles",
            "deleteteams",
        }:
            return True
    if target_sc in _expand_scenarios(cov_op, cov_sc):
        return True
    if cov_sc in _expand_scenarios(target_op, target_sc):
        return True
    return _scenario_matches(cov_sc, target_sc, channel=channel)


def _scenario_matches(covered_scenario: str, target_scenario: str, *, channel: str) -> bool:
    if covered_scenario == target_scenario:
        return True
    if channel == "app":
        equiv = _APP_SCENARIO_EQUIV.get(target_scenario, set())
        if covered_scenario in equiv:
            return True
        equiv_cov = _APP_SCENARIO_EQUIV.get(covered_scenario, set())
        if target_scenario in equiv_cov:
            return True
    return False


def _is_covered(
    event: str,
    scenario: str,
    covered: set[tuple[str, str]],
    *,
    canonical: dict[str, str],
    channel: str,
) -> bool:
    op, target_sc = (
        _norm_operation(event, canonical),
        _norm_scenario(str(scenario or "global")),
    )
    for cov_op, cov_sc in covered:
        if _literal_covers_target(cov_op, cov_sc, op, target_sc, channel=channel):
            return True
    return False


def _rows_to_keys(
    df: pd.DataFrame,
    *,
    canonical: dict[str, str],
    channel: str,
) -> set[tuple[str, str]]:
    keys: set[tuple[str, str]] = set()
    for _, row in df.iterrows():
        op = _norm_operation(str(row["event"]), canonical)
        sc = _norm_scenario(str(row["scenario"]))
        keys.add((op, sc))
        if channel == "app" and sc == "global":
            keys.add((op, "global_app"))
        if channel == "app" and sc == "global_app":
            keys.add((op, "global"))
    return keys


def _missing_display_scenario(raw: str) -> str:
    sc = _norm_scenario(raw)
    labels = {
        "favourite": "favourite",
        "list": "list",
        "global": "global",
        "project": "project",
        "project_list": "project_list",
        "global_app": "global_app",
        "manage": "manage",
        "pairing": "pairing",
        "roles": "roles",
        "teams": "teams",
        "app": "app",
    }
    return labels.get(sc, sc)


def _covered_by_event(covered: set[tuple[str, str]]) -> dict[str, list[str]]:
    by_event: dict[str, list[str]] = defaultdict(list)
    for op, sc in sorted(covered):
        label = _missing_display_scenario(sc)
        if label not in by_event[op]:
            by_event[op].append(label)
    return dict(by_event)


def _read_datasource(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="datasource")
    df.columns = [str(c).strip().lower() for c in df.columns]
    # Merged-style rows leave event_name blank — forward-fill from parent event.
    if "event_name" in df.columns:
        df.loc[:, "event_name"] = df["event_name"].ffill()
    return df


def _ok_only_rows(
    df: pd.DataFrame,
    *,
    canonical: dict[str, str],
) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["event", "scenario", "status"])
    work = df.copy()
    work = work[work["status"].astype(str).str.upper().isin({"OK", "PASS"})]
    out = work[["event_name", "scenario", "status"]].copy()
    out.loc[:, "event_name"] = out["event_name"].map(lambda v: _norm_operation(str(v), canonical))
    out.loc[:, "scenario"] = out["scenario"].map(lambda v: _norm_scenario(str(v)))
    out = out.dropna(subset=["event_name"])
    out = out[out["event_name"].astype(str).str.strip().str.lower().isin({"", "nan", "none"}) == False]
    out = out.drop_duplicates(subset=["event_name", "scenario"], keep="last")
    out.rename(columns={"event_name": "event"}, inplace=True)
    out.loc[:, "status"] = "OK"
    return out.sort_values(["event", "scenario"]).reset_index(drop=True)


def _pending_rows(df: pd.DataFrame, *, canonical: dict[str, str]) -> pd.DataFrame:
    if df.empty or "status" not in df.columns:
        return pd.DataFrame(columns=["event", "scenario", "status", "notes"])
    work = df[~df["status"].astype(str).str.upper().isin({"OK", "PASS", "NAN", ""})]
    work = work[work["status"].astype(str).str.upper().isin({"ERROR", "SKIPPED"})]
    cols = [c for c in ("event_name", "scenario", "status", "notes") if c in work.columns]
    out = work[cols].copy()
    out.loc[:, "event_name"] = out["event_name"].map(lambda v: _norm_operation(str(v), canonical))
    out.loc[:, "scenario"] = out["scenario"].map(lambda v: _norm_scenario(str(v)))
    out = out.dropna(subset=["event_name"])
    out = out.drop_duplicates(subset=["event_name", "scenario"], keep="last")
    out.rename(columns={"event_name": "event"}, inplace=True)
    out.loc[:, "status"] = out["status"].astype(str).str.upper()
    return out.sort_values(["event", "scenario"]).reset_index(drop=True)


def _sheet1_coverage_rows(path: Path, *, canonical: dict[str, str]) -> pd.DataFrame:
    try:
        df = pd.read_excel(path, sheet_name="Sheet1")
    except Exception:
        return pd.DataFrame(columns=["event", "scenario"])
    if "Event" not in df.columns:
        return pd.DataFrame(columns=["event", "scenario"])
    work = df.copy()
    work.loc[:, "Event"] = work["Event"].ffill()
    work = work.rename(columns={"Event": "event", "Source": "scenario"})
    out = work[["event", "scenario"]].copy()
    out.loc[:, "event"] = out["event"].map(lambda v: _norm_operation(str(v), canonical))
    out.loc[:, "scenario"] = out["scenario"].map(lambda v: _norm_scenario(str(v)))
    return out.dropna(subset=["event"]).drop_duplicates(subset=["event", "scenario"])


def _load_sheet1_ui(path: Path, *, canonical: dict[str, str]) -> dict[tuple[str, str], str]:
    try:
        df = pd.read_excel(path, sheet_name="Sheet1")
    except Exception:
        return {}
    if "Event" not in df.columns or "UI Location" not in df.columns:
        return {}
    work = df.copy()
    work.loc[:, "Event"] = work["Event"].ffill()
    out: dict[tuple[str, str], str] = {}
    for _, row in work.iterrows():
        if pd.isna(row.get("Event")):
            continue
        event = _norm_operation(str(row["Event"]), canonical)
        scenario = _norm_scenario(str(row.get("Source") or "global"))
        ui = str(row.get("UI Location") or "").strip()
        if event and ui:
            out[(event, scenario)] = ui
    return out


def _load_touchpoint_paths() -> dict[str, list[str]]:
    path = Path(__file__).resolve().parents[1] / "data" / "ui_navigation_touchpoints.json"
    if not path.is_file():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, list[str]] = {}
    for row in data.get("rows") or []:
        op = str(row.get("operation") or "").strip()
        if not op:
            continue
        paths = [str(p).strip() for p in (row.get("touchpoints") or []) if p]
        if paths:
            out[op] = paths
    return out


def _load_simple_nav() -> dict[str, list[str]]:
    path = Path(__file__).resolve().parents[1] / "data" / "ui_navigation.json"
    if not path.is_file():
        return {}
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, list[str]] = {}
    for op, meta in data.items():
        if isinstance(meta, dict):
            nav = [str(n) for n in (meta.get("navigation") or []) if n]
            if nav:
                out[op] = nav
    return out


def _format_ui_path(raw: str) -> str:
    s = str(raw or "").strip()
    s = re.sub(r"\s*>\s*", " → ", s)
    s = re.sub(r"\s+→\s+", " → ", s)
    return s


def _ui_location(
    event: str,
    scenario: str,
    *,
    sheet1_ui: dict[tuple[str, str], str],
    touchpoints: dict[str, list[str]],
    simple_nav: dict[str, list[str]],
) -> str:
    sc = _norm_scenario(scenario)
    key = (event, sc)
    if key in sheet1_ui:
        return _format_ui_path(sheet1_ui[key])

    hints: dict[str, list[str]] = {
        "global": ["search", "discover", "family card", "global", "dashboard"],
        "favourite": ["favour", "favorite"],
        "list": ["list", "library", "fontlist"],
        "project": ["project"],
        "project_list": ["project", "list"],
        "manage": ["manage", "role", "team"],
        "pairing": ["pair"],
        "reporting": ["report"],
        "company_library": ["company", "library"],
        "company_settings": ["company", "setting"],
        "library_assets": ["library", "asset"],
        "user_access": ["user", "access"],
        "app": ["app", "desktop", "connect"],
        "global_app": ["app", "desktop", "connect"],
    }
    wanted = hints.get(sc, [sc])

    for path in touchpoints.get(event, []):
        low = path.lower()
        if any(h in low for h in wanted):
            return _format_ui_path(path)

    for path in simple_nav.get(event, []):
        low = path.lower()
        if any(h in low for h in wanted):
            return _format_ui_path(path)

    if touchpoints.get(event):
        return _format_ui_path(touchpoints[event][0])
    if simple_nav.get(event):
        return _format_ui_path(simple_nav[event][0])
    return ""


def _covered_display_df(
    ok_rows: pd.DataFrame,
    *,
    sheet1_ui: dict[tuple[str, str], str],
    touchpoints: dict[str, list[str]],
    simple_nav: dict[str, list[str]],
) -> pd.DataFrame:
    rows = []
    for _, row in ok_rows.iterrows():
        event = str(row["event"])
        scenario = str(row["scenario"])
        rows.append(
            {
                "Event": event,
                "scenario": _display_scenario(scenario),
                "Status": "OK",
                "UI location": _ui_location(
                    event,
                    scenario,
                    sheet1_ui=sheet1_ui,
                    touchpoints=touchpoints,
                    simple_nav=simple_nav,
                ),
            }
        )
    if not rows:
        return pd.DataFrame(columns=["Event", "scenario", "Status", "UI location"])
    return pd.DataFrame(rows).sort_values(["Event", "scenario"]).reset_index(drop=True)


_SECTION_REASONS: dict[str, str] = {
    "Auth0 / Login & Identity": (
        "Must be done via Auth0 / identity service — login, logout, and workspace switch "
        "cannot be reliably driven from Playwright UI scripts"
    ),
    "Cron / Scheduler": (
        "Cron / scheduler payloads are backend-triggered — cannot be fired from UI automation scripts"
    ),
    "Plugin (Adobe/Creative Cloud)": (
        "Requires Adobe Creative Cloud plugin support — plugin events are not available via web/app UI scripts"
    ),
    "Font Bridge": (
        "Font Bridge events need the Font Bridge service / desktop bridge — not UI-scriptable"
    ),
    "Desktop Font Sync (local activation)": (
        "Local desktop font sync / activation via Connect ingress — not triggered by web UI scripts"
    ),
    "Desktop Ingress (non-app)": (
        "Desktop ingress event — use API generate, CasePilot, or manual trigger (not Playwright UI script)"
    ),
}


def _classify_non_automatable(item: dict) -> str | None:
    cid = str(item.get("id") or "")
    kind = str(item.get("kind") or "")
    op = str(item.get("operation") or "").lower()
    label = str(item.get("label") or "").lower()

    if kind == "cron" or cid.startswith("cron:"):
        return "Cron / Scheduler"
    if cid.startswith("ingress:plugin_") or "plugin" in cid:
        return "Plugin (Adobe/Creative Cloud)"
    if "fontbridge" in cid or "fontbridge" in op or op == "fontbridgeauthfailed":
        return "Font Bridge"
    if any(
        x in op or x in cid
        for x in (
            "userlogin",
            "userlogout",
            "userswitchworkspace",
            "identitylinked",
            "userlogintiated",
            "auth0",
        )
    ):
        return "Auth0 / Login & Identity"
    if cid.startswith("ingress:") and not cid.startswith("ingress:app_"):
        if "font.localfont" in label or "font_activation" in cid:
            return "Desktop Font Sync (local activation)"
        return "Desktop Ingress (non-app)"
    return None


def _graphql_targets(catalog: list[dict], *, channel: str) -> list[dict]:
    rows = []
    for item in catalog:
        if item.get("kind") != "graphql":
            continue
        op = str(item.get("operation") or "")
        touch = item.get("touchpoint")
        scenario = _norm_scenario(short_touchpoint(touch) if touch else "global")
        if channel == "web" and scenario in _APP_ONLY_SCENARIOS:
            continue
        if channel == "app" and scenario not in _APP_ONLY_SCENARIOS:
            continue
        rows.append({"event": op, "scenario": scenario, "channel": channel})
    return rows


def _automatable_web_targets(catalog: list[dict]) -> list[dict]:
    return _graphql_targets(catalog, channel="web")


def _automatable_app_targets(catalog: list[dict]) -> list[dict]:
    """App automatable set = ingress:app_* + same GraphQL web scenarios (app can fire them)."""
    rows = []
    for item in catalog:
        cid = str(item.get("id") or "")
        if not cid.startswith("ingress:app_"):
            continue
        rows.append(
            {
                "event": str(item.get("operation") or ""),
                "scenario": "app",
                "channel": "app",
            }
        )
    # App datasource covers the same GraphQL mutations as web.
    rows.extend(_graphql_targets(catalog, channel="web"))
    return rows


def _catalog_scenarios_by_event(targets: list[dict], *, canonical: dict[str, str]) -> dict[str, list[str]]:
    by_event: dict[str, list[str]] = defaultdict(list)
    for t in targets:
        op = _norm_operation(t["event"], canonical)
        sc = _norm_scenario(t["scenario"])
        if sc not in by_event[op]:
            by_event[op].append(sc)
    return dict(by_event)


def _coverage_buckets(
    ok_rows: pd.DataFrame,
    targets: list[dict],
    covered: set[tuple[str, str]],
    *,
    canonical: dict[str, str],
    channel: str,
) -> tuple[list[str], list[str]]:
    """Return (fully_covered_events, partially_covered_events) for events that have OK rows."""
    by_cat = _catalog_scenarios_by_event(targets, canonical=canonical)
    ok_events = sorted({_norm_operation(e, canonical) for e in ok_rows["event"].tolist()})
    fully: list[str] = []
    partial: list[str] = []
    for event in ok_events:
        catalog_scs = by_cat.get(event, [])
        if not catalog_scs:
            fully.append(event)
            continue
        missing = [
            sc
            for sc in catalog_scs
            if not _is_covered(event, sc, covered, canonical=canonical, channel=channel)
        ]
        if not missing:
            fully.append(event)
        else:
            partial.append(event)
    return fully, partial


def _aggregate_missing_by_event(
    targets: list[dict],
    covered: set[tuple[str, str]],
    pending: pd.DataFrame,
    *,
    canonical: dict[str, str],
    channel: str,
) -> pd.DataFrame:
    by_event_catalog = _catalog_scenarios_by_event(targets, canonical=canonical)
    events = set(by_event_catalog) | {
        _norm_operation(r["event"], canonical) for _, r in pending.iterrows()
    }
    covered_by = _covered_by_event(covered)
    rows = []

    for event in sorted(events):
        catalog_scenarios = by_event_catalog.get(event, [])
        missing_labels: list[str] = []
        pending_for_event: dict[str, str] = {}

        if not pending.empty:
            for _, prow in pending.iterrows():
                pop = _norm_operation(str(prow["event"]), canonical)
                psc = _norm_scenario(str(prow["scenario"]))
                if pop != event:
                    continue
                if _is_covered(pop, psc, covered, canonical=canonical, channel=channel):
                    continue
                pending_for_event[psc] = str(prow.get("status") or "").upper()

        for sc in catalog_scenarios:
            if _is_covered(event, sc, covered, canonical=canonical, channel=channel):
                continue
            label = _missing_display_scenario(sc)
            status = pending_for_event.get(sc, "")
            if status in {"ERROR", "SKIPPED"}:
                label = f"{label} ({status})"
            missing_labels.append(label)

        for psc, status in pending_for_event.items():
            if psc in catalog_scenarios:
                continue
            label = _missing_display_scenario(psc)
            if status in {"ERROR", "SKIPPED"}:
                label = f"{label} ({status})"
            if label not in missing_labels:
                missing_labels.append(label)

        if not missing_labels:
            continue

        covered_labels = covered_by.get(event, [])
        rows.append(
            {
                "event": event,
                "missing_scenario": ", ".join(missing_labels),
                "coverage_status": "partial" if covered_labels else "none",
                "covered_scenarios": ", ".join(covered_labels),
            }
        )

    if not rows:
        return pd.DataFrame(
            columns=["event", "missing_scenario", "coverage_status", "covered_scenarios"]
        )
    return pd.DataFrame(rows).sort_values(["event"]).reset_index(drop=True)


def _not_automatable_sheet(catalog: list[dict]) -> pd.DataFrame:
    """One row per event, grouped by section, with blank spacer rows between sections."""
    by_section: dict[str, list[str]] = defaultdict(list)
    for item in catalog:
        section = _classify_non_automatable(item)
        if not section:
            continue
        event = str(item.get("operation") or item.get("id") or "").strip()
        if event and event not in by_section[section]:
            by_section[section].append(event)

    section_order = [
        "Auth0 / Login & Identity",
        "Cron / Scheduler",
        "Plugin (Adobe/Creative Cloud)",
        "Font Bridge",
        "Desktop Font Sync (local activation)",
        "Desktop Ingress (non-app)",
    ]
    rows: list[dict] = []
    first = True
    for section in section_order:
        events = sorted(by_section.get(section, []), key=str.lower)
        if not events:
            continue
        if not first:
            rows.append(
                {
                    "section": "",
                    "event": "",
                    "kind": "",
                    "reason": "",
                }
            )
        first = False
        reason = _SECTION_REASONS.get(section, "Not automatable via Playwright UI script")
        for event in events:
            kind = "cron" if section.startswith("Cron") else "ingress"
            rows.append(
                {
                    "section": section,
                    "event": event,
                    "kind": kind,
                    "reason": reason,
                }
            )
    return pd.DataFrame(rows)


def _merge_event_cells(ws, *, event_col: int = 1, start_row: int = 2) -> None:
    """Merge Event column cells for consecutive rows with the same event."""
    max_row = ws.max_row
    if max_row < start_row:
        return
    block_start = start_row
    prev = ws.cell(row=start_row, column=event_col).value
    for row in range(start_row + 1, max_row + 2):
        cur = ws.cell(row=row, column=event_col).value if row <= max_row else None
        if row > max_row or cur != prev:
            if prev and row - 1 > block_start:
                ws.merge_cells(
                    start_row=block_start,
                    start_column=event_col,
                    end_row=row - 1,
                    end_column=event_col,
                )
                cell = ws.cell(row=block_start, column=event_col)
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            block_start = row
            prev = cur


def _autosize_columns(ws, *, max_width: int = 70) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        width = 12
        for row in range(1, ws.max_row + 1):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                width = max(width, min(max_width, len(str(val)) + 2))
        ws.column_dimensions[letter].width = width


def _style_sheet(
    ws,
    *,
    enable_filter: bool = True,
    alternate: bool = True,
    wrap_reason: bool = False,
) -> None:
    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    alt_fill = PatternFill("solid", fgColor="F2F2F2")
    section_fills = {
        "Auth0 / Login & Identity": PatternFill("solid", fgColor="FFF2CC"),
        "Cron / Scheduler": PatternFill("solid", fgColor="E2EFDA"),
        "Plugin (Adobe/Creative Cloud)": PatternFill("solid", fgColor="FCE4D6"),
        "Font Bridge": PatternFill("solid", fgColor="DDEBF7"),
        "Desktop Font Sync (local activation)": PatternFill("solid", fgColor="E4DFEC"),
        "Desktop Ingress (non-app)": PatternFill("solid", fgColor="D9E1F2"),
    }

    max_row, max_col = ws.max_row, ws.max_column
    if max_row < 1 or max_col < 1:
        return

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = thin

    for row in range(2, max_row + 1):
        section_val = str(ws.cell(row=row, column=1).value or "")
        is_spacer = all(
            (ws.cell(row=row, column=c).value in (None, "")) for c in range(1, max_col + 1)
        )
        row_fill = section_fills.get(section_val)
        if is_spacer:
            continue
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=wrap_reason and col == max_col)
            if row_fill is not None:
                cell.fill = row_fill
            elif alternate and row % 2 == 0:
                cell.fill = alt_fill

    if enable_filter and max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    ws.freeze_panes = "A2"


def _summary_df(
    *,
    web_fully: list[str],
    web_partial: list[str],
    app_fully: list[str],
    app_partial: list[str],
    web_ok: pd.DataFrame,
    app_ok: pd.DataFrame,
    not_auto: pd.DataFrame,
) -> pd.DataFrame:
    not_auto_events = (
        not_auto[not_auto["event"].astype(str).str.strip() != ""]["event"].nunique()
        if not not_auto.empty
        else 0
    )
    return pd.DataFrame(
        [
            {"metric": "Events fully covered (web)", "value": len(web_fully)},
            {"metric": "Events partially covered (web)", "value": len(web_partial)},
            {"metric": "Events fully covered (app)", "value": len(app_fully)},
            {"metric": "Events partially covered (app)", "value": len(app_partial)},
            {
                "metric": "Total Event + scenario covered (web)",
                "value": len(web_ok) if not web_ok.empty else 0,
            },
            {
                "metric": "Total Event + scenario covered (App)",
                "value": len(app_ok) if not app_ok.empty else 0,
            },
            {"metric": "Not automatable", "value": not_auto_events},
        ]
    )


def build_workbook(
    *,
    web_path: Path,
    app_path: Path,
    output: Path,
) -> Path:
    catalog = operation_source_report().get("catalog") or []
    canonical = _canonical_ops(catalog)

    web_raw = _read_datasource(web_path)
    app_raw = _read_datasource(app_path)

    web_ok = _ok_only_rows(web_raw, canonical=canonical)
    app_ok = _ok_only_rows(app_raw, canonical=canonical)
    web_pending = _pending_rows(web_raw, canonical=canonical)
    app_pending = _pending_rows(app_raw, canonical=canonical)

    sheet1_ui = _load_sheet1_ui(web_path, canonical=canonical)
    touchpoints = _load_touchpoint_paths()
    simple_nav = _load_simple_nav()

    web_coverage_keys = _rows_to_keys(web_ok, canonical=canonical, channel="web")
    app_coverage_keys = _rows_to_keys(app_ok, canonical=canonical, channel="app")

    web_covered = _covered_display_df(
        web_ok,
        sheet1_ui=sheet1_ui,
        touchpoints=touchpoints,
        simple_nav=simple_nav,
    )
    app_covered = _covered_display_df(
        app_ok,
        sheet1_ui=sheet1_ui,
        touchpoints=touchpoints,
        simple_nav=simple_nav,
    )

    web_targets = _automatable_web_targets(catalog)
    app_targets = _automatable_app_targets(catalog)

    web_fully, web_partial = _coverage_buckets(
        web_ok, web_targets, web_coverage_keys, canonical=canonical, channel="web"
    )
    app_fully, app_partial = _coverage_buckets(
        app_ok, app_targets, app_coverage_keys, canonical=canonical, channel="app"
    )

    missing_web = _aggregate_missing_by_event(
        web_targets,
        web_coverage_keys,
        web_pending,
        canonical=canonical,
        channel="web",
    )
    missing_app = _aggregate_missing_by_event(
        app_targets,
        app_coverage_keys,
        app_pending,
        canonical=canonical,
        channel="app",
    )

    not_auto = _not_automatable_sheet(catalog)
    summary = _summary_df(
        web_fully=web_fully,
        web_partial=web_partial,
        app_fully=app_fully,
        app_partial=app_partial,
        web_ok=web_ok,
        app_ok=app_ok,
        not_auto=not_auto,
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        web_covered.to_excel(writer, sheet_name="Web Covered", index=False)
        app_covered.to_excel(writer, sheet_name="App Covered", index=False)
        missing_web.to_excel(writer, sheet_name="Web Missing", index=False)
        missing_app.to_excel(writer, sheet_name="App Missing", index=False)
        not_auto.to_excel(writer, sheet_name="Not Script Automatable", index=False)

        for sheet in ("Web Covered", "App Covered"):
            ws = writer.sheets[sheet]
            _merge_event_cells(ws, event_col=1)
            _autosize_columns(ws)
            _style_sheet(ws, enable_filter=True, alternate=True)

        for sheet in ("Web Missing", "App Missing"):
            ws = writer.sheets[sheet]
            _autosize_columns(ws)
            _style_sheet(ws, enable_filter=True, alternate=True)

        ws_sum = writer.sheets["Summary"]
        _autosize_columns(ws_sum)
        _style_sheet(ws_sum, enable_filter=False, alternate=True)

        ws_na = writer.sheets["Not Script Automatable"]
        _autosize_columns(ws_na, max_width=90)
        _style_sheet(ws_na, enable_filter=True, alternate=False, wrap_reason=True)
        ws_na.column_dimensions["D"].width = 85

    return output


def main() -> None:
    root = Path(__file__).resolve().parents[3]
    parser = argparse.ArgumentParser(description="Build combined audit coverage workbook")
    parser.add_argument(
        "--web",
        type=Path,
        default=root.parent
        / "MT Connect NextGen/MTConnectAutomation/tests/AuditAutomation/App/web-audit/datasource-web.xlsx",
    )
    parser.add_argument(
        "--app",
        type=Path,
        default=root.parent
        / "MT Connect NextGen/MTConnectAutomation/tests/AuditAutomation/App/web-audit/datasource-app.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=root / "reports" / "audit-coverage-combined.xlsx",
    )
    args = parser.parse_args()
    out = build_workbook(web_path=args.web, app_path=args.app, output=args.output)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
