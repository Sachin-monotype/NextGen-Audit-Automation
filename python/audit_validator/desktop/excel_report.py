"""Excel report for desktop UI navigation automation."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from .validator import DesktopValidationReport


def build_desktop_validation_workbook(report: DesktopValidationReport) -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "Desktop UI Events"
    ws.append(
        [
            "event_name",
            "operation",
            "category",
            "navigation",
            "trigger_hint",
            "ui_status",
            "log_status",
            "validation_status",
            "xCorrelationId",
            "occurred_at",
            "log_file",
            "mongo_raw",
            "mongo_enriched",
            "remarks",
        ]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for row in report.events:
        ws.append(
            [
                row.event_name,
                row.operation,
                row.category,
                row.navigation,
                row.trigger_hint,
                row.ui_status,
                row.log_status,
                row.validation_status,
                row.x_correlation_id,
                row.occurred_at,
                row.log_file,
                "yes" if row.mongo_raw else "no",
                "yes" if row.mongo_enriched else "no",
                row.remarks,
            ]
        )

    ws2 = wb.create_sheet("Selectors")
    ws2.append(["operation", "action", "selector", "xpath", "description"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)

    from .navigation import load_desktop_events

    for ev in load_desktop_events():
        for step in ev.steps:
            ws2.append([ev.operation, step.action, step.selector, step.xpath, step.description])

    ws3 = wb.create_sheet("Summary")
    ws3.append(["metric", "value"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    s = report.summary
    ws3.append(["checked_at", report.checked_at])
    ws3.append(["log_dir", report.log_dir])
    ws3.append(["today_only", "yes" if report.today_only else "no"])
    ws3.append(["total_events", len(report.events)])
    ws3.append(["pass", s.pass_count])
    ws3.append(["fail", s.fail_count])
    ws3.append(["skip", s.skip_count])
    ws3.append(["manual", s.manual_count])

    return wb


def write_desktop_validation_xlsx(report: DesktopValidationReport, out_path: Path) -> Path:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_desktop_validation_workbook(report)
    wb.save(out_path)
    return out_path
