"""CSV / Excel report — one row per operation with simulation and queue verification."""

from __future__ import annotations

import csv
import json
from dataclasses import dataclass
from pathlib import Path

from .coverage.matrix import CoverageMatrix, PipelineStage
from .coverage.correlation_selection import best_validation_per_operation
from .models import ValidationResult, ValidationStatus
from .operation_catalog import expects_enriched_event, is_validator_pending
from .utility.operation_meta import (
    build_curl,
    build_curl_resolved,
    execute_operation_preview,
    load_curl_context,
    ui_navigation,
)
from .rabbitmq.resolver_routing_map import expected_routing_key
from .operation_registry import e2e_expected_operations, tracked_operations
from .template_registry import OPERATION_TEMPLATE_MAP

# runOperation label in flows -> audit operation name in template map
_SIMULATION_ALIASES: dict[str, str] = {
    "markAsProductionFont": "markProductionFonts",
    "createAsset (FontList)": "createAsset",
    "createAsset (Folder)": "createAsset",
    "updateAssetSharing (GRANT)": "updateAssetSharing",
    "updateAssetSharing (REVOKE)": "updateAssetSharing",
    "deleteProject (duplicate)": "deleteProject",
    "deleteProject (original)": "deleteProject",
    "deleteAssets (Copied FontList)": "deleteAssets",
    "deleteAssets (Folder)": "deleteAssets",
    "updatePrivateTagAssociations (disassociate)": "updatePrivateTagAssociations",
}


def _load_flows_simulation(path: Path) -> tuple[dict[str, str], dict[str, str]]:
    """operation -> (status, error message)."""
    if not path.is_file():
        return {}, {}
    data = json.loads(path.read_text(encoding="utf-8"))
    status_out: dict[str, str] = {}
    error_out: dict[str, str] = {}
    for flow in data.get("flows", []):
        for r in flow.get("results", []):
            op = r["operation"]
            audit_op = _SIMULATION_ALIASES.get(op, op)
            status_out[audit_op] = r.get("status", "UNKNOWN")
            if r.get("error"):
                error_out[audit_op] = str(r["error"])
    return status_out, error_out


def _load_cron_results(path: Path | None) -> list[dict]:
    if not path or not path.is_file():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    cases = data.get("cases", [])
    return [c for c in cases if isinstance(c, dict)]


def _load_ingress_results(path: Path | None) -> list[dict]:
    if not path or not path.is_file():
        return []
    data = json.loads(path.read_text(encoding="utf-8"))
    cases = data.get("cases", [])
    return [c for c in cases if isinstance(c, dict)]


def _cron_rows_by_operation(cases: list[dict]) -> dict[str, dict]:
    """Last cron case per operation (for Results tab merge)."""
    out: dict[str, dict] = {}
    for case in cases:
        op = str(case.get("operation") or "")
        if op:
            out[op] = case
    return out


def _ingress_rows_by_case(cases: list[dict]) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for case in cases:
        cid = str(case.get("case_id") or "")
        if cid:
            out[cid] = case
    return out


def _raw_status(row_stages: dict) -> str:
    if row_stages.get(PipelineStage.RAW_QUEUE):
        return "YES"
    return "NO"


def _enriched_status(row_stages: dict, operation: str) -> str:
    if not expects_enriched_event(operation):
        return "N/A"
    if row_stages.get(PipelineStage.ENRICHED_QUEUE):
        return "YES"
    if row_stages.get(PipelineStage.DEAD_LETTER):
        return "DEAD_LETTER"
    if row_stages.get(PipelineStage.RAW_QUEUE):
        return "NO"
    return "NO"


@dataclass(frozen=True)
class E2EOutcome:
    status: str
    reason: str


def _check_messages(val: ValidationResult | None, check: str) -> str | None:
    if not val:
        return None
    for c in val.checks:
        if c.check == check:
            return c.message
    return None


def _validation_detail(val: ValidationResult | None) -> str:
    if not val:
        return ""
    fails = [c for c in val.checks if c.status in {ValidationStatus.FAIL, ValidationStatus.WARN}]
    if not fails:
        return ""
    return "; ".join(f"{c.check}: {c.message}" for c in fails[:4])


def _validator_pending_reason(val: ValidationResult | None) -> bool:
    if not val:
        return False
    fail_checks = {c.check for c in val.checks if c.status == ValidationStatus.FAIL}
    return fail_checks <= {
        "template",
        "subject_enrichment_added",
        "missing_field",
        "subject_enrichment",
    }


def resolve_e2e_outcome(
    operation: str,
    *,
    simulation: str,
    sim_error: str,
    raw: str,
    enriched: str,
    val: ValidationResult | None,
) -> E2EOutcome:
    enrichment_expected = expects_enriched_event(operation)

    if is_validator_pending(operation) and simulation == "PASS" and raw == "YES":
        if enriched in ("YES", "N/A") and _validator_pending_reason(val):
            return E2EOutcome("SKIP", "Validator template not implemented yet")

    if simulation in ("SKIP", "NOT_RUN", "UNKNOWN", "XFAIL") and raw == "NO":
        if simulation == "NOT_RUN":
            return E2EOutcome("SKIP", "Operation not run in this flows session")
        if simulation == "XFAIL":
            return E2EOutcome("SKIP", sim_error or "Expected failure / not supported in automation")
        return E2EOutcome("SKIP", sim_error or f"Simulation {simulation}")

    if simulation == "FAIL":
        return E2EOutcome("FAIL", sim_error or "GraphQL simulation failed")

    if enrichment_expected and raw == "YES" and enriched == "NO":
        detail = _check_messages(val, "enriched_timeout") or "Raw on queue but enriched event missing"
        return E2EOutcome("FAIL", detail)

    if enrichment_expected and enriched == "DEAD_LETTER":
        detail = _check_messages(val, "dead_letter") or "Raw event routed to resolver DLQ"
        return E2EOutcome("FAIL", detail)

    if val and val.status == ValidationStatus.FAIL:
        return E2EOutcome("FAIL", _validation_detail(val) or "Validation failed")

    if val and val.status == ValidationStatus.WARN:
        hard = [
            c
            for c in val.checks
            if c.status in {ValidationStatus.FAIL, ValidationStatus.WARN}
            and c.check in {"dead_letter", "enriched_timeout", "subject_snap_keys", "missing_field"}
        ]
        if hard:
            return E2EOutcome("FAIL", hard[0].message)
        if raw == "YES" and enriched in ("YES", "N/A"):
            return E2EOutcome("PASS", "")

    if raw == "NO":
        return E2EOutcome("FAIL", "No raw audit event on queue")

    if raw == "YES" and enriched in ("YES", "N/A"):
        return E2EOutcome("PASS", "")

    return E2EOutcome("FAIL", "Did not meet pass criteria")


def flows_to_retry(project_root: Path, *, results_path: Path | None = None) -> frozenset[str] | None:
    """Map non-PASS ops from a previous results.csv to simulation flow names."""
    from .simulation.flow_catalog import flow_operations

    path = results_path or (project_root / "temp" / "results.csv")
    if not path.is_file():
        return None

    retry_ops: set[str] = set()
    with path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            overall = (row.get("overall") or row.get("status") or "").strip().upper()
            if overall and overall != "PASS":
                retry_ops.add(row["operation"])

    if not retry_ops:
        return frozenset()

    flows: set[str] = set()
    for fo in flow_operations():
        if fo.graphql_operation in retry_ops:
            flows.add(fo.flow)
    return frozenset(flows)


@dataclass(frozen=True)
class PassedRunKeys:
    """Operations/cases that fully passed in the latest result/result.xlsx."""

    gql_operations: frozenset[str]
    cron_operations: frozenset[str]
    ingress_case_ids: frozenset[str]


def passed_keys_from_result_workbook(
    project_root: Path,
    *,
    path: Path | None = None,
) -> PassedRunKeys | None:
    """Read Pass sheet from result/result.xlsx (or explicit path)."""
    from .report_paths import result_xlsx

    xlsx_path = path or result_xlsx(project_root)
    if not xlsx_path.is_file():
        return None

    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Pass" not in wb.sheetnames:
        wb.close()
        return None

    gql_ops: set[str] = set()
    cron_ops: set[str] = set()
    ingress_cases: set[str] = set()

    ws = wb["Pass"]
    rows = ws.iter_rows(min_row=2, values_only=True)
    for row in rows:
        if not row or len(row) < 7:
            continue
        key = str(row[0] or "").strip()
        trigger = str(row[2] or "").strip().upper()
        overall = str(row[6] or "").strip().upper()
        if not key or overall != "PASS":
            continue
        if trigger == "GQL":
            gql_ops.add(key)
        elif trigger == "CRON":
            cron_ops.add(key)
        elif trigger == "INGRESS":
            ingress_cases.add(key)

    wb.close()
    if not gql_ops and not cron_ops and not ingress_cases:
        return None

    return PassedRunKeys(
        gql_operations=frozenset(gql_ops),
        cron_operations=frozenset(cron_ops),
        ingress_case_ids=frozenset(ingress_cases),
    )


def flow_filter_skip_passed(project_root: Path) -> frozenset[str] | None:
    """Flows to run when skipping operations that passed in the latest result workbook."""
    from .simulation.flow_catalog import flow_operations

    passed = passed_keys_from_result_workbook(project_root)
    if passed is None:
        return None

    flows: set[str] = set()
    for fo in flow_operations():
        if fo.graphql_operation not in passed.gql_operations:
            flows.add(fo.flow)
    return frozenset(flows)


def cron_case_filter_skip_passed(project_root: Path) -> frozenset[str] | None:
    """Cron case_ids to run when skipping passed cron operations."""
    from .cron.payloads import load_cron_cases

    passed = passed_keys_from_result_workbook(project_root)
    if passed is None:
        return None

    remaining = {
        case.case_id
        for case in load_cron_cases()
        if case.operation not in passed.cron_operations
    }
    return frozenset(remaining)


def flows_to_retry_incomplete(
    project_root: Path,
    *,
    path: Path | None = None,
    include_skip: bool = True,
) -> frozenset[str] | None:
    """Map FAIL/SKIP/NOT_RUN GQL rows from result.xlsx to simulation flow names."""
    from .report_paths import result_xlsx
    from .simulation.flow_catalog import flow_operations

    xlsx_path = path or result_xlsx(project_root)
    if not xlsx_path.is_file():
        return None

    try:
        from openpyxl import load_workbook
    except ImportError:
        return None

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    retry_ops: set[str] = set()
    sheets = ["Fail"]
    if include_skip:
        sheets.append("Skip")

    for sheet_name in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 7:
                continue
            op = str(row[0] or "").strip()
            trigger = str(row[2] or "").strip().upper()
            overall = str(row[6] or "").strip().upper()
            sim = str(row[3] or "").strip().upper()
            if trigger != "GQL" and trigger != "NOT_RUN":
                continue
            if overall == "FAIL":
                retry_ops.add(op)
            elif include_skip and overall == "SKIP":
                if sim in {"SKIP", "NOT_RUN", "XFAIL", "FAIL"}:
                    retry_ops.add(op)

    wb.close()
    if not retry_ops:
        return frozenset()

    flows: set[str] = set()
    for fo in flow_operations():
        if fo.graphql_operation in retry_ops:
            flows.add(fo.flow)
    return frozenset(flows)


def ingress_case_filter_skip_passed(project_root: Path) -> frozenset[str] | None:
    """Ingress case_ids to run when skipping passed ingress cases."""
    from .ingress.payloads import load_ingress_cases

    passed = passed_keys_from_result_workbook(project_root)
    if passed is None:
        return None

    all_ids = {case.case_id for case in load_ingress_cases()}
    remaining = all_ids - passed.ingress_case_ids
    return frozenset(remaining)


@dataclass(frozen=True)
class PipelineStatusRow:
    operation: str
    bucket: str
    failure_category: str
    failure_detail: str
    expected_routing_key: str
    simulation: str
    raw_queue: str
    enriched_queue: str
    validation_status: str
    template_id: str
    correlation_id: str


_PIPELINE_COLUMNS = [
    "operation",
    "expected_routing_key",
    "simulation",
    "raw_queue",
    "enriched_queue",
    "validation_status",
    "template_id",
    "correlation_id",
    "failure_category",
    "failure_detail",
]


def classify_pipeline_status(
    operation: str,
    *,
    simulation: str,
    sim_error: str = "",
    stages: dict,
    val: ValidationResult | None,
    correlation_id: str = "",
) -> PipelineStatusRow:
    template_id = OPERATION_TEMPLATE_MAP.get(operation, "unknown")
    raw = _raw_status(stages)
    enriched = _enriched_status(stages, operation)
    validation_status = val.status.value if val else ""
    rk = expected_routing_key(operation) or ""

    outcome = resolve_e2e_outcome(
        operation,
        simulation=simulation,
        sim_error=sim_error,
        raw=raw,
        enriched=enriched,
        val=val,
    )

    if outcome.status == "PASS":
        detail = ""
        if not expects_enriched_event(operation) and raw == "YES":
            detail = "Query/read — raw only (enrichment N/A)"
        return PipelineStatusRow(
            operation=operation,
            bucket="validated",
            failure_category="",
            failure_detail=detail,
            expected_routing_key=rk,
            simulation=simulation,
            raw_queue=raw,
            enriched_queue=enriched,
            validation_status="PASS",
            template_id=template_id,
            correlation_id=correlation_id,
        )

    if outcome.status == "SKIP":
        category = "not_simulated"
        if is_validator_pending(operation):
            category = "validator_pending"
        return PipelineStatusRow(
            operation=operation,
            bucket="skipped",
            failure_category=category,
            failure_detail=outcome.reason,
            expected_routing_key=rk,
            simulation=simulation,
            raw_queue=raw,
            enriched_queue=enriched,
            validation_status="SKIP",
            template_id=template_id,
            correlation_id=correlation_id,
        )

    category = "other"
    if simulation == "FAIL":
        category = "simulation_fail"
    elif enriched == "DEAD_LETTER":
        category = "dead_letter"
    elif enriched == "NO" and expects_enriched_event(operation) and raw == "YES":
        category = "not_enriched"
    elif val and val.status == ValidationStatus.FAIL:
        category = "enriched_invalid"
        if any(c.check in {"font_family_name", "font_style_name", "font_family_id"} for c in val.checks):
            category = "notification_fields_missing"
        elif any(c.check == "notification_placeholder" for c in val.checks):
            category = "notification_placeholder"
    elif raw == "NO":
        category = "no_raw"

    return PipelineStatusRow(
        operation=operation,
        bucket="failed",
        failure_category=category,
        failure_detail=outcome.reason,
        expected_routing_key=rk,
        simulation=simulation,
        raw_queue=raw,
        enriched_queue=enriched,
        validation_status="FAIL",
        template_id=template_id,
        correlation_id=correlation_id,
    )


def build_pipeline_status_rows(
    *,
    validation_results: list[ValidationResult],
    coverage: CoverageMatrix | None,
    flows_results_path: Path | None = None,
    operations: list[str] | None = None,
) -> list[PipelineStatusRow]:
    sim, sim_errors = _load_flows_simulation(flows_results_path) if flows_results_path else ({}, {})
    val_by_op = best_validation_per_operation(validation_results)
    cov_by_op = {row.operation: row for row in coverage.operations} if coverage else {}
    op_list = operations or sorted(e2e_expected_operations())

    rows: list[PipelineStatusRow] = []
    for op in op_list:
        cov = cov_by_op.get(op)
        stages = cov.stages if cov else {}
        val = val_by_op.get(op)
        cid = (cov.x_correlation_id if cov else "") or ""
        rows.append(
            classify_pipeline_status(
                op,
                simulation=sim.get(op, "NOT_RUN"),
                sim_error=sim_errors.get(op, ""),
                stages=stages,
                val=val,
                correlation_id=cid or "",
            )
        )
    return rows


def write_pipeline_status_workbook(
    *,
    path: Path,
    validation_results: list[ValidationResult],
    coverage: CoverageMatrix | None,
    flows_results_path: Path | None = None,
    project_root: Path | None = None,
    operations: list[str] | None = None,
) -> dict[str, int]:
    """Excel workbook with Validated / Dead Letter / Failed pipeline buckets."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel export: pip install openpyxl") from exc

    rows = build_pipeline_status_rows(
        validation_results=validation_results,
        coverage=coverage,
        flows_results_path=flows_results_path,
        operations=operations,
    )
    passed = [r for r in rows if r.bucket == "validated"]
    failed = [r for r in rows if r.bucket == "failed"]
    skipped = [r for r in rows if r.bucket == "skipped"]

    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["bucket", "count", "description"])
    ws_summary.append(["Pass", len(passed), "Simulation + queue + validation OK"])
    ws_summary.append(["Fail", len(failed), "Simulation, enrichment, or validation gap"])
    ws_summary.append(["Skip", len(skipped), "Not run or validator not implemented"])

    def _append_sheet(title: str, data: list[PipelineStatusRow]) -> None:
        ws = wb.create_sheet(title)
        ws.append(_PIPELINE_COLUMNS)
        for row in data:
            ws.append(
                [
                    row.operation,
                    row.expected_routing_key,
                    row.simulation,
                    row.raw_queue,
                    row.enriched_queue,
                    row.validation_status,
                    row.template_id,
                    row.correlation_id,
                    row.failure_category,
                    row.failure_detail,
                ]
            )

    _append_sheet("Pass", passed)
    _append_sheet("Fail", failed)
    _append_sheet("Skip", skipped)

    if project_root is not None:
        ws_ops = wb.create_sheet("Operations")
        ws_ops.append(["operation", "UI_Navigation", "cURL"])
        for op in tracked_operations():
            ws_ops.append([op, ui_navigation(op), build_curl(op, project_root)])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return {
        "pass": len(passed),
        "fail": len(failed),
        "skip": len(skipped),
        # backward-compatible keys
        "validated": len(passed),
        "dead_letter": 0,
        "failed": len(failed),
    }


def write_e2e_csv(
    *,
    path: Path,
    validation_results: list[ValidationResult],
    coverage: CoverageMatrix | None,
    flows_results_path: Path | None = None,
) -> None:
    sim, sim_errors = _load_flows_simulation(flows_results_path) if flows_results_path else ({}, {})

    val_by_op = best_validation_per_operation(validation_results)

    cov_by_op = {row.operation: row for row in coverage.operations} if coverage else {}

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(
            [
                "operation",
                "expected_routing_key",
                "simulation",
                "raw_queue",
                "enriched_queue",
                "overall",
                "reason",
            ]
        )

        for op in sorted(e2e_expected_operations()):
            row = cov_by_op.get(op)
            stages = row.stages if row else {}
            val = val_by_op.get(op)
            simulation = sim.get(op, "NOT_RUN")
            raw = _raw_status(stages)
            enriched = _enriched_status(stages, op)
            outcome = resolve_e2e_outcome(
                op,
                simulation=simulation,
                sim_error=sim_errors.get(op, ""),
                raw=raw,
                enriched=enriched,
                val=val,
            )
            writer.writerow(
                [
                    op,
                    expected_routing_key(op) or "",
                    simulation,
                    raw,
                    enriched,
                    outcome.status,
                    outcome.reason,
                ]
            )


def write_operations_reference_csv(*, path: Path, project_root: Path) -> None:
    """Second sheet data: operation, UI_Navigation, cURL."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(["operation", "UI_Navigation", "cURL"])
        for op in tracked_operations():
            writer.writerow([op, ui_navigation(op), build_curl(op, project_root)])


def write_e2e_workbook(
    *,
    path: Path,
    validation_results: list[ValidationResult],
    coverage: CoverageMatrix | None,
    flows_results_path: Path | None = None,
    cron_results_path: Path | None = None,
    ingress_results_path: Path | None = None,
    project_root: Path,
) -> None:
    """Excel workbook: Results (GQL + cron + ingress) + Pass/Fail/Skip + detail tabs."""
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for Excel export: pip install openpyxl") from exc

    sim, sim_errors = _load_flows_simulation(flows_results_path) if flows_results_path else ({}, {})
    cron_cases = _load_cron_results(cron_results_path)
    cron_by_op = _cron_rows_by_operation(cron_cases)
    ingress_cases = _load_ingress_results(ingress_results_path)
    ingress_by_case = _ingress_rows_by_case(ingress_cases)
    val_by_op = best_validation_per_operation(validation_results)
    cov_by_op = {row.operation: row for row in coverage.operations} if coverage else {}

    all_ops = sorted(set(e2e_expected_operations()) | set(cron_by_op.keys()))

    rows: list[list[str | int]] = []
    for op in all_ops:
        row = cov_by_op.get(op)
        stages = row.stages if row else {}
        val = val_by_op.get(op)
        cron_case = cron_by_op.get(op)
        if cron_case and op not in sim:
            simulation = "CRON"
            sim_err = str(cron_case.get("error") or "")
            raw = "YES" if cron_case.get("publish_status") == "PASS" else "NO"
            enrich = (
                "YES"
                if cron_case.get("enrich_status") == "PASS"
                else "DEAD_LETTER"
                if cron_case.get("enrich_status") == "DLQ"
                else "NO"
            )
            outcome_status = str(cron_case.get("validation_status") or "FAIL")
            outcome_reason = sim_err or f"cron case {cron_case.get('case_id', '')}"
            trigger = "CRON"
        else:
            simulation = sim.get(op, "NOT_RUN")
            sim_err = sim_errors.get(op, "")
            raw = _raw_status(stages)
            enriched = _enriched_status(stages, op)
            outcome = resolve_e2e_outcome(
                op,
                simulation=simulation,
                sim_error=sim_err,
                raw=raw,
                enriched=enriched,
                val=val,
            )
            outcome_status = outcome.status
            outcome_reason = outcome.reason
            enrich = enriched
            trigger = "GQL" if simulation not in ("NOT_RUN", "SKIP", "") else "NOT_RUN"

        rows.append(
            [
                op,
                expected_routing_key(op) or str((cron_case or {}).get("routing_key") or ""),
                trigger,
                simulation,
                raw,
                enrich,
                outcome_status,
                outcome_reason,
            ]
        )

    for case_id, ingress_case in sorted(ingress_by_case.items()):
        raw = "YES" if ingress_case.get("raw_status") == "PASS" else "NO"
        enrich_st = str(ingress_case.get("enrich_status") or "")
        enrich = "YES" if enrich_st == "PASS" else "NO" if enrich_st else "NO"
        rows.append(
            [
                case_id,
                str(ingress_case.get("operation") or ""),
                "INGRESS",
                str(ingress_case.get("category") or "ingress-api"),
                raw,
                enrich,
                str(ingress_case.get("validation_status") or "FAIL"),
                str(ingress_case.get("error") or ingress_case.get("event_name") or ""),
            ]
        )

    header = [
        "operation",
        "expected_routing_key",
        "trigger",
        "simulation",
        "raw_queue",
        "enriched_queue",
        "overall",
        "reason",
    ]

    wb = Workbook()
    ws_all = wb.active
    ws_all.title = "Results"
    ws_all.append(header)
    for row in rows:
        ws_all.append(row)

    for title, status in (("Pass", "PASS"), ("Fail", "FAIL"), ("Skip", "SKIP")):
        ws = wb.create_sheet(title)
        ws.append(header)
        for row in rows:
            if row[6] == status:
                ws.append(row)

    if cron_cases:
        ws_cron = wb.create_sheet("Cron")
        cron_header = [
            "case_id",
            "operation",
            "routing_key",
            "service",
            "correlation_id",
            "publish_status",
            "enrich_status",
            "validation_status",
            "error",
            "jira_refs",
        ]
        ws_cron.append(cron_header)
        for case in cron_cases:
            refs = case.get("jira_refs") or []
            ws_cron.append(
                [
                    case.get("case_id", ""),
                    case.get("operation", ""),
                    case.get("routing_key", ""),
                    case.get("service", ""),
                    case.get("correlation_id", ""),
                    case.get("publish_status", ""),
                    case.get("enrich_status", ""),
                    case.get("validation_status", ""),
                    case.get("error", ""),
                    ", ".join(refs) if isinstance(refs, list) else str(refs),
                ]
            )

    if ingress_cases:
        from .ingress.test_cases import load_ingress_test_cases

        curl_by_case = {c.case_id: c.curl_file for c in load_ingress_test_cases()}
        ws_ingress = wb.create_sheet("Ingress")
        ingress_header = [
            "case_id",
            "event_name",
            "category",
            "operation",
            "service",
            "correlation_id",
            "publish_status",
            "raw_status",
            "enrich_status",
            "validation_status",
            "error",
            "curl_script",
        ]
        ws_ingress.append(ingress_header)
        for case in ingress_cases:
            cid = str(case.get("case_id") or "")
            ws_ingress.append(
                [
                    cid,
                    case.get("event_name", ""),
                    case.get("category", ""),
                    case.get("operation", ""),
                    case.get("service", ""),
                    case.get("correlation_id", ""),
                    case.get("publish_status", ""),
                    case.get("raw_status", ""),
                    case.get("enrich_status", ""),
                    case.get("validation_status", ""),
                    case.get("error", ""),
                    curl_by_case.get(cid, f"curls/{cid}.sh"),
                ]
            )

    ws_ops = wb.create_sheet("Operations")
    ws_ops.append(["operation", "UI_Navigation", "cURL"])
    for op in tracked_operations():
        ws_ops.append([op, ui_navigation(op), build_curl(op, project_root)])

    from .ingress.test_cases import load_ingress_test_cases

    for tc in load_ingress_test_cases():
        ws_ops.append([tc.case_id, f"Ingress API / {tc.category}", tc.curl_file])

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


_STAGE_BY_CATEGORY = {
    "simulation_fail": "GQL simulation",
    "no_raw": "Raw payload",
    "not_enriched": "Enrichment",
    "dead_letter": "Enrichment",
    "enriched_invalid": "Enrichment",
    "notification_fields_missing": "Enrichment",
    "notification_placeholder": "Enrichment",
    "enriched_warn": "Enrichment",
    "not_simulated": "Not simulated",
    "other": "Other",
}


def _failure_stage(row: PipelineStatusRow) -> str:
    if row.failure_category in _STAGE_BY_CATEGORY:
        return _STAGE_BY_CATEGORY[row.failure_category]
    if row.simulation == "FAIL":
        return "GQL simulation"
    if expects_enriched_event(row.operation) and row.raw_queue == "NO":
        return "Raw payload"
    if expects_enriched_event(row.operation) and row.raw_queue == "YES" and row.enriched_queue == "NO":
        return "Enrichment"
    return "Other"


def _truncate_failure(text: str, limit: int = 480) -> str:
    text = " ".join((text or "").split())
    if len(text) <= limit:
        return text
    return text[: limit - 3] + "..."


def _failure_reason(row: PipelineStatusRow, gql_errors: dict[str, str]) -> str:
    if row.failure_category == "simulation_fail":
        return gql_errors.get(row.operation) or row.failure_detail or "GraphQL simulation returned FAIL"
    if row.failure_detail:
        return _truncate_failure(row.failure_detail)
    if row.failure_category == "no_raw":
        return "No raw audit event captured on queue after simulation"
    if row.failure_category == "not_enriched":
        if row.raw_queue == "YES":
            return "Raw event captured but enriched event missing or timed out"
        return "Enriched event not received"
    return row.failure_category.replace("_", " ")


def _include_failure_row(row: PipelineStatusRow) -> bool:
    if row.bucket == "validated":
        return False
    if row.failure_category == "not_simulated":
        return False
    if row.simulation in ("SKIP", "XFAIL") and row.raw_queue == "NO" and row.enriched_queue in ("NO", "N/A"):
        return False
    return True

