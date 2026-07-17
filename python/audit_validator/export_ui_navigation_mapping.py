"""Build multi-sheet Excel mapping UI Navigation events ↔ automation FLOW_DEFS."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[2]


def _load_sheet_rows(xlsx: Path) -> list[dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(xlsx, data_only=True)
    rows_out: list[dict[str, Any]] = []

    # Touch Points — Event / TouchPoint / steps with curls
    if "Touch Points" in wb.sheetnames:
        ws = wb["Touch Points"]
        raw = list(ws.iter_rows(values_only=True))
        for r in raw[1:]:
            event = str(r[0] or "").strip()
            if not event:
                continue
            touch = str(r[1] or "").strip()
            steps = [str(c).strip() for c in r[3:8] if c and str(c).strip()]
            rows_out.append(
                {
                    "source_sheet": "Touch Points",
                    "event": event,
                    "touchpoint": touch,
                    "web_app": str(r[2] or "").strip(),
                    "step_count": len(steps),
                    "has_curl": any("curl" in s.lower() or "mutation" in s.lower() for s in steps),
                }
            )

    # UI Navigation by Section — Operation/Events + navigation paths
    if "UI Navigation by Section" in wb.sheetnames:
        ws = wb["UI Navigation by Section"]
        raw = list(ws.iter_rows(values_only=True))
        for r in raw[1:]:
            section = str(r[0] or "").strip()
            event = str(r[1] or "").strip()
            if not event:
                continue
            navs = [str(c).strip() for c in r[2:8] if c and str(c).strip() and str(c).strip() != "-"]
            rows_out.append(
                {
                    "source_sheet": "UI Navigation by Section",
                    "section": section,
                    "event": event,
                    "touchpoint": navs[0] if navs else "",
                    "navigation_paths": " | ".join(navs),
                    "path_count": len(navs),
                }
            )

    if "App based Events" in wb.sheetnames:
        ws = wb["App based Events"]
        raw = list(ws.iter_rows(values_only=True))
        for r in raw[1:]:
            event = str(r[1] or r[0] or "").strip()
            if not event:
                continue
            rows_out.append(
                {
                    "source_sheet": "App based Events",
                    "section": str(r[0] or "").strip(),
                    "event": event,
                    "touchpoint": str(r[2] or "").strip(),
                }
            )
    return rows_out


def _norm_event(name: str) -> str:
    s = re.sub(r"[^a-zA-Z0-9]", "", (name or "").lower())
    return s


def build_ui_navigation_mapping_workbook(
    *,
    project_root: Path | None = None,
    sheet_xlsx: Path | None = None,
) -> Workbook:
    root = project_root or _repo_root()
    xlsx = sheet_xlsx or (root / "docs" / "UI Navigation of Event (2).xlsx")

    from audit_validator.touchpoint.payloads import FLOW_DEFS
    from audit_validator.touchpoint.scenarios import list_scenarios
    from audit_validator.operation_sources import operation_source_report

    sheet_rows = _load_sheet_rows(xlsx) if xlsx.is_file() else []
    scenarios = list_scenarios()
    catalog = operation_source_report().get("catalog") or []

    flow_ops = {op.lower(): op for op in FLOW_DEFS}
    sheet_events = sorted(
        {_norm_event(r["event"]) for r in sheet_rows if r.get("event")},
        key=str,
    )
    sheet_by_norm: dict[str, list[dict[str, Any]]] = {}
    for r in sheet_rows:
        sheet_by_norm.setdefault(_norm_event(r["event"]), []).append(r)

    wb = Workbook()

    # ── Sheet 1: Automation FLOW_DEFS ──
    ws1 = wb.active
    ws1.title = "Automation FLOW_DEFS"
    ws1.append(
        ["operation", "touchpoint", "steps", "step_count", "scenario_id", "in_ui_sheet"]
    )
    for cell in ws1[1]:
        cell.font = Font(bold=True)
    for sc in scenarios:
        op = sc["operation"]
        norm = _norm_event(op)
        ws1.append(
            [
                op,
                sc["touchpoint"],
                " → ".join(sc["steps"]),
                len(sc["steps"]),
                sc["id"],
                "yes" if norm in sheet_by_norm else "no",
            ]
        )

    # ── Sheet 2: UI Navigation events ──
    ws2 = wb.create_sheet("UI Navigation Events")
    ws2.append(
        [
            "source_sheet",
            "section",
            "event",
            "touchpoint",
            "navigation_paths",
            "web_app",
            "has_curl",
            "automation_operation",
            "automation_touchpoints",
            "mapped",
        ]
    )
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    for r in sheet_rows:
        event = r.get("event") or ""
        norm = _norm_event(event)
        auto_op = flow_ops.get(norm) or flow_ops.get(norm.replace("update", "update")) 
        # fuzzy: match FLOW key where norm equals or contains
        if not auto_op:
            for k, v in flow_ops.items():
                if k == norm or norm in k or k in norm:
                    auto_op = v
                    break
        touches = list((FLOW_DEFS.get(auto_op) or {}).keys()) if auto_op else []
        ws2.append(
            [
                r.get("source_sheet", ""),
                r.get("section", ""),
                event,
                r.get("touchpoint", ""),
                r.get("navigation_paths", ""),
                r.get("web_app", ""),
                "yes" if r.get("has_curl") else "",
                auto_op or "",
                " | ".join(touches),
                "yes" if auto_op else "no",
            ]
        )

    # ── Sheet 3: Gaps ──
    ws3 = wb.create_sheet("Gaps")
    ws3.append(["gap_type", "name", "detail"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
    auto_norms = {_norm_event(op) for op in FLOW_DEFS}
    for norm, rows in sheet_by_norm.items():
        if norm not in auto_norms and not any(
            norm == a or norm in a or a in norm for a in auto_norms
        ):
            ws3.append(
                [
                    "sheet_event_without_FLOW_DEFS",
                    rows[0].get("event", ""),
                    f"{len(rows)} row(s) in UI Navigation workbook",
                ]
            )
    for op in sorted(FLOW_DEFS):
        if _norm_event(op) not in sheet_by_norm:
            ws3.append(
                [
                    "FLOW_DEFS_not_in_sheet",
                    op,
                    f"{len(FLOW_DEFS[op])} touchpoint(s)",
                ]
            )

    # ── Sheet 4: Full generate catalog ──
    ws4 = wb.create_sheet("Generate Catalog")
    ws4.append(["id", "label", "kind", "operation", "touchpoint", "steps"])
    for cell in ws4[1]:
        cell.font = Font(bold=True)
    for c in catalog:
        steps = c.get("steps") or []
        ws4.append(
            [
                c.get("id", ""),
                c.get("label", ""),
                c.get("kind", ""),
                c.get("operation", ""),
                c.get("touchpoint", ""),
                " → ".join(steps) if isinstance(steps, list) else str(steps or ""),
            ]
        )

    # ── Sheet 5: activateFamily detail (example of multi-touch) ──
    ws5 = wb.create_sheet("activateFamily paths")
    ws5.append(["touchpoint", "dependency_steps", "notes"])
    for cell in ws5[1]:
        cell.font = Font(bold=True)
    for touch, steps in (FLOW_DEFS.get("activateFamily") or {}).items():
        note = ""
        if "createProject" in steps and "createAsset" in steps:
            note = "Create project → create list → add family → activate"
        elif "createAsset" in steps:
            note = "Create list → add family → activate"
        elif "createProject" in steps:
            note = "Create project → add family → activate"
        elif steps == ["activateFamily"]:
            note = "Discovery global — activate only (no asset create)"
        ws5.append([touch, " → ".join(steps), note])

    # ── Sheet 6: summary ──
    ws6 = wb.create_sheet("Summary")
    ws6.append(["metric", "value"])
    for cell in ws6[1]:
        cell.font = Font(bold=True)
    ws6.append(["ui_navigation_xlsx", str(xlsx) if xlsx.is_file() else "missing"])
    ws6.append(["sheet_event_rows", len(sheet_rows)])
    ws6.append(["unique_sheet_events", len(sheet_by_norm)])
    ws6.append(["FLOW_DEFS_operations", len(FLOW_DEFS)])
    ws6.append(["automation_scenarios", len(scenarios)])
    ws6.append(["generate_catalog_items", len(catalog)])

    # silence unused
    _ = sheet_events
    return wb


def write_ui_navigation_mapping(
    *,
    project_root: Path | None = None,
    out_path: Path | None = None,
) -> Path:
    root = project_root or _repo_root()
    dest = out_path or (root / "reports" / "ui-navigation-event-mapping.xlsx")
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = build_ui_navigation_mapping_workbook(project_root=root)
    wb.save(dest)
    return dest
