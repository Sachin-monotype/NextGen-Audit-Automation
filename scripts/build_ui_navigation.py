#!/usr/bin/env python3
"""Regenerate ``ui_navigation.json`` from ``docs/UI Navigation of Event.xlsx``.

The workbook has section-header rows (only the first column filled) followed by
operation rows with one or more ``UI_Navigation*`` columns and an optional
``Remarks`` column. Every sheet is merged into a single
``{operation: {section, navigation: [...], remarks}}`` map.

Usage:  python scripts/build_ui_navigation.py
"""

from __future__ import annotations

import json
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX = ROOT / "docs" / "UI Navigation of Event.xlsx"
OUT = ROOT / "python" / "audit_validator" / "data" / "ui_navigation.json"


def _clean(value: object) -> str:
    return str(value).strip() if value is not None else ""


def build() -> dict[str, dict]:
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    result: dict[str, dict] = {}

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [_clean(h).lower() for h in rows[0]]
        nav_idx = [i for i, h in enumerate(header) if h.startswith("ui_navigation")]
        remarks_idx = next((i for i, h in enumerate(header) if h == "remarks"), None)

        current_section = ""
        for row in rows[1:]:
            section_cell = _clean(row[0]) if len(row) else ""
            op_cell = _clean(row[1]) if len(row) > 1 else ""
            if section_cell and not op_cell:
                current_section = section_cell
                continue
            if not op_cell:
                continue
            navigation = []
            for i in nav_idx:
                seg = _clean(row[i]) if len(row) > i else ""
                if seg and seg not in ("-", "—", "N/A", "NA"):
                    navigation.append(seg)
            remarks = (
                _clean(row[remarks_idx]) if remarks_idx is not None and len(row) > remarks_idx else ""
            )
            result[op_cell] = {
                "section": current_section,
                "navigation": navigation,
                "remarks": remarks,
            }
    return dict(sorted(result.items()))


def main() -> None:
    data = build()
    OUT.write_text(json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    with_nav = sum(1 for v in data.values() if v["navigation"])
    print(f"Wrote {len(data)} operation(s) ({with_nav} with navigation) → {OUT}")


if __name__ == "__main__":
    main()
