#!/usr/bin/env python3
"""Build module-split TouchPoint Excel + Postman under docs/touchpoint/.

Modules: activation, library, projects, favourites, imported_fonts, documents,
tags, teams_orgs, notifications, sharing, other, ingress, cron (Excel only).

Target: 160+ events with schema-correct variables (not _hint stubs).
"""

from __future__ import annotations

import json
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env")
except Exception:
    pass

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit("openpyxl required") from exc

import requests

from audit_validator.touchpoint.modules import MODULES, module_for
from audit_validator.touchpoint.payloads import FLOW_DEFS, SeedIds, variables_for
from audit_validator.utility.operation_graphql import (
    get_document_for_operation,
    is_mutation_operation,
)

OUT_ROOT = ROOT / "docs" / "touchpoint"
INGRESS_DIR = ROOT / "python" / "audit_validator" / "data" / "ingress_payloads"
CRON_DIR = ROOT / "python" / "audit_validator" / "data" / "cron_payloads"


def _bearer() -> str:
    raw = (os.getenv("BEARER_TOKEN") or "").strip()
    return raw[7:].strip() if raw.lower().startswith("bearer ") else raw


def _endpoint() -> str:
    return (
        os.getenv("NEXTGEN_GRAPHQL_ENDPOINT")
        or os.getenv("GRAPHQL_ENDPOINT")
        or "https://nextgen.monotype-pp.com/graph"
    ).rstrip("/")


def _jwt_gcid(token: str) -> str:
    try:
        import base64

        payload = token.split(".")[1]
        payload += "=" * (-len(payload) % 4)
        claims = json.loads(base64.urlsafe_b64decode(payload))
        info = claims.get("https://secure.monotype.com/info") or {}
        return str(
            info.get("parentCustomerId")
            or claims.get("https://api.monotype.com/gcid")
            or os.getenv("OAUTH_GCID")
            or ""
        )
    except Exception:
        return os.getenv("OAUTH_GCID") or os.getenv("INGRESS_DEFAULT_GCID") or ""


def resolve_seed() -> SeedIds:
    family = os.getenv("TOUCHPOINT_FAMILY_ID", "").strip() or "910130168"
    style = (
        os.getenv("TOUCHPOINT_STYLE_ID", "").strip()
        or os.getenv("SEED_STYLE_ID", "").strip()
        or "920374778"
    )
    md5 = os.getenv("SEED_VARIATION_MD5", "").strip() or "b783215634650cf0a55e0d723123d5e0"
    live = ROOT / "reports" / "touchpoint-live-seed.json"
    list_id = project_id = ""
    if live.exists():
        try:
            d = json.loads(live.read_text())
            list_id = d.get("list_id") or ""
            project_id = d.get("project_id") or ""
            family = d.get("family_id") or family
            style = d.get("style_id") or style
        except Exception:
            pass
    ts = int(time.time())
    token = _bearer()
    return SeedIds(
        family_id=family,
        style_id=style,
        md5=md5,
        list_id=list_id or os.getenv("TOUCHPOINT_LIST_ID", "").strip(),
        project_id=project_id or os.getenv("TOUCHPOINT_PROJECT_ID", "").strip(),
        list_name=f"QA_TP_List_{ts}",
        project_name=f"QA_TP_Project_{ts}",
        customer_id=_jwt_gcid(token),
        tag_id=os.getenv("TOUCHPOINT_TAG_ID", "").strip() or "QA_TAG_ID",
        team_id=os.getenv("TOUCHPOINT_TEAM_ID", "").strip() or "QA_TEAM_ID",
        role_id=os.getenv("TOUCHPOINT_ROLE_ID", "").strip() or "QA_ROLE_ID",
        profile_id=os.getenv("TOUCHPOINT_PROFILE_ID", "").strip() or "QA_PROFILE_ID",
        notification_id=os.getenv("TOUCHPOINT_NOTIFICATION_ID", "").strip()
        or "QA_NOTIFICATION_ID",
        contract_id=os.getenv("TOUCHPOINT_CONTRACT_ID", "").strip() or "QA_CONTRACT_ID",
        batch_id=os.getenv("TOUCHPOINT_BATCH_ID", "").strip() or "QA_BATCH_ID",
        sharee_id=os.getenv("TOUCHPOINT_SHAREE_ID", "").strip() or "",
        session_id=os.getenv("TOUCHPOINT_SESSION_ID", "").strip() or "QA_SESSION_ID",
        file_id=os.getenv("TOUCHPOINT_FILE_ID", "").strip() or "",
        document_id=os.getenv("TOUCHPOINT_DOCUMENT_ID", "").strip() or "QA_DOCUMENT_ID",
        attachment_id=os.getenv("TOUCHPOINT_ATTACHMENT_ID", "").strip() or "QA_ATTACHMENT_ID",
    )


def _op_name(doc: str, operation: str) -> str:
    m = re.search(r"(?:mutation|query)\s+(\w+)", doc or "", re.I)
    return m.group(1) if m else operation[:1].upper() + operation[1:]


def curl_for(operation: str, variables: dict, seed: SeedIds) -> str:
    doc = get_document_for_operation(operation)
    if not doc:
        return f"# No GraphQL document for {operation}\n# variables:\n# {json.dumps(variables, indent=2)}"
    payload = {
        "operationName": _op_name(doc, operation),
        "variables": variables,
        "extensions": {"clientLibrary": {"name": "@apollo/client", "version": "4.0.9"}},
        "query": doc,
    }
    body = json.dumps(payload, separators=(",", ":"))
    token = _bearer() or "{{BEARER_TOKEN}}"
    return (
        f"curl '{_endpoint()}' \\\n"
        f"  -H 'accept: application/graphql-response+json,application/json;q=0.9' \\\n"
        f"  -H 'authorization: Bearer {token}' \\\n"
        f"  -H 'content-type: application/json' \\\n"
        f"  -H 'origin: https://nextgen.monotype-pp.com' \\\n"
        f"  -X POST \\\n"
        f"  --data-raw '{body}'"
    )


def is_good_vars(variables: dict) -> bool:
    if not isinstance(variables, dict):
        return False
    blob = json.dumps(variables)
    if "_hint" in blob or "_unsupported" in blob:
        return False
    # empty vars OK for no-arg mutations
    if variables == {}:
        return True
    # must have some real field
    return any(k for k in variables.keys() if not k.startswith("_"))


def collect_graphql_ops(seed: SeedIds) -> dict[str, list[dict[str, Any]]]:
    """module → rows of {operation, touch, variables, steps, curl_cells}."""
    by_mod: dict[str, list[dict[str, Any]]] = {m: [] for m in MODULES if m not in ("ingress", "cron")}
    seen: set[tuple[str, str]] = set()

    # 1) Multi-touch FLOW_DEFS
    for op, touches in FLOW_DEFS.items():
        for touch, steps in touches.items():
            key = (op, touch)
            if key in seen:
                continue
            seen.add(key)
            vars_ = variables_for(op, seed, touch=touch)
            if not is_good_vars(vars_):
                continue
            step_curls = []
            for step_op in steps:
                sv = variables_for(step_op, seed, touch=touch)
                step_curls.append((step_op, curl_for(step_op, sv, seed), sv))
            by_mod.setdefault(module_for(op), []).append(
                {
                    "operation": op,
                    "touch": touch,
                    "variables": vars_,
                    "steps": steps,
                    "step_curls": step_curls,
                    "kind": "graphql",
                }
            )

    # 2) All other mutations with a document + variables_for builder
    from audit_validator.utility import operation_graphql as og

    # Prefer index of known ops
    try:
        catalog = list(getattr(og, "OPERATION_GRAPHQL_INDEX", {}) or {})
    except Exception:
        catalog = []
    if not catalog:
        # fall back: inventory hint+good
        inv = ROOT / "temp" / "inventory-report.json"
        if inv.exists():
            d = json.loads(inv.read_text())
            catalog = list(d.get("touchpoint_schema_good") or []) + list(
                d.get("touchpoint_hint_only") or []
            )

    for op in sorted(set(catalog)):
        if not get_document_for_operation(op):
            continue
        if not is_mutation_operation(op) and op not in {"sharingInfoForAssets"}:
            # skip pure queries except sharingInfo (listed in matrix)
            if op.startswith("get") or op.endswith("Viewed"):
                continue
        touch = "Discovery/Browse (global)"
        if (op, touch) in seen:
            continue
        # skip if already covered under any touch
        if any(r["operation"] == op for rows in by_mod.values() for r in rows):
            continue
        vars_ = variables_for(op, seed, touch=touch)
        if not is_good_vars(vars_):
            continue
        seen.add((op, touch))
        by_mod.setdefault(module_for(op), []).append(
            {
                "operation": op,
                "touch": touch,
                "variables": vars_,
                "steps": [op],
                "step_curls": [(op, curl_for(op, vars_, seed), vars_)],
                "kind": "graphql",
            }
        )
    return by_mod


def write_module_xlsx(module: str, rows: list[dict], seed: SeedIds) -> Path:
    folder = OUT_ROOT / module
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{module}_TouchPoint.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Touch Points"
    headers = [
        "Event",
        "Entity/TouchPoint",
        "Platform",
        "Step 1",
        "Step 2",
        "Step 3",
        "Step 4",
        "Step 5",
        "unique_input_json",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E2F3")

    for row in rows:
        cells = [row["operation"], row["touch"], "Web"]
        curls = row.get("step_curls") or []
        for i in range(5):
            if i < len(curls):
                op_name, curl, _ = curls[i]
                title = {
                    "createAsset": "Create List :",
                    "createProject": "Create Project :",
                    "addFontListFamilies": "Add family to list :",
                    "addFontProjectFamilies": "Add family to project :",
                    "addFavoriteFamilies": "Add family to favourites :",
                    "activateFamily": "Activate family :",
                }.get(op_name, f"{op_name} :")
                cells.append(f"{title}\n{curl}")
            else:
                cells.append("")
        cells.append(json.dumps(row["variables"], indent=2))
        ws.append(cells)

    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 24
    for col in range(4, 9):
        ws.column_dimensions[get_column_letter(col)].width = 48
    ws.column_dimensions["I"].width = 40
    ws.freeze_panes = "A2"

    # Legend
    leg = wb.create_sheet("Legend", 0)
    leg.append(["Module", MODULES.get(module, module)])
    leg.append(["Seed familyId", seed.family_id])
    leg.append(["Seed styleId", seed.style_id])
    leg.append(["Seed listId", seed.list_id])
    leg.append(["Seed projectId", seed.project_id])
    leg.append(["Seed customerId", seed.customer_id])
    leg.append(["Rows", len(rows)])
    leg.append(
        [
            "Note",
            "Curls use .env BEARER_TOKEN. Multi-step: run Create before Add/Activate. "
            "Verify mutation success==true and RabbitMQ raw landing via Generate verify.",
        ]
    )

    wb.save(path)
    return path


def write_module_postman(module: str, rows: list[dict], seed: SeedIds) -> Path:
    folder = OUT_ROOT / module
    folder.mkdir(parents=True, exist_ok=True)
    path = folder / f"{module}.postman_collection.json"

    def walk_sub(obj: Any) -> Any:
        if isinstance(obj, dict):
            return {k: walk_sub(v) for k, v in obj.items()}
        if isinstance(obj, list):
            return [walk_sub(x) for x in obj]
        if isinstance(obj, str):
            if obj == seed.list_id and seed.list_id:
                return "{{listId}}"
            if obj == seed.project_id and seed.project_id:
                return "{{projectId}}"
            if obj == f"project_{seed.project_id}" and seed.project_id:
                return "project_{{projectId}}"
            if obj == seed.family_id:
                return "{{familyId}}"
            if obj == seed.style_id:
                return "{{styleId}}"
        return obj

    folders = []
    for row in rows:
        items = []
        for idx, (op, _curl, vars_) in enumerate(row.get("step_curls") or [], 1):
            doc = get_document_for_operation(op) or ""
            body_vars = walk_sub(json.loads(json.dumps(vars_)))
            # Force chaining vars for known keys
            inp = body_vars.get("input") if isinstance(body_vars, dict) else None
            if isinstance(inp, dict):
                if "fontListId" in inp:
                    inp["fontListId"] = "{{listId}}"
                if "fontProjectId" in inp:
                    inp["fontProjectId"] = "{{projectId}}"
                if op == "activateFamily" and inp.get("listType") == "FONTLIST" and "listIds" in inp:
                    inp["listIds"] = ["{{listId}}"]
                if op == "activateFamily" and inp.get("listType") == "FONTPROJECT":
                    inp["listIds"] = ["project_{{projectId}}"]
                    inp["projectId"] = "{{projectId}}"
                if "families" in inp and isinstance(inp["families"], dict):
                    inp["families"]["familyIds"] = ["{{familyId}}"]
                if "familyIds" in inp:
                    inp["familyIds"] = ["{{familyId}}"]
            payload = {
                "operationName": _op_name(doc, op),
                "variables": body_vars,
                "extensions": {"clientLibrary": {"name": "@apollo/client", "version": "4.0.9"}},
                "query": doc,
            }
            events = []
            if op == "createAsset":
                events.append(
                    {
                        "listen": "test",
                        "script": {
                            "type": "text/javascript",
                            "exec": [
                                "const j=pm.response.json();",
                                "const id=j?.data?.createAsset?.asset?.id;",
                                "if(id){pm.collectionVariables.set('listId',id);}",
                                "const ok=j?.data?.createAsset?.success;",
                                "pm.test('createAsset success',()=>pm.expect(ok).to.eql(true));",
                            ],
                        },
                    }
                )
            if op == "createProject":
                events.append(
                    {
                        "listen": "test",
                        "script": {
                            "type": "text/javascript",
                            "exec": [
                                "const j=pm.response.json();",
                                "const id=j?.data?.createProject?.asset?.id;",
                                "if(id){pm.collectionVariables.set('projectId',id);}",
                                "pm.test('createProject success',()=>pm.expect(j?.data?.createProject?.success).to.eql(true));",
                            ],
                        },
                    }
                )
            if op not in {"createAsset", "createProject"}:
                events.append(
                    {
                        "listen": "test",
                        "script": {
                            "type": "text/javascript",
                            "exec": [
                                "const j=pm.response.json();",
                                "pm.test('no BAD_USER_INPUT',()=>{",
                                "  const errs=j.errors||[];",
                                "  pm.expect(errs.some(e=>String((e.extensions||{}).code||'')==='BAD_USER_INPUT')).to.eql(false);",
                                "});",
                                f"const node=j?.data?.{op};",
                                "if(node && Object.prototype.hasOwnProperty.call(node,'success')){",
                                "  console.log('success=', node.success, 'errors=', node.errors);",
                                "}",
                            ],
                        },
                    }
                )
            items.append(
                {
                    "name": f"{idx}. {op}",
                    "event": events,
                    "request": {
                        "method": "POST",
                        "header": [
                            {"key": "Authorization", "value": "Bearer {{bearerToken}}"},
                            {"key": "Content-Type", "value": "application/json"},
                            {
                                "key": "Accept",
                                "value": "application/graphql-response+json,application/json;q=0.9",
                            },
                        ],
                        "body": {
                            "mode": "raw",
                            "raw": json.dumps(payload, indent=2),
                            "options": {"raw": {"language": "json"}},
                        },
                        "url": "{{graphqlUrl}}",
                    },
                }
            )
        folders.append({"name": f"{row['operation']} / {row['touch']}", "item": items})

    collection = {
        "info": {
            "name": f"NextGen TouchPoint — {MODULES.get(module, module)}",
            "description": f"Module `{module}`. Run folder requests in order. Assert success where available.",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
        },
        "variable": [
            {"key": "graphqlUrl", "value": _endpoint()},
            {"key": "bearerToken", "value": _bearer()},
            {"key": "familyId", "value": seed.family_id},
            {"key": "styleId", "value": seed.style_id},
            {"key": "listId", "value": seed.list_id},
            {"key": "projectId", "value": seed.project_id},
            {"key": "customerId", "value": seed.customer_id},
        ],
        "item": folders,
    }
    path.write_text(json.dumps(collection, indent=2) + "\n", encoding="utf-8")
    return path


def write_ingress() -> tuple[Path, Path]:
    folder = OUT_ROOT / "ingress"
    folder.mkdir(parents=True, exist_ok=True)
    manifest = json.loads((INGRESS_DIR / "manifest.json").read_text())
    cases = manifest.get("cases") or []
    xlsx = folder / "ingress_TouchPoint.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Touch Points"
    ws.append(
        [
            "Event",
            "case_id",
            "Platform",
            "Trigger",
            "Payload JSON",
            "curl",
        ]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)
    postman_items = []
    ingress_url = os.getenv(
        "INGRESS_API_URL",
        "https://mt-audit-log-resolver-service-preprod.monotype-pp.com/v1/audit-events",
    )
    for case in cases:
        if case.get("skipped"):
            continue
        op = case.get("operation") or ""
        fid = case.get("file") or ""
        payload_path = INGRESS_DIR / fid
        payload = json.loads(payload_path.read_text()) if payload_path.exists() else {}
        curl_file = INGRESS_DIR / (case.get("curl_file") or "")
        curl_txt = curl_file.read_text() if curl_file.exists() else ""
        ws.append(
            [
                op,
                case.get("case_id"),
                "App/Plugin",
                "Ingress POST",
                json.dumps(payload, indent=2)[:32000],
                curl_txt[:32000],
            ]
        )
        postman_items.append(
            {
                "name": f"{case.get('case_id')} ({op})",
                "request": {
                    "method": "POST",
                    "header": [
                        {"key": "Authorization", "value": "Bearer {{bearerToken}}"},
                        {"key": "Content-Type", "value": "application/json"},
                        {"key": "x-request-source", "value": "MT_CONNECT_BS"},
                        {"key": "x-machine-id", "value": "{{machineId}}"},
                        {"key": "x-unique-id", "value": "{{uniqueId}}"},
                    ],
                    "body": {
                        "mode": "raw",
                        "raw": json.dumps(payload if isinstance(payload, list) else [payload], indent=2),
                        "options": {"raw": {"language": "json"}},
                    },
                    "url": "{{ingressUrl}}",
                },
            }
        )
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(xlsx)
    collection = {
        "info": {
            "name": "NextGen TouchPoint — Ingress APIs",
            "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json",
        },
        "variable": [
            {"key": "ingressUrl", "value": ingress_url},
            {"key": "bearerToken", "value": _bearer()},
            {
                "key": "machineId",
                "value": os.getenv("INGRESS_MACHINE_ID", "qa-touchpoint-machine"),
            },
            {
                "key": "uniqueId",
                "value": os.getenv("INGRESS_UNIQUE_ID", "qa-touchpoint-unique"),
            },
        ],
        "item": postman_items,
    }
    pjson = folder / "ingress.postman_collection.json"
    pjson.write_text(json.dumps(collection, indent=2) + "\n")
    return xlsx, pjson


def write_cron() -> Path:
    folder = OUT_ROOT / "cron"
    folder.mkdir(parents=True, exist_ok=True)
    xlsx = folder / "cron_TouchPoint.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Touch Points"
    ws.append(
        [
            "Event",
            "case_file",
            "Platform",
            "Trigger",
            "routingKey",
            "Payload JSON",
        ]
    )
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for path in sorted(CRON_DIR.glob("*.json")):
        payload = json.loads(path.read_text())
        op = (
            (payload.get("source") or {}).get("operation")
            or path.stem
        )
        rk = payload.get("routingKey") or ""
        ws.append(
            [
                op,
                path.name,
                "Scheduler",
                "Cron / RabbitMQ publish",
                rk,
                json.dumps(payload, indent=2)[:32000],
            ]
        )
    for r in ws.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    leg = wb.create_sheet("Legend", 0)
    leg.append(["Module", "Cron / Scheduler"])
    leg.append(["Note", "No Postman — publish via audit_validator.cron.runner / Generate cron path"])
    leg.append(["Queue", os.getenv("RAW_EVENTS_QUEUE", "mt.platform,resolver.raw_events_test_queue")])
    wb.save(xlsx)
    return xlsx


def write_index(summary: dict) -> Path:
    path = OUT_ROOT / "README.md"
    lines = [
        "# TouchPoint packs (module-split)",
        "",
        "Generated by `scripts/build_touchpoint_modules.py`.",
        "",
        "| Module | Events (rows) | Excel | Postman |",
        "|--------|--------------:|-------|---------|",
    ]
    for mod, info in summary.items():
        excel = info.get("excel", "")
        postman = info.get("postman", "—")
        lines.append(
            f"| `{mod}` — {MODULES.get(mod, mod)} | {info.get('rows', 0)} | "
            f"`{excel}` | `{postman}` |"
        )
    lines += [
        "",
        f"**Total covered event rows:** {sum(i.get('rows', 0) for i in summary.values())}",
        "",
        "## Verify",
        "",
        "```bash",
        "PYTHONPATH=python backend/.venv/bin/python scripts/verify_touchpoint_live.py",
        "```",
        "",
        "Checks: GraphQL `success: true` (where field exists), no `BAD_USER_INPUT`,",
        "and RabbitMQ/Mongo raw landing for owned `xCorrelationId` when ingestion is up.",
        "",
    ]
    path.write_text("\n".join(lines) + "\n")
    return path


def main() -> int:
    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    seed = resolve_seed()
    print("seed", seed)
    by_mod = collect_graphql_ops(seed)
    summary: dict[str, dict] = {}
    total_ops: set[str] = set()
    for mod, rows in sorted(by_mod.items()):
        if not rows:
            continue
        xlsx = write_module_xlsx(mod, rows, seed)
        pman = write_module_postman(mod, rows, seed)
        ops = {r["operation"] for r in rows}
        total_ops |= ops
        summary[mod] = {
            "rows": len(rows),
            "ops": len(ops),
            "excel": str(xlsx.relative_to(ROOT)),
            "postman": str(pman.relative_to(ROOT)),
        }
        print(f"  {mod}: {len(rows)} rows / {len(ops)} ops → {xlsx.name}")

    ix, ip = write_ingress()
    ingress_cases = json.loads((INGRESS_DIR / "manifest.json").read_text()).get("cases") or []
    ingress_ops = {c.get("operation") for c in ingress_cases if not c.get("skipped")}
    total_ops |= {o for o in ingress_ops if o}
    summary["ingress"] = {
        "rows": len([c for c in ingress_cases if not c.get("skipped")]),
        "ops": len(ingress_ops),
        "excel": str(ix.relative_to(ROOT)),
        "postman": str(ip.relative_to(ROOT)),
    }
    print(f"  ingress: {summary['ingress']['rows']} cases")

    cx = write_cron()
    cron_files = list(CRON_DIR.glob("*.json"))
    cron_ops = set()
    for p in cron_files:
        try:
            cron_ops.add((json.loads(p.read_text()).get("source") or {}).get("operation") or p.stem)
        except Exception:
            cron_ops.add(p.stem)
    total_ops |= cron_ops
    summary["cron"] = {
        "rows": len(cron_files),
        "ops": len(cron_ops),
        "excel": str(cx.relative_to(ROOT)),
        "postman": "—",
    }
    print(f"  cron: {len(cron_files)} payloads")

    write_index(summary)
    cov = {
        "generated_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
        "unique_operations": sorted(total_ops),
        "unique_operation_count": len(total_ops),
        "modules": summary,
    }
    (OUT_ROOT / "coverage.json").write_text(json.dumps(cov, indent=2) + "\n")
    print(f"UNIQUE OPS COVERED: {len(total_ops)}")
    if len(total_ops) < 160:
        print(f"WARNING: below 160 target ({len(total_ops)})")
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
