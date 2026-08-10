#!/usr/bin/env python3
"""Build App Event Sheet Status report across baseline 60 App Events (74 Scenarios).

Baseline App Events & Scenarios provided by user:
  - activateFamily: Favourite, List, global
  - activateList: List, Project, Project > List
  - activateStyle: Favourite, List, global
  - addFavoriteFamilies: Favourite, global
  - addFavoriteStyles: global
  - addFontListFamilies: List, Project > List
  - bulkActivateLists: List
  - bulkActivateStyles: Favourite, Project, global
  - bulkAddStylesToFavourites: global
  - bulkCopyAssets: global
  - bulkDeactivateLists: List
  - bulkDeactivateStyles: Favourite, global
  - bulkMoveAssets: global
  - bulkNotificationAction: global
  - bulkRemoveStylesFromFavourites: global
  - bulkTagStyles: global
  - bulkUntagStyles: global
  - bulkUpdatePreferences: global
  - createAsset: global
  - createCompanyLogoUploadUrl: global
  - createPrivateTags: manage
  - createProject: Project
  - createRole: manage
  - createTeam: manage
  - createUploadSession: Document
  - deActivateList: List, Project
  - deactivateFamilies: global
  - deactivateStyle: Favourite, global
  - deleteAllPrivateTags: global
  - deleteAssets: global
  - deletePrivateTags: manage
  - dismissNotification: global
  - exportActiveFonts: reporting
  - exportCompanyLibrary: company_library
  - exportFontUsers: manage
  - exportMyLibrary: My Library
  - exportNotifications: Notifications
  - exportRoles: manage
  - exportTags: manage
  - exportTeams: manage
  - exportUnassignedImportedFontsTemplate: global
  - exportUsers: manage
  - getActiveBatches: global
  - getImportedFonts: global
  - getStylesOfAllFontLists: global
  - markAllNotificationsRead: global
  - markCompanyLogoUploadSuccess: global
  - markNotificationRead: global
  - pinAsset: global
  - publishProject: Project
  - removeFavoriteFamilies: Favourite, global
  - removeFavoriteStyles: global
  - removeFontListFamilies: List
  - setLanguagePreference: Profile
  - sharingInfoForAssets: global
  - unpinAsset: global
  - updateAsset: global
  - updateAssetSharing: global
  - updateAssets: global
  - updatePrivateTag: manage

Status criteria:
  - covered: All baseline scenarios for the event are covered in App
  - partial cover: 1+ scenarios covered in App, but 1+ scenarios missing in App
  - not covered: 0 scenarios covered in App
    - Note: If present/covered in Web, specify 'Present in web (<scenarios>) but not in app'.
"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
REF_PATH = Path("/Users/sachinkoirala/Documents/CodeBases/QA /Auditdeviation/Coverage Status.xlsx")

TAG_RE = re.compile(r"\((BE|UI|app|APP|web|WEB)\)$", re.IGNORECASE)

USER_BASELINE_APP_SCENARIOS = {
    "activateFamily": ["Favourite", "List", "global"],
    "activateList": ["List", "Project", "Project > List"],
    "activateStyle": ["Favourite", "List", "global"],
    "addFavoriteFamilies": ["Favourite", "global"],
    "addFavoriteStyles": ["global"],
    "addFontListFamilies": ["List", "Project > List"],
    "bulkActivateLists": ["List"],
    "bulkActivateStyles": ["Favourite", "Project", "global"],
    "bulkAddStylesToFavourites": ["global"],
    "bulkCopyAssets": ["global"],
    "bulkDeactivateLists": ["List"],
    "bulkDeactivateStyles": ["Favourite", "global"],
    "bulkMoveAssets": ["global"],
    "bulkNotificationAction": ["global"],
    "bulkRemoveStylesFromFavourites": ["global"],
    "bulkTagStyles": ["global"],
    "bulkUntagStyles": ["global"],
    "bulkUpdatePreferences": ["global"],
    "createAsset": ["global"],
    "createCompanyLogoUploadUrl": ["global"],
    "createPrivateTags": ["manage"],
    "createProject": ["Project"],
    "createRole": ["manage"],
    "createTeam": ["manage"],
    "createUploadSession": ["Document"],
    "deActivateList": ["List", "Project"],
    "deactivateFamilies": ["global"],
    "deactivateStyle": ["Favourite", "global"],
    "deleteAllPrivateTags": ["global"],
    "deleteAssets": ["global"],
    "deletePrivateTags": ["manage"],
    "dismissNotification": ["global"],
    "exportActiveFonts": ["reporting"],
    "exportCompanyLibrary": ["company_library"],
    "exportFontUsers": ["manage"],
    "exportMyLibrary": ["My Library"],
    "exportNotifications": ["Notifications"],
    "exportRoles": ["manage"],
    "exportTags": ["manage"],
    "exportTeams": ["manage"],
    "exportUnassignedImportedFontsTemplate": ["global"],
    "exportUsers": ["manage"],
    "getActiveBatches": ["global"],
    "getImportedFonts": ["global"],
    "getStylesOfAllFontLists": ["global"],
    "markAllNotificationsRead": ["global"],
    "markCompanyLogoUploadSuccess": ["global"],
    "markNotificationRead": ["global"],
    "pinAsset": ["global"],
    "publishProject": ["Project"],
    "removeFavoriteFamilies": ["Favourite", "global"],
    "removeFavoriteStyles": ["global"],
    "removeFontListFamilies": ["List"],
    "setLanguagePreference": ["Profile"],
    "sharingInfoForAssets": ["global"],
    "unpinAsset": ["global"],
    "updateAsset": ["global"],
    "updateAssetSharing": ["global"],
    "updateAssets": ["global"],
    "updatePrivateTag": ["manage"],
}

ALIASES = {
    "exportuserfonts": "exportFontUsers",
    "exportuserprojects": "exportFontProjects",
    "exportuserassets": "exportFontAssets",
    "deactivatelist": "deActivateList",
    "bulkaddpairstofavorite": "bulkAddStylesToFavourites",
    "bulkremovepairsfromfavorite": "bulkRemoveStylesFromFavourites",
}

SCENARIO_MAP = {
    "favourite": "favourite",
    "favorites": "favourite",
    "fav": "favourite",
    "list": "list",
    "fontlist": "list",
    "project": "project",
    "project_list": "project_list",
    "project > list": "project_list",
    "global": "global",
    "default": "default",
    "app": "app",
    "preferences": "preferences",
    "font": "font",
    "login": "login",
    "document": "document",
    "pairing": "pairing",
    "roles": "roles",
    "teams": "teams",
    "manage": "manage",
    "reporting": "reporting",
    "user_access": "user_access",
    "my library": "my_library",
    "notifications": "notifications",
    "company_library": "company_library",
    "profile": "profile",
}


def norm_sc(raw: str) -> str:
    s = (raw or "").strip().lower()
    s = s.replace(">", "_").replace(" ", "_")
    while "__" in s:
        s = s.replace("__", "_")
    return SCENARIO_MAP.get(s, s or "default")


def parse_op(raw: str) -> tuple[str, str, str]:
    raw_str = (raw or "").strip()
    channel = ""
    m = TAG_RE.search(raw_str)
    if m:
        channel = m.group(1).lower()
        raw_str = raw_str[: m.start()].rstrip()
    parts = []
    while True:
        m2 = re.search(r"\(([^()]+)\)$", raw_str)
        if not m2:
            break
        parts.insert(0, m2.group(1))
        raw_str = raw_str[: m2.start()].rstrip()
    event = raw_str or raw
    sc = "/".join(parts) if parts else ("default" if not channel else "default")
    canonical = ALIASES.get(event.lower(), event)
    return canonical, norm_sc(sc), channel


def build_app_sheet_rows() -> list[dict]:
    # 1. Load comparison-latest-qa.json and Reference sheet
    qa_app_covered = defaultdict(set)
    qa_web_covered = defaultdict(set)

    qa_path = ROOT / "reports" / "comparison-latest-qa.json"
    if qa_path.exists():
        store = json.loads(qa_path.read_text(encoding="utf-8"))
        for op, block in store.items():
            if not isinstance(block, dict):
                continue
            ev, sc, ch = parse_op(op)
            b_summary = block.get("summary") or {}
            passed = int(b_summary.get("passed") or 0)
            if passed > 0:
                if ch == "app":
                    qa_app_covered[ev].add(sc)
                else:
                    qa_web_covered[ev].add(sc)

    if REF_PATH.exists():
        wb_ref = openpyxl.load_workbook(REF_PATH, data_only=True)
        if "App Event and Scenario" in wb_ref.sheetnames:
            ws_a = wb_ref["App Event and Scenario"]
            curr_ev = None
            for r in range(2, ws_a.max_row + 1):
                ev = ws_a.cell(r, 1).value
                sc = ws_a.cell(r, 2).value
                st = ws_a.cell(r, 3).value
                if ev:
                    curr_ev = str(ev).strip()
                if sc and curr_ev:
                    canonical = ALIASES.get(curr_ev.lower(), curr_ev)
                    sc_norm = norm_sc(str(sc))
                    st_clean = str(st or "").strip().lower()
                    if st_clean in {"covered", "partial"}:
                        qa_app_covered[canonical].add(sc_norm)

        if "Web Event and Scenario" in wb_ref.sheetnames:
            ws_w = wb_ref["Web Event and Scenario"]
            curr_ev = None
            for r in range(2, ws_w.max_row + 1):
                ev = ws_w.cell(r, 1).value
                sc = ws_w.cell(r, 2).value
                st = ws_w.cell(r, 3).value
                if ev:
                    curr_ev = str(ev).strip()
                if sc and curr_ev:
                    canonical = ALIASES.get(curr_ev.lower(), curr_ev)
                    sc_norm = norm_sc(str(sc))
                    st_clean = str(st or "").strip().lower()
                    if st_clean in {"covered", "partial"}:
                        qa_web_covered[canonical].add(sc_norm)

    rows: list[dict] = []
    for ev, req_scenarios in USER_BASELINE_APP_SCENARIOS.items():
        req_norm = [norm_sc(s) for s in req_scenarios]
        app_covered = qa_app_covered.get(ev, set())
        web_covered = qa_web_covered.get(ev, set())

        covered_scs = sorted(list(app_covered.intersection(req_norm)))
        missing_scs = sorted([s for s in req_scenarios if norm_sc(s) not in app_covered])

        if not covered_scs:
            status = "not covered"
            if web_covered:
                w_list = sorted(list(web_covered))
                note = f"Present in web ({', '.join(w_list)}) but not in app"
            else:
                note = "Not covered in web or app"
        elif missing_scs:
            status = "partial cover"
            note = f"Covered in app: {', '.join(covered_scs)}; Missing scenario(s) in app: {', '.join(missing_scs)}"
        else:
            status = "covered"
            note = f"Fully covered in app ({', '.join(req_scenarios)})"

        rows.append(
            {
                "event": ev,
                "category": "App Event Suite",
                "app_status": status,
                "app_covered_scenarios": ", ".join(covered_scs) if covered_scs else "None",
                "app_missing_scenarios": ", ".join(missing_scs) if missing_scs else "None",
                "web_covered_scenarios": ", ".join(sorted(list(web_covered))) if web_covered else "None",
                "note": note,
            }
        )

    return rows


def generate_excel_report(rows: list[dict]) -> Path:
    out_path = ROOT / "reports" / "app_event_sheet_status.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "App Event Status"

    ws.append(["App Event Coverage Sheet Status Report (Baseline App Suite)"])
    ws.append(
        [
            f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | "
            f"Total App Events: {len(rows)} | "
            f"Covered: {sum(1 for r in rows if r['app_status'] == 'covered')} | "
            f"Partial Cover: {sum(1 for r in rows if r['app_status'] == 'partial cover')} | "
            f"Not Covered: {sum(1 for r in rows if r['app_status'] == 'not covered')}"
        ]
    )
    ws.append([])

    headers = [
        "Event Name",
        "Category / Section",
        "App Status",
        "App Covered Scenarios",
        "App Missing Scenarios",
        "Web Covered Scenarios",
        "Note / Remarks",
    ]
    ws.append(headers)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    for cell in ws[4]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    fills = {
        "covered": PatternFill("solid", fgColor="DCFCE7"),
        "partial cover": PatternFill("solid", fgColor="FEF3C7"),
        "not covered": PatternFill("solid", fgColor="FEE2E2"),
    }
    status_fonts = {
        "covered": Font(bold=True, color="166534"),
        "partial cover": Font(bold=True, color="92400E"),
        "not covered": Font(bold=True, color="991B1B"),
    }

    for r in rows:
        ws.append(
            [
                r["event"],
                r["category"],
                r["app_status"],
                r["app_covered_scenarios"],
                r["app_missing_scenarios"],
                r["web_covered_scenarios"],
                r["note"],
            ]
        )
        row_idx = ws.max_row
        st = r["app_status"]
        for col in range(1, 8):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin
            if col == 3:
                cell.fill = fills[st]
                cell.font = status_fonts[st]
                cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A1:G1")
    ws.merge_cells("A2:G2")
    ws["A1"].font = Font(bold=True, size=14, color="1F2937")
    ws["A2"].font = Font(italic=True, color="4B5563")

    widths = [32, 24, 16, 28, 28, 28, 55]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.auto_filter.ref = f"A4:G{ws.max_row}"

    comb_path = ROOT / "reports" / "audit-coverage-combined.xlsx"
    if comb_path.exists():
        try:
            wb_comb = openpyxl.load_workbook(comb_path)
            if "App Event Status" in wb_comb.sheetnames:
                del wb_comb["App Event Status"]
            ws_c = wb_comb.create_sheet(title="App Event Status")
            ws_c.append(["App Event Coverage Sheet Status Report (Baseline App Suite)"])
            ws_c.append(
                [
                    f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')} | "
                    f"Total App Events: {len(rows)} | "
                    f"Covered: {sum(1 for r in rows if r['app_status'] == 'covered')} | "
                    f"Partial Cover: {sum(1 for r in rows if r['app_status'] == 'partial cover')} | "
                    f"Not Covered: {sum(1 for r in rows if r['app_status'] == 'not covered')}"
                ]
            )
            ws_c.append([])
            ws_c.append(headers)
            for cell in ws_c[4]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for r in rows:
                ws_c.append(
                    [
                        r["event"],
                        r["category"],
                        r["app_status"],
                        r["app_covered_scenarios"],
                        r["app_missing_scenarios"],
                        r["web_covered_scenarios"],
                        r["note"],
                    ]
                )
                row_idx = ws_c.max_row
                st = r["app_status"]
                for col in range(1, 8):
                    cell = ws_c.cell(row=row_idx, column=col)
                    cell.border = thin
                    if col == 3:
                        cell.fill = fills[st]
                        cell.font = status_fonts[st]
                        cell.alignment = Alignment(horizontal="center")

            ws_c.merge_cells("A1:G1")
            ws_c.merge_cells("A2:G2")
            ws_c["A1"].font = Font(bold=True, size=14, color="1F2937")
            ws_c["A2"].font = Font(italic=True, color="4B5563")

            for i, w in enumerate(widths, 1):
                ws_c.column_dimensions[get_column_letter(i)].width = w

            ws_c.auto_filter.ref = f"A4:G{ws_c.max_row}"
            wb_comb.save(comb_path)
            print(f"Updated {comb_path} with 'App Event Status' tab.")
        except Exception as e:
            print(f"Error updating audit-coverage-combined.xlsx: {e}")

    wb.save(out_path)
    print(f"Wrote {out_path}")
    return out_path


def main() -> None:
    rows = build_app_sheet_rows()
    out_path = generate_excel_report(rows)
    counts = Counter(r["app_status"] for r in rows)

    print("\n========================================================")
    print("      APP EVENT COVERAGE SHEET STATUS REPORT SUMMARY    ")
    print("========================================================")
    print(f"Total App Events Evaluated: {len(rows)}")
    print(f"  🟢 covered       : {counts.get('covered', 0)}")
    print(f"  🟡 partial cover : {counts.get('partial cover', 0)}")
    print(f"  🔴 not covered   : {counts.get('not covered', 0)}")
    print("========================================================\n")


if __name__ == "__main__":
    main()
