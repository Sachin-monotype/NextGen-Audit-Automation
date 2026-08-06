#!/usr/bin/env python3
"""Build QA Results coverage workbook (Web + App sheets) from comparison-latest-qa.json."""

from __future__ import annotations

import argparse
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
TAG_RE = re.compile(r"\((BE|UI|app|APP|web|WEB)\)$", re.IGNORECASE)


def split_event_scenario(operation: str) -> tuple[str, str, str]:
    """Return (event, scenario, channel) where channel is app|web|be|ui|''."""
    raw = (operation or "").strip()
    channel = ""
    m = TAG_RE.search(raw)
    if m:
        channel = m.group(1).lower()
        raw = raw[: m.start()].rstrip()
    # Peel trailing (touchpoint) groups: op(a)(b) → event=op, scenario=a/b
    parts: list[str] = []
    while True:
        m2 = re.search(r"\(([^()]+)\)$", raw)
        if not m2:
            break
        parts.insert(0, m2.group(1))
        raw = raw[: m2.start()].rstrip()
    event = raw or operation
    scenario = "/".join(parts) if parts else ("—" if channel else "default")
    return event, scenario, channel


def outcome(summary: dict) -> str:
    failed = int(summary.get("failed") or 0)
    skipped = int(summary.get("skipped") or 0)
    if failed > 0:
        return "FAILED"
    if skipped > 0:
        return "PARTIAL"
    return "PASSED"


def build_rows(store: dict) -> tuple[list[dict], list[dict]]:
    web: list[dict] = []
    app: list[dict] = []
    for op, block in store.items():
        if not isinstance(block, dict):
            continue
        summary = block.get("summary") or {}
        if not summary and block.get("rows"):
            c = Counter(str(r.get("match_status") or "") for r in block["rows"])
            summary = {
                "passed": c.get("PASS", 0),
                "failed": c.get("FAIL", 0),
                "skipped": c.get("SKIP", 0),
                "na": c.get("N/A", 0),
            }
        event, scenario, channel = split_event_scenario(op)
        passed = int(summary.get("passed") or 0)
        failed = int(summary.get("failed") or 0)
        skipped = int(summary.get("skipped") or 0)
        na = int(summary.get("na") or 0)
        row = {
            "event": event,
            "scenario": scenario,
            "operation": op,
            "status": outcome(summary),
            "passed": passed,
            "failed": failed,
            "skipped": skipped,
            "na": na,
            "fields": passed + failed + skipped + na,
            "compared_at": str(block.get("compared_at") or ""),
            "job_id": str(block.get("job_id") or ""),
            "channel": channel,
        }
        if channel == "app":
            app.append(row)
        else:
            web.append(row)
    web.sort(key=lambda r: (r["event"].lower(), r["scenario"].lower(), r["operation"].lower()))
    app.sort(key=lambda r: (r["event"].lower(), r["scenario"].lower(), r["operation"].lower()))
    return web, app


def _write_sheet(wb: Workbook, title: str, label: str, rows: list[dict], source_name: str) -> None:
    ws = wb.create_sheet(title=title)
    events = {r["event"] for r in rows}
    counts = Counter(r["status"] for r in rows)
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    ws.append([f"QA Results coverage — {label}"])
    ws.append(
        [
            f"Generated {generated} · source: {source_name} · "
            f"{len(events)} events · {len(rows)} scenarios · "
            f"PASSED {counts.get('PASSED', 0)} · FAILED {counts.get('FAILED', 0)} · "
            f"PARTIAL {counts.get('PARTIAL', 0)}"
        ]
    )
    ws.append([])
    headers = [
        "Event",
        "Scenario",
        "Full operation",
        "Status",
        "PASS",
        "FAIL",
        "SKIP",
        "N/A",
        "Fields",
        "Compared at (UTC)",
        "Job id",
    ]
    ws.append(headers)
    fills = {
        "PASSED": PatternFill("solid", fgColor="DCFCE7"),
        "FAILED": PatternFill("solid", fgColor="FEE2E2"),
        "PARTIAL": PatternFill("solid", fgColor="FEF3C7"),
    }
    header_font = Font(bold=True)
    thin = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    for cell in ws[4]:
        cell.font = header_font
        cell.border = thin
    for r in rows:
        ws.append(
            [
                r["event"],
                r["scenario"],
                r["operation"],
                r["status"],
                r["passed"],
                r["failed"],
                r["skipped"],
                r["na"],
                r["fields"],
                r["compared_at"],
                r["job_id"],
            ]
        )
        fill = fills.get(r["status"])
        row_idx = ws.max_row
        for col in range(1, 12):
            cell = ws.cell(row=row_idx, column=col)
            cell.border = thin
            if fill and col == 4:
                cell.fill = fill
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=11)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].alignment = Alignment(wrap_text=True)
    widths = [22, 18, 42, 10, 8, 8, 8, 8, 10, 28, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--store",
        type=Path,
        default=ROOT / "reports" / "comparison-latest-qa.json",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Default: reports/qa-results-coverage-YYYYMMDD.xlsx",
    )
    args = parser.parse_args()
    store_path = args.store
    if not store_path.is_file():
        print(f"Missing store: {store_path}")
        return 1
    data = json.loads(store_path.read_text(encoding="utf-8"))
    web, app = build_rows(data)
    out = args.out or (
        ROOT / "reports" / f"qa-results-coverage-{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    )
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb, "Web", "Web", web, store_path.name)
    _write_sheet(wb, "App", "App", app, store_path.name)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote {out}")
    print(f"  Web: {len(web)} scenarios")
    print(f"  App: {len(app)} scenarios")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
