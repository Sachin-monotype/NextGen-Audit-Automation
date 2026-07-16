#!/usr/bin/env python3
"""Build the GQL touchpoint × unique-input sheet (TouchPoint.xlsx).

Sources
-------
- audit-event-context-matrix.xlsx (Matrix tab) → operation × global/project/document/list
- TouchPoint.xlsx examples → real activateFamily / activateList input shapes
- mtf-graphql-schema Activate*Input fields
- mtconnect-api list-scope routing (FONTLIST / FAVORITE / FONTPROJECT)

Output
------
docs/mappings/TouchPoint.xlsx with sheets:
  Legend | UniqueInputShapes | TouchPoints | AutomationFlows | Gaps
Also refreshes docs/mappings/event_trigger_sheet.csv and
python/audit_validator/data/trigger_sequences.json (input-aware).
"""

from __future__ import annotations

import csv
import json
import os
import sys
import time
from copy import deepcopy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT / "python"))

try:
    from dotenv import load_dotenv

    load_dotenv(ROOT / ".env")
except Exception:
    pass

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError as exc:  # pragma: no cover
    raise SystemExit("openpyxl required: pip install openpyxl") from exc

import requests  # noqa: E402

from audit_validator.operation_sources import operation_source_report  # noqa: E402
from audit_validator.utility.operation_graphql import (  # noqa: E402
    get_document_for_operation,
    is_mutation_operation,
)

sys.path.insert(0, str(ROOT / "scripts"))
from touchpoint_payloads import FLOW_DEFS, SeedIds, variables_for  # noqa: E402

MATRIX_XLSX = Path("/Users/sachinkoirala/Downloads/audit-event-context-matrix.xlsx")
TOUCHPOINT_SRC = Path("/Users/sachinkoirala/Downloads/TouchPoint.xlsx")
OUT_XLSX = ROOT / "docs" / "mappings" / "TouchPoint.xlsx"
OUT_CSV = ROOT / "docs" / "mappings" / "event_trigger_sheet.csv"
SEQ_OUT = ROOT / "python" / "audit_validator" / "data" / "trigger_sequences.json"
GRAPHQL_DOCS = ROOT / "python" / "audit_validator" / "data" / "graphql_documents.json"

# Seed values — updated by resolve_live_context() before rows are emitted.
PH: dict[str, str] = {
    "family_id": "{{FAMILY_ID}}",
    "style_id": "{{STYLE_ID}}",
    "list_id": "{{LIST_ID}}",
    "list_id_prefixed": "list_{{LIST_ID}}",
    "project_id": "{{PROJECT_ID}}",
    "project_list_id": "project_{{PROJECT_ID}}",
    "md5": "{{VARIATION_MD5}}",
    "session_id": "{{SESSION_ID}}",
    "file_id": "{{FILE_ID}}",
    "document_id": "{{DOCUMENT_ID}}",
    "endpoint": "https://nextgen.monotype-pp.com/graph",
    "token": "{{BEARER_TOKEN}}",
    "list_name": "QA_TouchPoint_List",
    "project_name": "QA_TouchPoint_Project",
}


@dataclass
class LiveContext:
    family_id: str
    style_id: str
    md5: str
    bearer: str
    endpoint: str
    list_id: str = ""
    project_id: str = ""
    list_name: str = ""
    project_name: str = ""
    notes: list[str] = field(default_factory=list)


def _bearer_from_env() -> str:
    for key in ("BEARER_TOKEN", "BEARER_TOKEN_PP", "NEXTGEN_BEARER_TOKEN", "INGRESS_BEARER_TOKEN"):
        raw = (os.getenv(key) or "").strip()
        if raw:
            return raw[7:].strip() if raw.lower().startswith("bearer ") else raw
    return ""


def _id_from_enrich(op: str, *fallback_keys: str) -> str:
    for folder in ("enrich", "raw"):
        path = ROOT / "payload" / folder / f"{op}.json"
        if not path.is_file():
            continue
        try:
            data = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        subject = data.get("subject") or {}
        ids = subject.get("id")
        if isinstance(ids, list) and ids:
            return str(ids[0])
        if isinstance(ids, str) and ids:
            return ids
        inp = ((subject.get("metadata") or {}).get("input") or {})
        for key in fallback_keys:
            val = inp.get(key)
            if isinstance(val, list) and val:
                return str(val[0])
            if isinstance(val, str) and val:
                return val
    return ""


def _gql(endpoint: str, bearer: str, payload: dict[str, Any]) -> dict[str, Any]:
    resp = requests.post(
        endpoint,
        headers={
            "Authorization": f"Bearer {bearer}",
            "Content-Type": "application/json",
            "Accept": "application/graphql-response+json,application/json;q=0.9",
            "Origin": "https://nextgen.monotype-pp.com",
            "Referer": "https://nextgen.monotype-pp.com/search",
        },
        json=payload,
        timeout=60,
    )
    resp.raise_for_status()
    body = resp.json()
    if body.get("errors"):
        raise RuntimeError(json.dumps(body["errors"])[:500])
    return body.get("data") or {}


def resolve_live_context(*, provision_assets: bool = True) -> LiveContext:
    """Real bearer + family/style from .env / enrich JSON; live list/project IDs via API."""
    # Prefer TOUCHPOINT_FAMILY_ID / live-seed / UI-proven 910130168. Avoid 794981
    # (Discovery timeout on this PP org for addFontListFamilies).
    live_seed_path = ROOT / "reports" / "touchpoint-live-seed.json"
    live_seed: dict[str, Any] = {}
    if live_seed_path.is_file():
        try:
            live_seed = json.loads(live_seed_path.read_text(encoding="utf-8"))
        except Exception:
            live_seed = {}
    family = (
        os.getenv("TOUCHPOINT_FAMILY_ID", "").strip()
        or str(live_seed.get("family_id") or "").strip()
        or "910130168"  # UI/Apollo-proven on PP (prefer over enrich seed)
        or _id_from_enrich("activateFamily", "familyIds")
    )
    style = (
        os.getenv("TOUCHPOINT_STYLE_ID", "").strip()
        or str(live_seed.get("style_id") or "").strip()
        or os.getenv("SEED_STYLE_ID", "").strip()
        or _id_from_enrich("activateStyle", "styleIds")
        or "920374778"
    )
    md5 = (
        os.getenv("SEED_VARIATION_MD5", "").strip()
        or _id_from_enrich("activateVariation", "md5s")
        or "b783215634650cf0a55e0d723123d5e0"
    )
    bearer = _bearer_from_env()
    endpoint = (
        os.getenv("NEXTGEN_GRAPHQL_ENDPOINT")
        or os.getenv("GRAPHQL_ENDPOINT")
        or "https://nextgen.monotype-pp.com/graph"
    ).rstrip("/")
    ts = int(time.time())
    list_name = f"QA_TouchPoint_List_{ts}"
    project_name = f"QA_TouchPoint_Project_{ts}"
    ctx = LiveContext(
        family_id=family,
        style_id=style,
        md5=md5,
        bearer=bearer or "{{BEARER_TOKEN}}",
        endpoint=endpoint,
        list_name=list_name,
        project_name=project_name,
    )
    if not bearer:
        ctx.notes.append("BEARER_TOKEN missing — curls use {{BEARER_TOKEN}} placeholder")
    if provision_assets and bearer:
        try:
            create_doc = get_document_for_operation("createAsset") or (
                "mutation CreateAsset($input: CreateAssetInput!) { createAsset(input: $input) { "
                "success errors { code message } asset { ... on FontList { id } ... on FontProject { id } } } }"
            )
            list_data = _gql(
                endpoint,
                bearer,
                {
                    "operationName": "CreateAsset",
                    "variables": {
                        "input": {
                            "name": list_name,
                            "assetType": "FontList",
                            "accessRight": "FullAccess",
                        }
                    },
                    "query": create_doc,
                },
            )
            asset = (list_data.get("createAsset") or {}).get("asset") or {}
            ctx.list_id = str(asset.get("id") or "")

            # FontProject must use createProject (createAsset rejects assetType FontProject)
            project_doc = get_document_for_operation("createProject") or (
                "mutation CreateProject($input: CreateProjectInput!) { "
                "createProject(input: $input) { success asset { id name } errors { message code } } }"
            )
            proj_data = _gql(
                endpoint,
                bearer,
                {
                    "operationName": "CreateProject",
                    "variables": {
                        "input": {
                            "name": project_name[:50],
                            "description": "QA TouchPoint sheet seed project",
                            "allowFontAdditionsByCollaborators": False,
                            "allowFontDownloadsByCollaborators": False,
                            "allowFontImportsByCollaborators": False,
                            "enableProjectLevelImportedFonts": False,
                            "autoActivateFontsForMembers": False,
                        }
                    },
                    "query": project_doc,
                },
            )
            passet = (proj_data.get("createProject") or {}).get("asset") or {}
            ctx.project_id = str(passet.get("id") or "")
            ctx.notes.append(f"Provisioned FontList={ctx.list_id} FontProject={ctx.project_id}")
        except Exception as exc:  # noqa: BLE001
            ctx.notes.append(f"Asset provision failed ({exc}) — list/project ids left blank")
    PH.update(
        {
            "family_id": ctx.family_id,
            "style_id": ctx.style_id,
            "md5": ctx.md5,
            "list_id": ctx.list_id or "{{LIST_ID}}",
            "list_id_prefixed": f"list_{ctx.list_id}" if ctx.list_id else "list_{{LIST_ID}}",
            "project_id": ctx.project_id or "{{PROJECT_ID}}",
            "project_list_id": f"project_{ctx.project_id}" if ctx.project_id else "project_{{PROJECT_ID}}",
            "endpoint": ctx.endpoint,
            "token": ctx.bearer,
            "list_name": ctx.list_name,
            "project_name": ctx.project_name,
        }
    )
    return ctx


def _fill_placeholders(value: Any) -> Any:
    if isinstance(value, str):
        out = value
        for k, v in {
            "{{FAMILY_ID}}": PH["family_id"],
            "{{STYLE_ID}}": PH["style_id"],
            "{{LIST_ID}}": PH["list_id"],
            "{{PROJECT_ID}}": PH["project_id"],
            "{{VARIATION_MD5}}": PH["md5"],
            "{{LIST_NAME}}": PH["list_name"],
            "{{PROJECT_NAME}}": PH["project_name"],
            "{{SESSION_ID}}": PH["session_id"],
            "{{FILE_ID}}": PH["file_id"],
            "{{DOCUMENT_ID}}": PH["document_id"],
            "{{GRAPHQL_ENDPOINT}}": PH["endpoint"],
            "{{BEARER_TOKEN}}": PH["token"],
        }.items():
            out = out.replace(k, v)
        return out
    if isinstance(value, list):
        return [_fill_placeholders(x) for x in value]
    if isinstance(value, dict):
        return {k: _fill_placeholders(v) for k, v in value.items()}
    return value


def step_title(operation: str, variables: dict[str, Any] | None = None) -> str:
    vars_ = variables or {}
    asset_type = str(vars_.get("assetType") or "")
    titles = {
        "createAsset": {
            "FontList": "Create List",
            "FontProject": "Create Project",
            "Folder": "Create Folder",
        }.get(asset_type, "Create Asset"),
        "createProject": "Create Project",
        "addFontListFamilies": "Add family to list",
        "addFontListStyles": "Add style to list",
        "addFontProjectFamilies": "Add family to project",
        "addFontProjectStyles": "Add style to project",
        "addFavoriteFamilies": "Add family to favourites",
        "addFavoriteStyles": "Add style to favourites",
        "activateFamily": "Activate family",
        "deactivateFamilies": "Deactivate family",
        "activateStyle": "Activate style",
        "deactivateStyle": "Deactivate style",
        "activateVariation": "Activate variation",
        "deactivateVariation": "Deactivate variation",
        "activateList": "Activate list",
        "deActivateList": "Deactivate list",
        "activateFontProject": "Activate project fonts",
        "deActivateFontProject": "Deactivate project fonts",
        "bulkActivateStyles": "Bulk activate styles",
        "bulkDeactivateStyles": "Bulk deactivate styles",
        "bulkActivateLists": "Bulk activate lists",
        "bulkDeactivateLists": "Bulk deactivate lists",
    }
    return titles.get(operation, operation)


def format_step_cell(title: str, curl: str) -> str:
    return f"{title} :\n\n{curl}"

# ---------------------------------------------------------------------------
# Curated unique input shapes for FontActivation (from schema + TouchPoint.xlsx)
# ---------------------------------------------------------------------------
FONT_SCOPED_FAMILY = {
    "Discovery/Browse (global)": {
        "input_shape": {
            "familyIds": [PH["family_id"]],
            "activationType": "PERMANENT",
        },
        "diff_keys": "familyIds (+ optional activationType)",
        "prereq": [],
        "final_op": "activateFamily",
    },
    "List (FONTLIST)": {
        "input_shape": {
            "familyIds": [PH["family_id"]],
            "listIds": [PH["list_id_prefixed"]],
            "listType": "FONTLIST",
            "activationType": "PERMANENT",
        },
        "diff_keys": "listIds + listType=FONTLIST",
        "prereq": [
            ("createAsset", {"assetType": "FontList", "name": "{{LIST_NAME}}", "accessRight": "FullAccess"}),
            ("addFontListFamilies", {"fontListId": PH["list_id"], "families": {"familyIds": [PH["family_id"]]}}),
        ],
        "final_op": "activateFamily",
    },
    "Favourite": {
        "input_shape": {
            "familyIds": [PH["family_id"]],
            "listType": "FAVORITE",
            "activationType": "PERMANENT",
        },
        "diff_keys": "listType=FAVORITE (no listIds)",
        "prereq": [
            ("addFavoriteFamilies", {"familyIds": [PH["family_id"]]}),
        ],
        "final_op": "activateFamily",
    },
    "Project": {
        "input_shape": {
            "familyIds": [PH["family_id"]],
            "listIds": [PH["project_list_id"]],
            "listType": "FONTPROJECT",
            "activationType": "PERMANENT",
            "projectId": PH["project_id"],
        },
        "diff_keys": "listIds=project_{id} + listType=FONTPROJECT + projectId",
        "prereq": [
            ("createProject", {
                    "name": "{{PROJECT_NAME}}",
                    "description": "QA TouchPoint sheet seed project",
                    "allowFontAdditionsByCollaborators": False,
                    "allowFontDownloadsByCollaborators": False,
                    "allowFontImportsByCollaborators": False,
                    "enableProjectLevelImportedFonts": False,
                    "autoActivateFontsForMembers": False,
                }),
            (
                "addFontProjectFamilies",
                {
                    "projectId": PH["project_id"],
                    "families": {"familyIds": [PH["family_id"]]},
                },
            ),
        ],
        "final_op": "activateFamily",
    },
    "Project > List": {
        "input_shape": {
            "familyIds": [PH["family_id"]],
            "listIds": [PH["list_id_prefixed"]],
            "listType": "FONTLIST",
            "activationType": "PERMANENT",
            "projectId": PH["project_id"],
        },
        "diff_keys": "listIds (FONTLIST) + projectId (seat/context)",
        # Matches TouchPoint.xlsx: project → add family → list in project → add family → activate
        "prereq": [
            ("createProject", {
                    "name": "{{PROJECT_NAME}}",
                    "description": "QA TouchPoint sheet seed project",
                    "allowFontAdditionsByCollaborators": False,
                    "allowFontDownloadsByCollaborators": False,
                    "allowFontImportsByCollaborators": False,
                    "enableProjectLevelImportedFonts": False,
                    "autoActivateFontsForMembers": False,
                }),
            (
                "addFontProjectFamilies",
                {
                    "projectId": PH["project_id"],
                    "families": {"familyIds": [PH["family_id"]]},
                },
            ),
            ("createAsset", {"assetType": "FontList", "name": "{{LIST_NAME}}", "accessRight": "FullAccess", "parentId": PH["project_id"]}),
            ("addFontListFamilies", {"fontListId": PH["list_id"], "families": {"familyIds": [PH["family_id"]]}}),
        ],
        "final_op": "activateFamily",
    },
}

# Mirror deactivateFamilies / activateStyle / deactivateStyle with same scopes
def _swap_family_to_style(shape: dict[str, Any]) -> dict[str, Any]:
    out = deepcopy(shape)
    inp = out["input_shape"]
    if "familyIds" in inp:
        inp["styleIds"] = [PH["style_id"]]
        del inp["familyIds"]
    out["diff_keys"] = out["diff_keys"].replace("familyIds", "styleIds")
    # rewrite prereq family ops → style where needed
    new_prereq = []
    for op, vars_ in out.get("prereq") or []:
        if op == "addFontListFamilies":
            new_prereq.append(
                (
                    "addFontListStyles",
                    {"fontListId": PH["list_id"], "styles": [{"styleId": PH["style_id"]}]},
                )
            )
        elif op == "addFavoriteFamilies":
            new_prereq.append(("addFavoriteStyles", {"styleIds": [PH["style_id"]]}))
        elif op == "addFontProjectFamilies":
            new_prereq.append(
                (
                    "addFontProjectStyles",
                    {
                        "projectId": PH["project_id"],
                        "styles": [{"styleId": PH["style_id"]}],
                    },
                )
            )
        else:
            new_prereq.append((op, vars_))
    out["prereq"] = new_prereq
    return out


def _to_deactivate(shape: dict[str, Any], *, family: bool) -> dict[str, Any]:
    out = deepcopy(shape)
    inp = out["input_shape"]
    if "activationType" in inp:
        del inp["activationType"]
    inp["deactivationType"] = "PERMANENT"
    out["final_op"] = "deactivateFamilies" if family else "deactivateStyle"
    return out


SCOPED_OPS: dict[str, dict[str, dict[str, Any]]] = {
    "activateFamily": {k: {**v, "final_op": "activateFamily"} for k, v in FONT_SCOPED_FAMILY.items()},
    "deactivateFamilies": {
        k: _to_deactivate(v, family=True) for k, v in FONT_SCOPED_FAMILY.items()
    },
    "activateStyle": {
        k: {**_swap_family_to_style(v), "final_op": "activateStyle"}
        for k, v in FONT_SCOPED_FAMILY.items()
    },
    "deactivateStyle": {
        k: _to_deactivate(_swap_family_to_style(v), family=False)
        for k, v in FONT_SCOPED_FAMILY.items()
    },
}

# activateVariation / deactivateVariation — schema has listIds/listType/projectId;
# resolver largely uses projectId; still document UI-possible scopes.
for op, id_key, id_ph, act_key in (
    ("activateVariation", "variations", {"styleId": PH["style_id"], "md5": PH["md5"]}, "activationType"),
    ("deactivateVariation", "md5s", PH["md5"], "deactivationType"),
):
    SCOPED_OPS[op] = {}
    for touch, base in FONT_SCOPED_FAMILY.items():
        shape = deepcopy(base)
        inp: dict[str, Any] = {}
        if op == "activateVariation":
            inp["variations"] = [id_ph]
            inp["activationType"] = "PERMANENT"
        else:
            inp["md5s"] = [PH["md5"]]
            inp["deactivationType"] = "PERMANENT"
        # Carry scope fields from family shape except familyIds
        for k in ("listIds", "listType", "projectId"):
            if k in shape["input_shape"]:
                inp[k] = shape["input_shape"][k]
        shape["input_shape"] = inp
        shape["final_op"] = op
        shape["diff_keys"] = f"{id_key} + scope fields ({', '.join(k for k in inp if k not in {id_key, act_key}) or 'none'})"
        SCOPED_OPS[op][touch] = shape

SCOPED_OPS["activateList"] = {
    "List (FONTLIST)": {
        "input_shape": {
            "listId": PH["list_id"],
            "listType": "FONTLIST",
            "activationType": "PERMANENT",
        },
        "diff_keys": "listId + listType=FONTLIST",
        "prereq": [
            ("createAsset", {"assetType": "FontList", "name": "{{LIST_NAME}}", "accessRight": "FullAccess"}),
            ("addFontListFamilies", {"fontListId": PH["list_id"], "families": {"familyIds": [PH["family_id"]]}}),
        ],
        "final_op": "activateList",
    },
    "Favourite": {
        "input_shape": {
            "listType": "FAVORITE",
            "activationType": "PERMANENT",
        },
        "diff_keys": "listType=FAVORITE (listId forbidden)",
        "prereq": [("addFavoriteFamilies", {"familyIds": [PH["family_id"]]})],
        "final_op": "activateList",
    },
    "Project > List": {
        "input_shape": {
            "listId": PH["list_id"],
            "listType": "FONTLIST",
            "activationType": "PERMANENT",
            "projectId": PH["project_id"],
        },
        "diff_keys": "listId + projectId",
        "prereq": [
            ("createProject", {
                    "name": "{{PROJECT_NAME}}",
                    "description": "QA TouchPoint sheet seed project",
                    "allowFontAdditionsByCollaborators": False,
                    "allowFontDownloadsByCollaborators": False,
                    "allowFontImportsByCollaborators": False,
                    "enableProjectLevelImportedFonts": False,
                    "autoActivateFontsForMembers": False,
                }),
            ("createAsset", {"assetType": "FontList", "name": "{{LIST_NAME}}", "accessRight": "FullAccess", "parentId": PH["project_id"]}),
            ("addFontListFamilies", {"fontListId": PH["list_id"], "families": {"familyIds": [PH["family_id"]]}}),
        ],
        "final_op": "activateList",
    },
}
SCOPED_OPS["deActivateList"] = {
    k: {
        **deepcopy(v),
        "final_op": "deActivateList",
        "input_shape": {
            **{kk: vv for kk, vv in v["input_shape"].items() if kk != "activationType"},
            "deactivationType": "PERMANENT",
        },
    }
    for k, v in SCOPED_OPS["activateList"].items()
}

SCOPED_OPS["activateFontProject"] = {
    "Project": {
        "input_shape": {"projectId": PH["project_id"], "activationType": "PERMANENT"},
        "diff_keys": "projectId required (whole project)",
        "prereq": [
            ("createProject", {
                    "name": "{{PROJECT_NAME}}",
                    "description": "QA TouchPoint sheet seed project",
                    "allowFontAdditionsByCollaborators": False,
                    "allowFontDownloadsByCollaborators": False,
                    "allowFontImportsByCollaborators": False,
                    "enableProjectLevelImportedFonts": False,
                    "autoActivateFontsForMembers": False,
                }),
            (
                "addFontProjectFamilies",
                {
                    "projectId": PH["project_id"],
                    "families": {"familyIds": [PH["family_id"]]},
                },
            ),
        ],
        "final_op": "activateFontProject",
    }
}
SCOPED_OPS["deActivateFontProject"] = {
    "Project": {
        "input_shape": {"projectId": PH["project_id"], "deactivationType": "PERMANENT"},
        "diff_keys": "projectId required",
        "prereq": SCOPED_OPS["activateFontProject"]["Project"]["prereq"],
        "final_op": "deActivateFontProject",
    }
}

# Bulk activation — optional projectId
for op, body_key, item_ph in (
    ("bulkActivateStyles", "styles", PH["style_id"]),
    ("bulkDeactivateStyles", "styles", PH["style_id"]),
    ("bulkActivateLists", "lists", PH["list_id"]),
    ("bulkDeactivateLists", "lists", PH["list_id"]),
):
    is_deact = "Deactiv" in op
    type_key = "deactivationType" if is_deact else "activationType"
    base_items = {body_key: [{"id": item_ph}], type_key: "PERMANENT"}
    SCOPED_OPS[op] = {
        "Discovery/Browse (global)": {
            "input_shape": dict(base_items),
            "diff_keys": f"{body_key}[] only",
            "prereq": [],
            "final_op": op,
        },
        "Project": {
            "input_shape": {**base_items, "projectId": PH["project_id"]},
            "diff_keys": "projectId scopes bulk batch",
            "prereq": [
                (
                    "createAsset",
                    {
                        "assetType": "FontProject",
                        "name": "{{PROJECT_NAME}}",
                        "accessRight": "FullAccess",
                    },
                ),
            ],
            "final_op": op,
        },
    }


def _y(v: Any) -> bool:
    return str(v or "").strip().upper() in {"Y", "YES"}


def _opt(v: Any) -> bool:
    return str(v or "").strip().upper() in {"OPT", "OPTIONAL", "O"}


def _active(v: Any) -> bool:
    return _y(v) or _opt(v)


def parse_matrix(path: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Matrix"]
    rows = list(ws.iter_rows(values_only=True))
    ops: list[dict[str, Any]] = []
    cur: dict[str, Any] | None = None
    for r in rows[3:]:
        if r[0]:
            cur = {
                "operation": str(r[0]).strip(),
                "module": r[6],
                "trigger": str(r[7] or ""),
                "notes": str(r[8] or ""),
                "sources": {},
            }
            ops.append(cur)
        if not cur or not r[1]:
            continue
        cur["sources"][str(r[1])] = {
            "global": r[2],
            "project": r[3],
            "document": r[4],
            "list": r[5],
        }
    out = []
    for o in ops:
        trig = o["trigger"].lower()
        if "graphql" not in trig and "mutation" not in trig:
            continue
        web = o["sources"].get("web") or {}
        if not web:
            continue
        name = o["operation"]
        # Skip queries
        if name.startswith("get") or name.startswith("Get"):
            continue
        if get_document_for_operation(name) and not is_mutation_operation(name):
            # document is a query
            if name.startswith("download") or name.startswith("search"):
                continue
        out.append(
            {
                "operation": name,
                "module": o["module"],
                "trigger": o["trigger"],
                "notes": o["notes"],
                **web,
            }
        )
    return out


# Explicit multi-step flows (schema-correct variables via touchpoint_payloads.variables_for)
def touchpoints_for(m: dict[str, Any]) -> list[str]:
    """Derive UI touchpoints from matrix Y/Opt flags + known FontActivation set."""
    op = m["operation"]
    if op in FLOW_DEFS:
        return list(FLOW_DEFS[op].keys())
    if op in SCOPED_OPS:
        return list(SCOPED_OPS[op].keys())

    combos: list[str] = []
    g, p, d, l = m["global"], m["project"], m["document"], m["list"]
    notes = (m.get("notes") or "").lower()

    if _active(g):
        combos.append("Discovery/Browse (global)")
    if _active(l):
        combos.append("List (FONTLIST)")
        if "favorite" in notes or "favourite" in notes or op.startswith("addFavorite"):
            combos.append("Favourite")
    if _active(p):
        combos.append("Project")
    if _active(p) and _active(l):
        combos.append("Project > List")
    if _active(d):
        combos.append("Document")
        if _active(p):
            combos.append("Document > Project")
    if not combos:
        combos.append("default")
    return combos


def default_shape(op: str, touch: str, m: dict[str, Any]) -> dict[str, Any]:
    """Fallback input shape when not in SCOPED_OPS."""
    inp: dict[str, Any] = {"_note": "Fill from subject.metadata.input / GraphQL schema"}
    diff = "default / see schema"
    prereq: list[tuple[str, dict]] = []

    if "Document" in touch:
        inp = {"documentId": PH["document_id"]}
        if "Project" in touch or _active(m["project"]):
            inp["projectId"] = PH["project_id"]
        diff = "documentId" + (" + projectId" if "projectId" in inp else "")
    elif touch == "Project":
        inp = {"projectId": PH["project_id"]}
        diff = "projectId"
        prereq = [
            ("createProject", {
                    "name": "{{PROJECT_NAME}}",
                    "description": "QA TouchPoint sheet seed project",
                    "allowFontAdditionsByCollaborators": False,
                    "allowFontDownloadsByCollaborators": False,
                    "allowFontImportsByCollaborators": False,
                    "enableProjectLevelImportedFonts": False,
                    "autoActivateFontsForMembers": False,
                })
        ]
    elif "List" in touch:
        inp = {"listId": PH["list_id"]}
        if "Project" in touch:
            inp["projectId"] = PH["project_id"]
        diff = "listId" + (" + projectId" if "projectId" in inp else "")
        prereq = [
            ("createAsset", {"assetType": "FontList", "name": "{{LIST_NAME}}", "accessRight": "FullAccess"})
        ]
    elif touch == "Favourite":
        inp = {"listType": "FAVORITE"}
        diff = "listType=FAVORITE"
    else:
        # global — leave note; still useful for sheet completeness
        inp = {"_touchpoint": "global", "_hint": "Use captured raw subject.metadata.input"}
        diff = "global (no list/project/document scope)"

    return {
        "input_shape": inp,
        "diff_keys": diff,
        "prereq": prereq,
        "final_op": op,
    }


def _operation_name(operation: str, query: str) -> str:
    """Apollo operationName — prefer PascalCase from the document."""
    import re

    m = re.search(r"(?:mutation|query)\s+(\w+)", query or "", re.I)
    if m:
        return m.group(1)
    return operation[:1].upper() + operation[1:] if operation else "Operation"


def _normalize_variables(operation: str, variables: dict[str, Any]) -> dict[str, Any]:
    """Accept either a full variables dict or a bare input body; never invent fields."""
    vars_ = _fill_placeholders(variables or {})
    vars_ = {k: v for k, v in vars_.items() if not str(k).startswith("_")}
    # Already a full variables object from touchpoint_payloads.variables_for
    if "input" in vars_ or "styleFilterInput" in vars_ or "sessionId" in vars_:
        return vars_
    if operation == "processUploadSessionFonts":
        return {
            "sessionId": PH["session_id"],
            "fileId": PH["file_id"],
            "projectId": PH["project_id"],
        }
    return {"input": vars_} if vars_ else {}


def build_curl(operation: str, variables: dict[str, Any]) -> str:
    """Postman-ready curl: real bearer, Apollo body, schema-correct variables."""
    doc = get_document_for_operation(operation)
    if not doc:
        query = (
            f"mutation {_operation_name(operation, '')}($input: JSON) {{\n"
            f"  {operation}(input: $input) {{ __typename }}\n}}"
        )
    else:
        query = doc.strip()

    op_name = _operation_name(operation, query)
    gql_vars = _normalize_variables(operation, variables)

    # Hard guard: addFontListFamilies must never ship with listId / without families
    if operation == "addFontListFamilies":
        inp = gql_vars.get("input") or {}
        if "listId" in inp and "fontListId" not in inp:
            inp["fontListId"] = inp.pop("listId")
        if "families" not in inp:
            inp["families"] = {"familyIds": [PH["family_id"]]}
        if "fontListId" not in inp:
            inp["fontListId"] = PH["list_id"]
        gql_vars["input"] = inp
        gql_vars.setdefault(
            "styleFilterInput", {"pagination": {"skip": 0, "limit": 10}}
        )

    if operation == "addFontProjectFamilies":
        inp = gql_vars.get("input") or {}
        if "projectId" in inp and "fontProjectId" not in inp:
            inp["fontProjectId"] = inp.pop("projectId")
        if "families" not in inp:
            inp["families"] = {"familyIds": [PH["family_id"]]}
        gql_vars["input"] = inp

    payload = {
        "operationName": op_name,
        "variables": gql_vars,
        "extensions": {"clientLibrary": {"name": "@apollo/client", "version": "4.0.9"}},
        "query": query,
    }
    body = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    body_escaped = body.replace("'", "'\\''")
    return (
        f"curl '{PH['endpoint']}' \\\n"
        f"  -H 'accept: application/graphql-response+json,application/json;q=0.9' \\\n"
        f"  -H 'accept-language: en' \\\n"
        f"  -H 'authorization: Bearer {PH['token']}' \\\n"
        f"  -H 'content-type: application/json' \\\n"
        f"  -H 'origin: https://nextgen.monotype-pp.com' \\\n"
        f"  -H 'referer: https://nextgen.monotype-pp.com/search' \\\n"
        f"  -X POST \\\n"
        f"  --data-raw '{body_escaped}'"
    )


def build_titled_curl(operation: str, variables: dict[str, Any]) -> str:
    """Title + curl. ``variables`` should be from ``variables_for`` when possible."""
    gql_vars = _normalize_variables(operation, variables)
    title_vars = gql_vars.get("input") if isinstance(gql_vars.get("input"), dict) else gql_vars
    title = step_title(operation, title_vars if isinstance(title_vars, dict) else {})
    return format_step_cell(title, build_curl(operation, variables))


def ensure_missing_graphql_docs() -> list[str]:
    """Add GQL docs for matrix mutations missing from graphql_documents.json."""
    docs = json.loads(GRAPHQL_DOCS.read_text(encoding="utf-8"))
    added: list[str] = []
    extras = {
        "BULK_MARK_AS_PRODUCTION_FONTS_REQUEST": (
            "mutation BulkMarkAsProductionFontsRequest($input: BulkMarkAsProductionFontsRequestInput!) { "
            "bulkMarkAsProductionFontsRequest(input: $input) { batchId actionType status progressPercent "
            "actionCounts { totalFonts fontsProcessed fontsFailed fontsSkipped } createdAt updatedAt } }"
        ),
        "PROCESS_UPLOAD_SESSION_FONTS": (
            "mutation ProcessUploadSessionFonts($sessionId: UUID!, $fileId: UUID, $projectId: UUID) { "
            "processUploadSessionFonts(sessionId: $sessionId, fileId: $fileId, projectId: $projectId) { "
            "id status } }"
        ),
        "DENY_INTENT_FOR_PRODUCTION": (
            "mutation DenyIntentForProduction($input: DenyIntentForProductionInput!) { "
            "denyIntentForProduction(input: $input) { success errors { code message } } }"
        ),
        "BULK_UNMARK_PRODUCTION_FONTS_REQUEST": (
            "mutation BulkUnmarkProductionFontsRequest($input: BulkUnmarkRequestInput!) { "
            "bulkUnmarkProductionFontsRequest(input: $input) { batchId actionType status } }"
        ),
        "KEEP_IN_PRODUCTION": (
            "mutation KeepInProduction($input: KeepInProductionInput!) { "
            "keepInProduction(input: $input) }"
        ),
    }
    for key, doc in extras.items():
        if key not in docs:
            docs[key] = doc
            added.append(key)
    if added:
        GRAPHQL_DOCS.write_text(json.dumps(docs, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        # Clear lru cache so sheet build sees new docs
        from audit_validator.utility import operation_graphql as og

        og.load_operation_index.cache_clear()
    return added


def _style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _autosize(ws, max_width: int = 60) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min(max_width, max((len(str(c.value or "")) for c in col), default=10) + 2)
        ws.column_dimensions[letter].width = max(12, width)


def build() -> None:
    added_docs = ensure_missing_graphql_docs()
    live = resolve_live_context(provision_assets=True)
    print("Live seed:", {k: getattr(live, k) for k in ("family_id", "style_id", "list_id", "project_id", "list_name", "project_name")})
    for note in live.notes:
        print(" ", note)

    matrix = parse_matrix(MATRIX_XLSX)
    catalog = operation_source_report()["catalog"]
    catalog_ops = {c["operation"] for c in catalog}
    catalog_gql = {c["operation"] for c in catalog if c["kind"] == "graphql"}

    touch_rows: list[dict[str, Any]] = []
    shape_rows: list[dict[str, Any]] = []
    flow_rows: list[dict[str, Any]] = []
    seq_json: dict[str, Any] = {}

    seen_shapes: set[tuple[str, str]] = set()

    for m in sorted(matrix, key=lambda x: x["operation"]):
        op = m["operation"]
        touches = touchpoints_for(m)
        seed = SeedIds(
            family_id=live.family_id,
            style_id=live.style_id,
            md5=PH["md5"],
            list_id=live.list_id,
            project_id=live.project_id,
            list_name=live.list_name,
            project_name=live.project_name,
        )

        for touch in touches:
            if op in FLOW_DEFS and touch not in FLOW_DEFS[op]:
                continue
            if op in FLOW_DEFS:
                step_ops = list(FLOW_DEFS[op][touch])
                final_vars = variables_for(op, seed, touch=touch)
                input_shape = (final_vars.get("input") or final_vars)
                diff_keys = json.dumps(sorted((input_shape or {}).keys()))
            elif op in SCOPED_OPS and touch in SCOPED_OPS[op]:
                # Legacy fallback — still re-resolve final vars via schema builders when known
                step_ops = [p[0] for p in (SCOPED_OPS[op][touch].get("prereq") or [])] + [
                    SCOPED_OPS[op][touch]["final_op"]
                ]
                final_vars = variables_for(op, seed, touch=touch)
                if (final_vars.get("input") or {}).get("_unsupported"):
                    final_vars = {"input": _fill_placeholders(SCOPED_OPS[op][touch]["input_shape"])}
                input_shape = final_vars.get("input") or {}
                diff_keys = SCOPED_OPS[op][touch].get("diff_keys", "")
            else:
                spec = default_shape(op, touch, m)
                step_ops = [p[0] for p in (spec.get("prereq") or [])] + [spec["final_op"]]
                final_vars = variables_for(op, seed, touch=touch)
                if (final_vars.get("input") or {}).get("_unsupported"):
                    final_vars = {"input": _fill_placeholders(spec["input_shape"])}
                input_shape = final_vars.get("input") or {}
                diff_keys = spec.get("diff_keys", "")

            shape_key = (op, json.dumps(input_shape, sort_keys=True, default=str))
            if shape_key not in seen_shapes:
                seen_shapes.add(shape_key)
                shape_rows.append(
                    {
                        "Event": op,
                        "TouchPoint": touch,
                        "unique_input_json": json.dumps(input_shape, indent=2, default=str),
                        "diff_keys_vs_global": diff_keys,
                        "matrix_global": m["global"],
                        "matrix_project": m["project"],
                        "matrix_document": m["document"],
                        "matrix_list": m["list"],
                        "module": m.get("module") or "",
                        "notes": (m.get("notes") or "")[:300],
                        "in_catalog": "Y" if op in catalog_ops else "N",
                        "has_gql_doc": "Y" if get_document_for_operation(op) else "N",
                    }
                )

            # Build titled curls — variables always from schema builders when possible
            steps: list[tuple[str, str]] = []
            for step_op in step_ops:
                step_vars = variables_for(step_op, seed, touch=touch)
                if (step_vars.get("input") or {}).get("_unsupported"):
                    # last resort for unmapped ops
                    step_vars = {"input": {"_note": f"fill from schema for {step_op}"}}
                cell = build_titled_curl(step_op, step_vars)
                if step_op == "createAsset" and live.list_id:
                    cell = (
                        f"Create List :\n\n"
                        f"# Sheet-build list id (used in later steps): {live.list_id}\n"
                        f"# Prefer Postman folder scripts to create+capture a fresh id.\n\n"
                        f"{build_curl(step_op, step_vars)}"
                    )
                elif step_op == "createProject" and live.project_id:
                    cell = (
                        f"Create Project :\n\n"
                        f"# Sheet-build project id (used in later steps): {live.project_id}\n"
                        f"# Prefer Postman folder scripts to create+capture a fresh id.\n\n"
                        f"{build_curl(step_op, step_vars)}"
                    )
                steps.append((step_op, cell))

            row: dict[str, Any] = {
                "Event": op,
                "TouchPoint": touch,
                "Platform": "Web",
                "matrix_global": m["global"],
                "matrix_project": m["project"],
                "matrix_document": m["document"],
                "matrix_list": m["list"],
                "unique_input_json": json.dumps(input_shape, indent=2, default=str),
                "diff_keys_vs_global": diff_keys,
                "enrich_note": (
                    "Schema-correct GraphQL variables (mtf-graphql-schema). "
                    f"family={live.family_id} style={live.style_id} "
                    f"list={live.list_id or 'n/a'} project={live.project_id or 'n/a'}. "
                    "addFontListFamilies uses fontListId + families.familyIds (never listId alone)."
                ),
                "module": m.get("module") or "",
                "in_catalog": "Y" if op in catalog_ops else "N",
            }
            for i in range(5):
                if i < len(steps):
                    row[f"step {i + 1} op"] = steps[i][0]
                    row[f"step {i + 1}"] = steps[i][1]
                else:
                    row[f"step {i + 1} op"] = ""
                    row[f"step {i + 1}"] = ""
            touch_rows.append(row)

            flow_rows.append(
                {
                    "Event": op,
                    "TouchPoint": touch,
                    "flow": " → ".join(s[0] for s in steps),
                    "step_ops": " | ".join(s[0] for s in steps),
                    "unique_input_json": json.dumps(input_shape, indent=2, default=str),
                }
            )

            seq_json.setdefault(op, {})
            seq_json[op][touch] = [s[0] for s in steps]

        # Always keep a default sequence
        if op in seq_json and "default" not in seq_json[op]:
            first = next(iter(seq_json[op].values()), [op])
            seq_json[op]["default"] = first

    # Gaps
    matrix_ops = {m["operation"] for m in matrix}
    gaps = []
    for op in sorted(matrix_ops - catalog_gql):
        kind = next((c["kind"] for c in catalog if c["operation"] == op), "")
        gaps.append(
            {
                "Event": op,
                "in_matrix": "Y",
                "in_gql_catalog": "N",
                "elsewhere": kind or ("cron/ingress/registered" if op in catalog_ops else "MISSING"),
                "has_gql_doc": "Y" if get_document_for_operation(op) else "N",
                "action": (
                    "Add GraphQL document + simulation flow"
                    if not get_document_for_operation(op)
                    else "Enable in catalog (enricher / should_simulate)"
                ),
            }
        )

    # Write Excel
    OUT_XLSX.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()

    # Legend
    ws = wb.active
    ws.title = "Legend"
    legend = [
        ["GQL TouchPoint × Unique Input sheet"],
        [],
        ["How to read"],
        ["UniqueInputShapes", "One row per Event × distinct GraphQL input shape (what differs in raw metadata.input)"],
        ["TouchPoints", "Full Web curls (env placeholders) + multi-step prereq flows"],
        ["AutomationFlows", "Compact flow summary for dynamic generate: create → seed → trigger"],
        ["Gaps", "Matrix GQL mutations not yet in our GraphQL generate catalog"],
        [],
        ["Context places (from matrix)"],
        ["global / Discovery", "No listIds / projectId / documentId — browse/search activate"],
        ["List (FONTLIST)", "listIds + listType=FONTLIST"],
        ["Favourite", "listType=FAVORITE (usually no listIds)"],
        ["Project", "listIds=project_{uuid} + listType=FONTPROJECT + projectId  OR  activateFontProject"],
        ["Project > List", "listIds (FONTLIST under project) + projectId"],
        ["Document", "documentId (+ projectId when linked) — not used on activateFamily"],
        [],
        ["Placeholders"],
        ["{{GRAPHQL_ENDPOINT}}", "e.g. https://nextgen.monotype-pp.com/graph"],
        ["{{BEARER_TOKEN}}", "PP access token"],
        ["{{FAMILY_ID}} / {{STYLE_ID}} / {{LIST_ID}} / {{PROJECT_ID}}", "Seed from env / discovery"],
        [],
        ["Sources"],
        ["Matrix", str(MATRIX_XLSX)],
        ["TouchPoint examples", str(TOUCHPOINT_SRC)],
        ["Schema", "mtf-graphql-schema ActivateFamilyInput (+ list/project scope)"],
        ["API routing", "mtconnect-api fonts.resolver + listScopedActivation.helper"],
        [],
        ["Docs added this run", ", ".join(added_docs) or "(none)"],
        ["Row counts", f"shapes={len(shape_rows)} touch_rows={len(touch_rows)} gaps={len(gaps)}"],
    ]
    for r in legend:
        ws.append(r)
    ws["A1"].font = Font(bold=True, size=14)

    def write_sheet(name: str, rows: list[dict[str, Any]], cols: list[str]) -> None:
        w = wb.create_sheet(name)
        w.append(cols)
        _style_header(w)
        for row in rows:
            w.append([row.get(c, "") for c in cols])
        for r in w.iter_rows(min_row=2):
            for c in r:
                c.alignment = Alignment(wrap_text=True, vertical="top")
        _autosize(w)
        w.freeze_panes = "A2"

    write_sheet(
        "UniqueInputShapes",
        shape_rows,
        [
            "Event",
            "TouchPoint",
            "unique_input_json",
            "diff_keys_vs_global",
            "matrix_global",
            "matrix_project",
            "matrix_document",
            "matrix_list",
            "module",
            "notes",
            "in_catalog",
            "has_gql_doc",
        ],
    )
    write_sheet(
        "TouchPoints",
        touch_rows,
        [
            "Event",
            "TouchPoint",
            "Platform",
            "matrix_global",
            "matrix_project",
            "matrix_document",
            "matrix_list",
            "unique_input_json",
            "diff_keys_vs_global",
            "enrich_note",
            "step 1 op",
            "step 1",
            "step 2 op",
            "step 2",
            "step 3 op",
            "step 3",
            "step 4 op",
            "step 4",
            "step 5 op",
            "step 5",
            "module",
            "in_catalog",
        ],
    )
    write_sheet(
        "AutomationFlows",
        flow_rows,
        ["Event", "TouchPoint", "flow", "step_ops", "unique_input_json"],
    )
    write_sheet(
        "Gaps",
        gaps,
        ["Event", "in_matrix", "in_gql_catalog", "elsewhere", "has_gql_doc", "action"],
    )

    # Compatibility sheet matching original TouchPoint.xlsx column layout
    ws_tp = wb.create_sheet("Touch Points", 1)
    ws_tp.append(["Event", "TouchPoint", "Web / App", "step 1", "step 2", "step 3", "step 4", "step 5"])
    _style_header(ws_tp)
    last_event = None
    for row in touch_rows:
        ev = row["Event"] if row["Event"] != last_event else None
        last_event = row["Event"]
        ws_tp.append(
            [
                ev or "",
                row["TouchPoint"],
                row["Platform"],
                row.get("step 1") or "",
                row.get("step 2") or "",
                row.get("step 3") or "",
                row.get("step 4") or "",
                row.get("step 5") or "",
            ]
        )
    for r in ws_tp.iter_rows(min_row=2):
        for c in r:
            c.alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(ws_tp, max_width=80)
    ws_tp.freeze_panes = "A2"

    wb.save(OUT_XLSX)
    # Also refresh user's Downloads copy
    try:
        wb.save(TOUCHPOINT_SRC)
    except Exception as exc:  # noqa: BLE001
        print(f"Note: could not overwrite Downloads TouchPoint.xlsx ({exc})")

    # Postman collection (create → capture id → mutate)
    try:
        from build_touchpoint_postman import build as build_postman

        build_postman()
    except Exception as exc:  # noqa: BLE001
        print(f"Note: Postman collection build failed ({exc})")

    # CSV (Event / TouchPoint / steps) — shorter for git
    with OUT_CSV.open("w", encoding="utf-8", newline="") as fh:
        cols = [
            "Event",
            "TouchPoint",
            "unique_input_json",
            "diff_keys_vs_global",
            "step 1 op",
            "step 2 op",
            "step 3 op",
            "step 4 op",
            "step 5 op",
            "flow",
        ]
        w = csv.DictWriter(fh, fieldnames=cols, extrasaction="ignore")
        w.writeheader()
        for row, flow in zip(touch_rows, flow_rows):
            w.writerow({**row, "flow": flow["flow"]})

    SEQ_OUT.write_text(json.dumps(seq_json, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")

    print(f"Wrote {OUT_XLSX}")
    print(f"  UniqueInputShapes={len(shape_rows)} TouchPoints={len(touch_rows)} Gaps={len(gaps)}")
    print(f"  GraphQL docs added: {added_docs or 'none'}")
    print(f"  CSV → {OUT_CSV}")
    print(f"  sequences → {SEQ_OUT}")


if __name__ == "__main__":
    build()
